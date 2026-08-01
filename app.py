from __future__ import annotations

import base64
import hashlib
import hmac
import html
import os
import secrets
import sqlite3
from io import BytesIO
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlencode, urlparse


APP_NAME = "Mémoire coutumière"
DB_PATH = Path("genealogie_coutumiere.sqlite3")
HOST = "127.0.0.1"
PORT = 8765
SESSION_COOKIE = "mc_session"
SESSION_SECRET_FILE = Path(".session_secret")


def get_secret() -> bytes:
    if not SESSION_SECRET_FILE.exists():
        SESSION_SECRET_FILE.write_bytes(secrets.token_bytes(32))
    return SESSION_SECRET_FILE.read_bytes()


SESSION_SECRET = get_secret()


def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return (
        base64.urlsafe_b64encode(salt).decode("ascii")
        + "$"
        + base64.urlsafe_b64encode(digest).decode("ascii")
    )


def verify_password(password: str, stored: str) -> bool:
    try:
        salt_text, digest_text = stored.split("$", 1)
        salt = base64.urlsafe_b64decode(salt_text.encode("ascii"))
        expected = base64.urlsafe_b64decode(digest_text.encode("ascii"))
    except Exception:
        return False
    actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return hmac.compare_digest(actual, expected)


def make_session(username: str) -> str:
    nonce = secrets.token_urlsafe(18)
    payload = f"{username}|{nonce}"
    sig = hmac.new(SESSION_SECRET, payload.encode("utf-8"), hashlib.sha256).hexdigest()
    token = base64.urlsafe_b64encode(f"{payload}|{sig}".encode("utf-8")).decode("ascii")
    return token


def read_session(token: str | None) -> str | None:
    if not token:
        return None
    try:
        decoded = base64.urlsafe_b64decode(token.encode("ascii")).decode("utf-8")
        username, nonce, sig = decoded.split("|", 2)
    except Exception:
        return None
    payload = f"{username}|{nonce}"
    expected = hmac.new(SESSION_SECRET, payload.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(sig, expected):
        return None
    return username


def init_db() -> None:
    with db_connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS clans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                variants TEXT,
                region TEXT,
                description TEXT,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS persons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                first_names TEXT NOT NULL,
                current_name TEXT NOT NULL,
                birth_name TEXT,
                other_names TEXT,
                gender TEXT,
                birth_date TEXT,
                birth_place TEXT,
                death_date TEXT,
                current_clan_id INTEGER,
                origin_clan_id INTEGER,
                notes TEXT,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (current_clan_id) REFERENCES clans(id),
                FOREIGN KEY (origin_clan_id) REFERENCES clans(id)
            );

            CREATE TABLE IF NOT EXISTS relationships (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subject_person_id INTEGER NOT NULL,
                object_person_id INTEGER,
                object_clan_id INTEGER,
                relation_type TEXT NOT NULL,
                comment TEXT,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (subject_person_id) REFERENCES persons(id),
                FOREIGN KEY (object_person_id) REFERENCES persons(id),
                FOREIGN KEY (object_clan_id) REFERENCES clans(id)
            );

            CREATE TABLE IF NOT EXISTS inferred_relationship_exclusions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_a_id INTEGER NOT NULL,
                person_b_id INTEGER NOT NULL,
                relation_type TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(person_a_id, person_b_id, relation_type),
                FOREIGN KEY (person_a_id) REFERENCES persons(id) ON DELETE CASCADE,
                FOREIGN KEY (person_b_id) REFERENCES persons(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS person_group_links (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_id INTEGER NOT NULL,
                group_id INTEGER NOT NULL,
                link_type TEXT NOT NULL DEFAULT 'rattache_a',
                comment TEXT,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (person_id) REFERENCES persons(id) ON DELETE CASCADE,
                FOREIGN KEY (group_id) REFERENCES customary_groups(id)
            );

            CREATE TABLE IF NOT EXISTS person_customary_name_links (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_id INTEGER NOT NULL,
                customary_name_id INTEGER NOT NULL,
                link_type TEXT NOT NULL DEFAULT 'rattache_a',
                comment TEXT,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (person_id) REFERENCES persons(id) ON DELETE CASCADE,
                FOREIGN KEY (customary_name_id) REFERENCES customary_names(id)
            );

            CREATE TABLE IF NOT EXISTS customary_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT NOT NULL,
                title TEXT NOT NULL,
                event_date TEXT,
                place TEXT,
                main_person_id INTEGER,
                description TEXT NOT NULL,
                effects TEXT,
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (main_person_id) REFERENCES persons(id),
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS event_clans (
                event_id INTEGER NOT NULL,
                clan_id INTEGER NOT NULL,
                PRIMARY KEY (event_id, clan_id),
                FOREIGN KEY (event_id) REFERENCES customary_events(id) ON DELETE CASCADE,
                FOREIGN KEY (clan_id) REFERENCES clans(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS event_person_links (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER NOT NULL,
                person_id INTEGER NOT NULL,
                role TEXT NOT NULL,
                comment TEXT,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (event_id) REFERENCES customary_events(id) ON DELETE CASCADE,
                FOREIGN KEY (person_id) REFERENCES persons(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS sources (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_type TEXT NOT NULL,
                title TEXT NOT NULL,
                witness_name TEXT,
                collection_date TEXT,
                collected_by TEXT,
                summary TEXT NOT NULL,
                consent TEXT NOT NULL DEFAULT 'a_confirmer',
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                reliability TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS lands (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                place_type TEXT,
                location_text TEXT,
                clan_id INTEGER,
                known_rights TEXT,
                status TEXT NOT NULL DEFAULT 'a_verifier',
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'sensible',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (clan_id) REFERENCES clans(id),
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS customary_groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                group_type TEXT NOT NULL DEFAULT 'a_verifier',
                variants TEXT,
                region TEXT,
                description TEXT,
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS customary_names (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                name_type TEXT NOT NULL DEFAULT 'a_verifier',
                variants TEXT,
                region TEXT,
                description TEXT,
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS customary_name_relations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name_a_id INTEGER NOT NULL,
                name_b_id INTEGER NOT NULL,
                relation_type TEXT NOT NULL DEFAULT 'allie_a',
                context TEXT,
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (name_a_id) REFERENCES customary_names(id),
                FOREIGN KEY (name_b_id) REFERENCES customary_names(id),
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS group_relations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_a_id INTEGER NOT NULL,
                group_b_id INTEGER NOT NULL,
                relation_type TEXT NOT NULL,
                context TEXT,
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (group_a_id) REFERENCES customary_groups(id),
                FOREIGN KEY (group_b_id) REFERENCES customary_groups(id),
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS viva_lists (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                area TEXT,
                collector_author TEXT,
                collection_date TEXT,
                description TEXT,
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS viva_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                viva_list_id INTEGER NOT NULL,
                position INTEGER,
                section_name TEXT,
                raw_text TEXT NOT NULL,
                group_a_id INTEGER,
                group_b_id INTEGER,
                translation TEXT,
                note TEXT,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (viva_list_id) REFERENCES viva_lists(id) ON DELETE CASCADE,
                FOREIGN KEY (group_a_id) REFERENCES customary_groups(id),
                FOREIGN KEY (group_b_id) REFERENCES customary_groups(id)
            );

            CREATE TABLE IF NOT EXISTS change_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_id INTEGER NOT NULL,
                change_type TEXT NOT NULL,
                old_value TEXT,
                new_value TEXT,
                effective_date TEXT,
                event_id INTEGER,
                source_id INTEGER,
                comment TEXT,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (person_id) REFERENCES persons(id),
                FOREIGN KEY (event_id) REFERENCES customary_events(id),
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS research_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subject_type TEXT NOT NULL,
                subject_id INTEGER,
                title TEXT NOT NULL,
                statement TEXT NOT NULL,
                interpretation TEXT,
                author_view TEXT,
                source_id INTEGER,
                evidence_level TEXT NOT NULL DEFAULT 'document_ecrit',
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS lineage_genealogies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id INTEGER,
                title TEXT NOT NULL,
                author_or_collector TEXT,
                chain_text TEXT NOT NULL,
                interpretation TEXT,
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (group_id) REFERENCES customary_groups(id),
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );

            CREATE TABLE IF NOT EXISTS customary_functions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id INTEGER,
                person_id INTEGER,
                function_type TEXT NOT NULL,
                title TEXT NOT NULL,
                place TEXT,
                description TEXT NOT NULL,
                source_id INTEGER,
                confidentiality TEXT NOT NULL DEFAULT 'familial',
                certainty TEXT NOT NULL DEFAULT 'a_verifier',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (group_id) REFERENCES customary_groups(id),
                FOREIGN KEY (person_id) REFERENCES persons(id),
                FOREIGN KEY (source_id) REFERENCES sources(id)
            );
            """
        )
        seed_documentary_data(conn)


def seed_demo_data(conn: sqlite3.Connection) -> None:
    count = conn.execute("SELECT COUNT(*) FROM clans").fetchone()[0]
    if count:
        return

    conn.execute(
        "INSERT INTO clans (name, variants, region, description, confidentiality, certainty) VALUES (?, ?, ?, ?, ?, ?)",
        (
            "Clan Mwââru",
            "Mwaru; Mwâru",
            "Exemple - aire Sud",
            "Clan fictif utilisé pour tester les liens claniques et les sources.",
            "familial",
            "a_verifier",
        ),
    )
    conn.execute(
        "INSERT INTO clans (name, variants, region, description, confidentiality, certainty) VALUES (?, ?, ?, ?, ?, ?)",
        (
            "Clan Nékö",
            "Neko",
            "Exemple - aire Centre",
            "Clan fictif allié dans les données de démonstration.",
            "familial",
            "probable",
        ),
    )

    source_id = conn.execute(
        """
        INSERT INTO sources
        (source_type, title, witness_name, collection_date, collected_by, summary, consent, confidentiality, reliability)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "temoignage_oral",
            "Entretien fictif avec un ancien",
            "Ancien fictif",
            "2026-07-15",
            "Utilisateur test",
            "Source fictive créée pour vérifier la saisie et les liens entre personnes, clans et événements.",
            "a_confirmer",
            "familial",
            "a_verifier",
        ),
    ).lastrowid

    person_a = conn.execute(
        """
        INSERT INTO persons
        (first_names, current_name, birth_name, other_names, gender, birth_date, birth_place,
         current_clan_id, origin_clan_id, notes, confidentiality, certainty)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "Téâ Jean",
            "Wamytan",
            "Wamytan",
            "Jean-Téâ",
            "homme",
            "vers 1975",
            "Village fictif",
            1,
            1,
            "Personne fictive pour tester une fiche complète.",
            "familial",
            "probable",
        ),
    ).lastrowid

    person_b = conn.execute(
        """
        INSERT INTO persons
        (first_names, current_name, birth_name, gender, birth_date, birth_place,
         current_clan_id, origin_clan_id, notes, confidentiality, certainty)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "Malia",
            "Nékö",
            "Nékö",
            "femme",
            "vers 1978",
            "Lieu fictif",
            2,
            2,
            "Personne fictive liée par mariage dans la base de démonstration.",
            "familial",
            "probable",
        ),
    ).lastrowid

    child = conn.execute(
        """
        INSERT INTO persons
        (first_names, current_name, birth_name, gender, birth_date, birth_place,
         current_clan_id, origin_clan_id, notes, confidentiality, certainty)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "Noé",
            "Wamytan",
            "Wamytan",
            "homme",
            "vers 2002",
            "Lieu fictif",
            1,
            1,
            "Enfant fictif permettant de tester les relations parents/enfant.",
            "familial",
            "a_verifier",
        ),
    ).lastrowid

    conn.executemany(
        """
        INSERT INTO relationships
        (subject_person_id, object_person_id, relation_type, comment, confidentiality, certainty)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            (person_a, person_b, "conjoint", "Union fictive pour tester les liens.", "familial", "probable"),
            (person_a, child, "pere", "Lien fictif père-enfant.", "familial", "probable"),
            (person_b, child, "mere", "Lien fictif mère-enfant.", "familial", "probable"),
        ],
    )

    event_id = conn.execute(
        """
        INSERT INTO customary_events
        (event_type, title, event_date, place, main_person_id, description, effects, source_id, confidentiality, certainty)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "souffle_oncles_maternels",
            "Souffle des oncles maternels - exemple fictif",
            "vers 2002",
            "Lieu fictif",
            child,
            "Événement fictif décrivant le principe de consignation du souffle des oncles maternels.",
            "Lien maternel documenté dans la démonstration.",
            source_id,
            "familial",
            "a_verifier",
        ),
    ).lastrowid
    conn.execute("INSERT INTO event_clans (event_id, clan_id) VALUES (?, ?)", (event_id, 2))

    conn.execute(
        """
        INSERT INTO lands
        (name, place_type, location_text, clan_id, known_rights, status, source_id, confidentiality)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "Terre fictive de démonstration",
            "terre",
            "Localisation volontairement fictive.",
            1,
            "Droits fictifs, uniquement pour tester le marquage sensible.",
            "a_verifier",
            source_id,
            "sensible",
        ),
    )


def seed_documentary_data(conn: sqlite3.Connection) -> None:
    # Public repository copy: no real documentary, cultural, or genealogical
    # material is seeded automatically. The private local V1 keeps its data.
    return


def ensure_research_item(
    conn: sqlite3.Connection,
    subject_type: str,
    subject_id: int,
    title: str,
    statement: str,
    interpretation: str,
    author_view: str,
    source_id: int,
    evidence_level: str,
    certainty: str,
) -> int:
    row = conn.execute(
        "SELECT id FROM research_items WHERE subject_type = ? AND subject_id = ? AND title = ?",
        (subject_type, subject_id, title),
    ).fetchone()
    if row:
        return int(row["id"])
    return int(
        conn.execute(
            """
            INSERT INTO research_items
            (subject_type, subject_id, title, statement, interpretation, author_view, source_id, evidence_level, confidentiality, certainty)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (subject_type, subject_id, title, statement, interpretation, author_view, source_id, evidence_level, "familial", certainty),
        ).lastrowid
    )


def ensure_lineage_genealogy(
    conn: sqlite3.Connection,
    group_id: int,
    title: str,
    author_or_collector: str,
    chain_text: str,
    interpretation: str,
    source_id: int,
    certainty: str,
) -> int:
    row = conn.execute(
        "SELECT id FROM lineage_genealogies WHERE group_id = ? AND title = ?",
        (group_id, title),
    ).fetchone()
    if row:
        return int(row["id"])
    return int(
        conn.execute(
            """
            INSERT INTO lineage_genealogies
            (group_id, title, author_or_collector, chain_text, interpretation, source_id, confidentiality, certainty)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (group_id, title, author_or_collector, chain_text, interpretation, source_id, "familial", certainty),
        ).lastrowid
    )


def ensure_customary_function(
    conn: sqlite3.Connection,
    group_id: int,
    function_type: str,
    title: str,
    place: str,
    description: str,
    source_id: int,
    certainty: str,
) -> int:
    row = conn.execute(
        "SELECT id FROM customary_functions WHERE group_id = ? AND title = ?",
        (group_id, title),
    ).fetchone()
    if row:
        return int(row["id"])
    return int(
        conn.execute(
            """
            INSERT INTO customary_functions
            (group_id, function_type, title, place, description, source_id, confidentiality, certainty)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (group_id, function_type, title, place, description, source_id, "familial", certainty),
        ).lastrowid
    )


def ensure_group(
    conn: sqlite3.Connection,
    name: str,
    group_type: str,
    variants: str,
    region: str,
    description: str,
    source_id: int,
    certainty: str,
) -> int:
    row = conn.execute("SELECT id FROM customary_groups WHERE name = ?", (name,)).fetchone()
    if row:
        return int(row["id"])
    return int(
        conn.execute(
            """
            INSERT INTO customary_groups
            (name, group_type, variants, region, description, source_id, confidentiality, certainty)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (name, group_type, variants, region, description, source_id, "familial", certainty),
        ).lastrowid
    )


def users_exist() -> bool:
    with db_connect() as conn:
        return conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] > 0


def create_first_user(password: str) -> None:
    with db_connect() as conn:
        if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] > 0:
            return
        conn.execute(
            "INSERT INTO users (username, password_hash) VALUES (?, ?)",
            ("admin", hash_password(password)),
        )


def authenticate(password: str) -> bool:
    with db_connect() as conn:
        row = conn.execute("SELECT password_hash FROM users WHERE username = ?", ("admin",)).fetchone()
    return bool(row and verify_password(password, row["password_hash"]))


def e(value: Any) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def option_list(rows: list[sqlite3.Row], selected: Any = None, empty_label: str = "Non renseigné") -> str:
    selected_text = "" if selected is None else str(selected)
    parts = [f'<option value="">{e(empty_label)}</option>']
    for row in rows:
        value = str(row["id"])
        mark = " selected" if value == selected_text else ""
        label = row["label"] if "label" in row.keys() else row["name"]
        parts.append(f'<option value="{e(value)}"{mark}>{e(label)}</option>')
    return "\n".join(parts)


def select_options(values: list[tuple[str, str]], selected: str | None = None) -> str:
    return "\n".join(
        f'<option value="{e(value)}"{" selected" if value == selected else ""}>{e(label)}</option>'
        for value, label in values
    )


CONFIDENTIALITY = [
    ("public", "Public / partageable"),
    ("familial", "Familial"),
    ("sensible", "Sensible"),
    ("reserve", "Réservé"),
]

CERTAINTY = [
    ("a_verifier", "À vérifier"),
    ("probable", "Probable"),
    ("confirme", "Confirmé"),
    ("conteste", "Contesté"),
]

SEXES = [
    ("", "Non renseigné"),
    ("M", "Masculin"),
    ("F", "Féminin"),
]

EVENT_TYPES = [
    ("naissance", "Naissance"),
    ("souffle_oncles_maternels", "Souffle des oncles maternels"),
    ("mariage", "Mariage"),
    ("deces", "Décès"),
    ("enfant_redonne_oncles_maternels", "Enfant redonné aux oncles maternels"),
    ("changement_nom", "Changement de nom"),
    ("changement_clan", "Changement de clan"),
    ("transmission_terre", "Transmission ou droit de terre"),
    ("autre", "Autre événement coutumier"),
]

EVENT_PERSON_ROLES = [
    ("epoux", "Époux"),
    ("epouse", "Épouse"),
    ("enfant_redonne", "Enfant redonné"),
    ("oncle_maternel_concerne", "Oncle maternel concerné"),
    ("enfant_naissance", "Enfant concerné"),
    ("defunt", "Défunt"),
    ("participant", "Participant"),
    ("responsable_coutumier", "Responsable coutumier"),
    ("delegation", "Délégation"),
    ("famille_presente", "Famille présente"),
    ("clan_present", "Clan présent"),
    ("autre", "Autre rôle"),
]

PERSON_EVENT_VISIBLE_ROLES = {
    "epoux",
    "epouse",
    "enfant_redonne",
    "oncle_maternel_concerne",
    "enfant_naissance",
    "defunt",
}

RELATION_TYPES = [
    ("pere", "Père"),
    ("mere", "Mère"),
    ("enfant", "Enfant"),
    ("conjoint", "Conjoint"),
    ("frere_soeur", "Frère / sœur"),
    ("oncle_maternel", "Oncle maternel"),
    ("tante_maternelle", "Tante maternelle"),
    ("parent_coutumier", "Parent coutumier"),
    ("referent_memoire", "Référent mémoire"),
    ("appartenance_clan", "Appartenance à un clan"),
    ("clan_origine", "Clan d'origine"),
    ("autre", "Autre relation"),
]

GROUP_TYPES = [
    ("a_verifier", "À vérifier"),
    ("clan", "Clan"),
    ("lignage", "Lignage"),
    ("nom_famille", "Nom de famille"),
    ("branche", "Branche familiale"),
    ("chefferie", "Chefferie"),
    ("groupe_viva", "Groupe cité dans un viva"),
    ("lignage_ou_nom_viva", "Lignage / nom cité dans un viva"),
    ("lieu_associe", "Lieu associé"),
    ("autre", "Autre"),
]

GROUP_RELATION_TYPES = [
    ("apparie_dans_viva", "Apparié dans un viva"),
    ("aine_de", "Aîné de"),
    ("cadet_de", "Cadet de"),
    ("rattache_a", "Rattaché à"),
    ("allie_a", "Allié à"),
    ("accueille", "Accueille"),
    ("accueilli_par", "Accueilli par"),
    ("meme_regroupement", "Même regroupement"),
    ("hierarchie_reference", "Hiérarchie de référence"),
    ("conteste", "Relation contestée"),
    ("autre", "Autre relation"),
]

PERSON_GROUP_LINK_TYPES = [
    ("rattache_a", "Rattaché à"),
    ("appartenance_probable", "Appartenance probable"),
    ("nom_famille_associe", "Nom de famille associé"),
    ("lignage_associe", "Lignage associé"),
    ("origine", "Origine"),
    ("actuel", "Actuel"),
    ("a_verifier", "À vérifier"),
]

SUBJECT_TYPES = [
    ("group", "Groupe coutumier"),
    ("person", "Personne"),
    ("clan", "Clan"),
    ("land", "Terre / lieu"),
    ("event", "Evenement"),
    ("general", "General"),
]

EVIDENCE_LEVELS = [
    ("document_ecrit", "Document ecrit"),
    ("temoignage_oral", "Temoignage oral"),
    ("memoire_familiale", "Memoire familiale"),
    ("validation_coutumiere", "Validation coutumiere"),
    ("hypothese_travail", "Hypothese de travail"),
    ("contradiction", "Version contradictoire"),
]

FUNCTION_TYPES = [
    ("garde_lieu", "Garde d'un lieu"),
    ("maitre_element", "Maitre d'un element"),
    ("pierre_pouvoir", "Pierre / pouvoir"),
    ("guerrier", "Guerrier"),
    ("serviteur", "Serviteur"),
    ("porte_parole", "Porte-parole"),
    ("maitre_ignames", "Maitre des ignames"),
    ("peche", "Peche / reserve de peche"),
    ("rituel", "Fonction rituelle"),
    ("autre", "Autre fonction"),
]

SOURCE_TYPES = [
    ("temoignage_oral", "Témoignage oral"),
    ("reunion", "Réunion familiale ou coutumière"),
    ("document", "Document"),
    ("photo", "Photo"),
    ("audio", "Audio"),
    ("video", "Vidéo"),
    ("archive", "Archive"),
]

CONSENT = [
    ("oui", "Oui"),
    ("non", "Non"),
    ("inconnu", "Inconnu"),
    ("a_confirmer", "À confirmer"),
]


def layout(title: str, body: str, username: str | None = "admin", flash: str = "") -> bytes:
    nav = ""
    if username:
        nav = """
        <nav>
          <div class="nav-group">
            <span class="nav-label">Principal</span>
            <a href="/">Accueil</a>
            <a href="/persons">Personnes</a>
            <a href="/customary-names">Noms coutumiers</a>
            <a href="/events">Événements</a>
          </div>
          <div class="nav-group">
            <span class="nav-label">Outils</span>
            <a href="/bulk">Saisie en masse</a>
            <a href="/quick-relations">Relations rapides</a>
            <a href="/search">Recherche</a>
          </div>
          <div class="nav-group">
            <span class="nav-label">Vues</span>
            <a href="/visualisations">Visualisations</a>
          </div>
          <div class="nav-group secondary-nav">
            <span class="nav-label">Autres</span>
            <a href="/clans">Clans</a>
            <a href="/groups">Groupes</a>
            <a href="/viva">Viva</a>
            <a href="/research">Infos</a>
            <a href="/lineages">Généalogies</a>
            <a href="/functions">Fonctions</a>
            <a href="/sources">Sources</a>
            <a href="/lands">Terres</a>
          </div>
          <div class="nav-group account-nav">
            <a href="/logout">Déconnexion</a>
          </div>
        </nav>
        """
    html_text = f"""<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{e(title)} - {APP_NAME}</title>
  <style>
    :root {{
      --ink: #202124;
      --muted: #5f6368;
      --line: #d7dde5;
      --soft: #f5f7fa;
      --accent: #22577a;
      --accent-2: #2f7d62;
      --warn: #8a5b00;
      --danger: #9a3412;
      --bg: #ffffff;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--ink);
      background: #eef2f6;
      line-height: 1.45;
    }}
    header {{
      background: var(--bg);
      border-bottom: 1px solid var(--line);
      padding: 16px 28px 0;
      position: sticky;
      top: 0;
      z-index: 5;
    }}
    .brand {{
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 14px;
    }}
    h1 {{ font-size: 24px; margin: 0; color: var(--accent); }}
    h2 {{ font-size: 20px; margin: 26px 0 12px; color: var(--accent); }}
    h3 {{ font-size: 16px; margin: 20px 0 10px; }}
    .subtitle {{ color: var(--muted); font-size: 14px; }}
    nav {{ display: flex; flex-wrap: wrap; align-items: stretch; gap: 8px; padding-bottom: 8px; }}
    .nav-group {{
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      gap: 2px;
      padding: 4px 8px 4px 0;
      border-right: 1px solid var(--line);
    }}
    .nav-group:last-child {{ border-right: 0; }}
    .nav-label {{
      color: var(--muted);
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0;
      padding: 10px 6px 10px 0;
    }}
    .secondary-nav a {{ color: #3f4a55; }}
    .account-nav {{ margin-left: auto; border-right: 0; }}
    nav a {{
      color: var(--ink);
      text-decoration: none;
      padding: 9px 10px;
      border-radius: 6px;
      font-size: 14px;
    }}
    nav a:hover {{ background: var(--soft); color: var(--accent); }}
    main {{ max-width: 1560px; margin: 24px auto; padding: 0 24px 60px; }}
    .panel {{
      background: var(--bg);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 20px;
      margin-bottom: 18px;
      box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
    }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 14px; }}
    .metric {{ border: 1px solid var(--line); border-radius: 8px; padding: 16px; background: var(--soft); }}
    .metric strong {{ display: block; font-size: 28px; color: var(--accent); }}
    .actions {{ display: flex; flex-wrap: wrap; gap: 10px; margin: 12px 0; }}
    a.button, button {{
      display: inline-block;
      border: 1px solid var(--accent);
      background: var(--accent);
      color: white;
      text-decoration: none;
      border-radius: 6px;
      padding: 9px 12px;
      font-size: 14px;
      cursor: pointer;
    }}
    a.secondary, button.secondary {{ background: white; color: var(--accent); }}
    table {{ width: 100%; border-collapse: collapse; background: white; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 10px; text-align: left; vertical-align: top; }}
    th {{ background: var(--soft); color: #30363d; font-size: 13px; }}
    tr:hover td {{ background: #fafcff; }}
    .badge {{ display: inline-block; border-radius: 999px; padding: 3px 8px; font-size: 12px; background: var(--soft); border: 1px solid var(--line); }}
    .sensible, .reserve {{ color: var(--danger); border-color: #fed7aa; background: #fff7ed; }}
    .confirme {{ color: var(--accent-2); border-color: #bbf7d0; background: #f0fdf4; }}
    .conteste {{ color: var(--warn); border-color: #fde68a; background: #fffbeb; }}
    label {{ display: block; font-size: 13px; color: #30363d; margin-bottom: 5px; }}
    input, select, textarea {{
      width: 100%;
      border: 1px solid #c7ced8;
      border-radius: 6px;
      padding: 9px 10px;
      font: inherit;
      background: white;
    }}
    textarea {{ min-height: 110px; resize: vertical; }}
    .form-grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 14px; }}
    .full {{ grid-column: 1 / -1; }}
    .flash {{ padding: 12px 14px; border-radius: 8px; border: 1px solid #bbf7d0; background: #f0fdf4; margin-bottom: 16px; }}
    .empty {{ color: var(--muted); padding: 14px; background: var(--soft); border-radius: 8px; }}
    .detail-list {{ display: grid; grid-template-columns: 190px 1fr; gap: 8px 16px; }}
    .detail-list dt {{ color: var(--muted); }}
    .detail-list dd {{ margin: 0; }}
    @media (max-width: 720px) {{
      header {{ padding: 14px 16px 0; }}
      .brand {{ display: block; }}
      .form-grid {{ grid-template-columns: 1fr; }}
      .detail-list {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="brand">
      <div>
        <h1>{APP_NAME}</h1>
        <div class="subtitle">Prototype local - données stockées sur ce PC</div>
      </div>
      <div class="subtitle">{e(title)}</div>
    </div>
    {nav}
  </header>
  <main>
    {f'<div class="flash">{e(flash)}</div>' if flash else ''}
    {body}
  </main>
</body>
</html>"""
    return html_text.encode("utf-8")


class AppHandler(BaseHTTPRequestHandler):
    username: str | None

    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"{self.address_string()} - {fmt % args}")

    def do_GET(self) -> None:
        self.route()

    def do_POST(self) -> None:
        self.route()

    def route(self) -> None:
        init_db()
        self.username = read_session(self.cookie(SESSION_COOKIE))
        path = urlparse(self.path).path
        if not users_exist() and path not in {"/setup"}:
            self.redirect("/setup")
            return
        if path in {"/setup", "/login"}:
            getattr(self, f"handle_{path.strip('/') or 'index'}")()
            return
        if not self.username:
            self.redirect("/login")
            return
        routes = {
            "/": self.handle_index,
            "/persons": self.handle_persons,
            "/persons/new": self.handle_person_new,
            "/clans": self.handle_clans,
            "/clans/new": self.handle_clan_new,
            "/groups": self.handle_groups,
            "/groups/new": self.handle_group_new,
            "/customary-names": self.handle_customary_names,
            "/customary-names/new": self.handle_customary_name_new,
            "/viva": self.handle_viva,
            "/viva/new": self.handle_viva_new,
            "/research": self.handle_research,
            "/research/new": self.handle_research_new,
            "/lineages": self.handle_lineages,
            "/lineages/new": self.handle_lineage_new,
            "/functions": self.handle_functions,
            "/functions/new": self.handle_function_new,
            "/events": self.handle_events,
            "/events/new": self.handle_event_new,
            "/sources": self.handle_sources,
            "/sources/new": self.handle_source_new,
            "/lands": self.handle_lands,
            "/lands/new": self.handle_land_new,
            "/bulk": self.handle_bulk,
            "/bulk/persons": self.handle_bulk_persons,
            "/bulk/customary-names": self.handle_bulk_customary_names,
            "/bulk/relationships": self.handle_bulk_relationships,
            "/bulk/template.xlsx": self.handle_bulk_template,
            "/bulk/excel": self.handle_bulk_excel,
            "/quick-relations": self.handle_quick_relations,
            "/quick-relations/couple": self.handle_quick_relations_couple,
            "/quick-relations/central": self.handle_quick_relations_central,
            "/visualisations": self.handle_visualisations,
            "/visualisations/extended-tree": self.handle_visualisation_extended_tree,
            "/search": self.handle_search,
            "/logout": self.handle_logout,
        }
        if path in routes:
            routes[path]()
        elif path.startswith("/persons/") and path.endswith("/groups/new"):
            self.handle_person_group_new(path)
        elif path.startswith("/persons/") and path.endswith("/delete"):
            self.handle_person_delete(path)
        elif path.startswith("/persons/") and path.endswith("/edit"):
            self.handle_person_edit(path)
        elif path.startswith("/persons/") and path.endswith("/extended-tree/print"):
            self.handle_person_extended_tree_print(path)
        elif path.startswith("/persons/") and path.endswith("/extended-tree"):
            self.handle_person_extended_tree(path)
        elif path.startswith("/persons/") and path.endswith("/tree"):
            self.handle_person_tree(path)
        elif path.startswith("/persons/") and path.endswith("/relationships/new"):
            self.handle_relationship_new(path)
        elif path.startswith("/relationships/") and path.endswith("/edit"):
            self.handle_relationship_edit(path)
        elif path.startswith("/relationships/") and path.endswith("/delete"):
            self.handle_relationship_delete(path)
        elif path == "/inferred-relationships/delete":
            self.handle_inferred_relationship_delete()
        elif path.startswith("/person-groups/") and path.endswith("/edit"):
            self.handle_person_group_edit(path)
        elif path.startswith("/person-customary-names/") and path.endswith("/edit"):
            self.handle_person_customary_name_edit(path)
        elif path.startswith("/persons/"):
            self.handle_person_detail(path)
        elif path.startswith("/clans/") and path.endswith("/delete"):
            self.handle_clan_delete(path)
        elif path.startswith("/clans/") and path.endswith("/edit"):
            self.handle_clan_edit(path)
        elif path.startswith("/clans/"):
            self.handle_clan_detail(path)
        elif path.startswith("/groups/") and path.endswith("/delete"):
            self.handle_group_delete(path)
        elif path.startswith("/groups/") and path.endswith("/edit"):
            self.handle_group_edit(path)
        elif path.startswith("/groups/") and path.endswith("/map"):
            self.handle_group_map(path)
        elif path.startswith("/groups/"):
            self.handle_group_detail(path)
        elif path.startswith("/customary-names/") and path.endswith("/delete"):
            self.handle_customary_name_delete(path)
        elif path.startswith("/customary-names/") and path.endswith("/relations/new"):
            self.handle_customary_name_relation_new(path)
        elif path.startswith("/customary-names/") and path.endswith("/edit"):
            self.handle_customary_name_edit(path)
        elif path.startswith("/customary-names/"):
            self.handle_customary_name_detail(path)
        elif path.startswith("/viva/") and path.endswith("/delete"):
            self.handle_viva_delete(path)
        elif path.startswith("/viva/") and path.endswith("/edit"):
            self.handle_viva_edit(path)
        elif path.startswith("/viva/"):
            self.handle_viva_detail(path)
        elif path.startswith("/events/") and path.endswith("/delete"):
            self.handle_event_delete(path)
        elif path.startswith("/events/") and path.endswith("/edit"):
            self.handle_event_edit(path)
        elif path.startswith("/events/") and path.endswith("/people/new"):
            self.handle_event_person_new(path)
        elif path.startswith("/event-people/") and path.endswith("/edit"):
            self.handle_event_person_edit(path)
        elif path.startswith("/sources/") and path.endswith("/edit"):
            self.handle_source_edit(path)
        elif path.startswith("/lands/") and path.endswith("/edit"):
            self.handle_land_edit(path)
        elif path.startswith("/research/") and path.endswith("/edit"):
            self.handle_research_edit(path)
        elif path.startswith("/lineages/") and path.endswith("/edit"):
            self.handle_lineage_edit(path)
        elif path.startswith("/functions/") and path.endswith("/edit"):
            self.handle_function_edit(path)
        else:
            self.not_found()

    def cookie(self, name: str) -> str | None:
        raw = self.headers.get("Cookie", "")
        for part in raw.split(";"):
            if "=" in part:
                key, value = part.strip().split("=", 1)
                if key == name:
                    return unquote(value)
        return None

    def form(self) -> dict[str, str]:
        length = int(self.headers.get("Content-Length", "0") or 0)
        raw = self.rfile.read(length).decode("utf-8") if length else ""
        parsed = parse_qs(raw, keep_blank_values=True)
        return {key: values[0].strip() for key, values in parsed.items()}

    def form_multi(self) -> dict[str, list[str]]:
        length = int(self.headers.get("Content-Length", "0") or 0)
        raw = self.rfile.read(length).decode("utf-8") if length else ""
        parsed = parse_qs(raw, keep_blank_values=True)
        return {key: [value.strip() for value in values if value.strip()] for key, values in parsed.items()}

    def send_html(self, title: str, body: str, status: HTTPStatus = HTTPStatus.OK, flash: str = "") -> None:
        data = layout(title, body, self.username, flash)
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_file(self, filename: str, data: bytes, content_type: str) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def uploaded_file(self, field_name: str) -> bytes | None:
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type or "boundary=" not in content_type:
            return None
        boundary = content_type.split("boundary=", 1)[1].strip().strip('"').encode("utf-8")
        length = int(self.headers.get("Content-Length", "0") or 0)
        raw = self.rfile.read(length)
        marker = b"--" + boundary
        for part in raw.split(marker):
            if not part or part in {b"--\r\n", b"--"}:
                continue
            header_blob, _, body = part.partition(b"\r\n\r\n")
            if not body:
                continue
            headers_text = header_blob.decode("utf-8", errors="ignore")
            if f'name="{field_name}"' not in headers_text:
                continue
            if body.endswith(b"\r\n"):
                body = body[:-2]
            return body
        return None

    def redirect(self, path: str, cookie: str | None = None, clear_cookie: bool = False) -> None:
        self.send_response(HTTPStatus.SEE_OTHER)
        self.send_header("Location", path)
        if cookie:
            self.send_header("Set-Cookie", f"{SESSION_COOKIE}={quote(cookie)}; HttpOnly; SameSite=Lax; Path=/")
        if clear_cookie:
            self.send_header("Set-Cookie", f"{SESSION_COOKIE}=; HttpOnly; SameSite=Lax; Path=/; Max-Age=0")
        self.end_headers()

    def not_found(self) -> None:
        self.send_html("Introuvable", "<div class='panel'><h2>Page introuvable</h2></div>", HTTPStatus.NOT_FOUND)

    def handle_setup(self) -> None:
        self.username = None
        if self.command == "POST":
            data = self.form()
            password = data.get("password", "")
            confirm = data.get("confirm", "")
            if len(password) < 8:
                self.send_html("Création du mot de passe", setup_form("Le mot de passe doit contenir au moins 8 caractères."))
                return
            if password != confirm:
                self.send_html("Création du mot de passe", setup_form("Les deux mots de passe ne correspondent pas."))
                return
            create_first_user(password)
            self.redirect("/login")
            return
        self.send_html("Création du mot de passe", setup_form())

    def handle_login(self) -> None:
        self.username = None
        if self.command == "POST":
            password = self.form().get("password", "")
            if authenticate(password):
                self.redirect("/", make_session("admin"))
                return
            self.send_html("Connexion", login_form("Mot de passe incorrect."))
            return
        self.send_html("Connexion", login_form())

    def handle_logout(self) -> None:
        self.redirect("/login", clear_cookie=True)

    def handle_index(self) -> None:
        with db_connect() as conn:
            counts = {
                "personnes": conn.execute("SELECT COUNT(*) FROM persons").fetchone()[0],
                "clans": conn.execute("SELECT COUNT(*) FROM clans").fetchone()[0],
                "groupes": conn.execute("SELECT COUNT(*) FROM customary_groups").fetchone()[0],
                "noms_coutumiers": conn.execute("SELECT COUNT(*) FROM customary_names").fetchone()[0],
                "viva": conn.execute("SELECT COUNT(*) FROM viva_lists").fetchone()[0],
                "infos": conn.execute("SELECT COUNT(*) FROM research_items").fetchone()[0],
                "genealogies": conn.execute("SELECT COUNT(*) FROM lineage_genealogies").fetchone()[0],
                "fonctions": conn.execute("SELECT COUNT(*) FROM customary_functions").fetchone()[0],
                "evenements": conn.execute("SELECT COUNT(*) FROM customary_events").fetchone()[0],
                "sources": conn.execute("SELECT COUNT(*) FROM sources").fetchone()[0],
                "terres": conn.execute("SELECT COUNT(*) FROM lands").fetchone()[0],
                "a_verifier": conn.execute(
                    """
                    SELECT
                      (SELECT COUNT(*) FROM persons WHERE certainty='a_verifier') +
                      (SELECT COUNT(*) FROM clans WHERE certainty='a_verifier') +
                      (SELECT COUNT(*) FROM customary_groups WHERE certainty='a_verifier') +
                      (SELECT COUNT(*) FROM customary_names WHERE certainty='a_verifier') +
                      (SELECT COUNT(*) FROM viva_lists WHERE certainty='a_verifier') +
                      (SELECT COUNT(*) FROM research_items WHERE certainty='a_verifier') +
                      (SELECT COUNT(*) FROM lineage_genealogies WHERE certainty='a_verifier') +
                      (SELECT COUNT(*) FROM customary_functions WHERE certainty='a_verifier') +
                      (SELECT COUNT(*) FROM customary_events WHERE certainty='a_verifier')
                    """
                ).fetchone()[0],
                "sensibles": conn.execute(
                    """
                    SELECT
                      (SELECT COUNT(*) FROM persons WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM clans WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM customary_groups WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM customary_names WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM viva_lists WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM research_items WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM lineage_genealogies WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM customary_functions WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM customary_events WHERE confidentiality IN ('sensible','reserve')) +
                      (SELECT COUNT(*) FROM lands WHERE confidentiality IN ('sensible','reserve'))
                    """
                ).fetchone()[0],
            }
        body = f"""
        <section class="panel">
          <h2>Actions rapides</h2>
          <div class="actions">
            <a class="button" href="/persons/new">Ajouter une personne</a>
            <a class="button" href="/clans/new">Ajouter un clan</a>
            <a class="button" href="/groups/new">Ajouter un groupe</a>
            <a class="button" href="/customary-names/new">Ajouter un nom coutumier</a>
            <a class="button" href="/viva/new">Ajouter une liste viva</a>
            <a class="button" href="/research/new">Ajouter une info</a>
            <a class="button" href="/lineages/new">Ajouter une genealogie</a>
            <a class="button" href="/functions/new">Ajouter une fonction</a>
            <a class="button" href="/events/new">Ajouter un événement</a>
            <a class="button" href="/sources/new">Ajouter une source</a>
            <a class="button secondary" href="/search">Rechercher</a>
          </div>
        </section>
        <section class="panel">
          <h2>Tableau de bord</h2>
          <div class="grid">
            {metric("Personnes", counts["personnes"])}
            {metric("Clans", counts["clans"])}
            {metric("Groupes", counts["groupes"])}
            {metric("Noms coutumiers", counts["noms_coutumiers"])}
            {metric("Viva", counts["viva"])}
            {metric("Infos", counts["infos"])}
            {metric("Genealogies", counts["genealogies"])}
            {metric("Fonctions", counts["fonctions"])}
            {metric("Événements", counts["evenements"])}
            {metric("Sources", counts["sources"])}
            {metric("Terres", counts["terres"])}
            {metric("À vérifier", counts["a_verifier"])}
            {metric("Sensibles", counts["sensibles"])}
          </div>
        </section>
        """
        self.send_html("Accueil", body)

    def handle_bulk(self, flash: str = "") -> None:
        with db_connect() as conn:
            persons = conn.execute(
                """
                SELECT
                  p.id,
                  p.first_names || ' - ' || COALESCE(NULLIF(p.birth_name, ''), 'Inconnu') ||
                    CASE WHEN p.current_name IS NOT NULL AND p.current_name != '' THEN ' (' || p.current_name || ')' ELSE '' END ||
                    CASE WHEN p.birth_date IS NOT NULL AND p.birth_date != '' THEN ' | naissance: ' || p.birth_date ELSE '' END ||
                    ' | fiche #' || p.id AS label
                FROM persons p
                ORDER BY p.birth_name, p.first_names, p.current_name
                """
            ).fetchall()
        body = bulk_form(persons)
        self.send_html("Saisie en masse", body, flash=flash)

    def handle_bulk_persons(self) -> None:
        if self.command != "POST":
            self.handle_bulk()
            return
        data = self.form()
        created = 0
        skipped = 0
        errors: list[str] = []
        with db_connect() as conn:
            for index in bulk_indices(data, "person_birth_name_"):
                birth_name = data.get(f"person_birth_name_{index}", "").strip()
                first_names = data.get(f"person_first_names_{index}", "").strip()
                if not any(
                    data.get(f"person_{field}_{index}", "").strip()
                    for field in ("birth_name", "first_names", "other_names", "gender", "current_name", "birth_date", "birth_place", "death_date", "notes")
                ):
                    continue
                if not birth_name or not first_names:
                    skipped += 1
                    errors.append(f"Ligne personne {index + 1}: nom de naissance et prénom obligatoires.")
                    continue
                gender = data.get(f"person_gender_{index}", "")
                if gender not in {"", "M", "F"}:
                    skipped += 1
                    errors.append(f"Ligne personne {index + 1}: sexe invalide.")
                    continue
                conn.execute(
                    """
                    INSERT INTO persons
                    (first_names, current_name, birth_name, other_names, gender, birth_date, birth_place, death_date,
                     current_clan_id, origin_clan_id, notes, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        first_names,
                        data.get(f"person_current_name_{index}", "").strip(),
                        birth_name,
                        data.get(f"person_other_names_{index}", "").strip(),
                        gender,
                        data.get(f"person_birth_date_{index}", "").strip(),
                        data.get(f"person_birth_place_{index}", "").strip(),
                        data.get(f"person_death_date_{index}", "").strip(),
                        None,
                        None,
                        data.get(f"person_notes_{index}", "").strip(),
                        "familial",
                        "a_verifier",
                    ),
                )
                created += 1
        self.handle_bulk(bulk_flash("Personnes", created, skipped, errors))

    def handle_bulk_customary_names(self) -> None:
        if self.command != "POST":
            self.handle_bulk()
            return
        data = self.form()
        created = 0
        skipped = 0
        errors: list[str] = []
        with db_connect() as conn:
            for index in bulk_indices(data, "customary_name_"):
                name = data.get(f"customary_name_{index}", "").strip()
                if not any(
                    data.get(f"customary_{field}_{index}", "").strip()
                    for field in ("name", "variants", "region", "description")
                ):
                    continue
                if not name:
                    skipped += 1
                    errors.append(f"Ligne nom coutumier {index + 1}: nom obligatoire.")
                    continue
                exists = conn.execute("SELECT 1 FROM customary_names WHERE lower(name) = lower(?)", (name,)).fetchone()
                if exists:
                    skipped += 1
                    errors.append(f"Ligne nom coutumier {index + 1}: {name} existe déjà.")
                    continue
                conn.execute(
                    """
                    INSERT INTO customary_names
                    (name, name_type, variants, region, description, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        name,
                        data.get(f"customary_type_{index}", "a_verifier"),
                        data.get(f"customary_variants_{index}", "").strip(),
                        data.get(f"customary_region_{index}", "").strip(),
                        data.get(f"customary_description_{index}", "").strip(),
                        None,
                        "familial",
                        "a_verifier",
                    ),
                )
                created += 1
        self.handle_bulk(bulk_flash("Noms coutumiers", created, skipped, errors))

    def handle_bulk_relationships(self) -> None:
        if self.command != "POST":
            self.handle_bulk()
            return
        data = self.form()
        created = 0
        skipped = 0
        errors: list[str] = []
        with db_connect() as conn:
            for index in bulk_indices(data, "rel_subject_"):
                subject_id = none_if_empty(data.get(f"rel_subject_{index}", ""))
                object_id = none_if_empty(data.get(f"rel_object_{index}", ""))
                relation_type = data.get(f"rel_type_{index}", "")
                if not any([subject_id, object_id, data.get(f"rel_comment_{index}", "").strip()]):
                    continue
                if not subject_id or not object_id or not relation_type:
                    skipped += 1
                    errors.append(f"Ligne relation {index + 1}: personne A, rôle et personne B sont obligatoires.")
                    continue
                if subject_id == object_id:
                    skipped += 1
                    errors.append(f"Ligne relation {index + 1}: une personne ne peut pas être reliée à elle-même.")
                    continue
                duplicate = conn.execute(
                    """
                    SELECT 1 FROM relationships
                    WHERE subject_person_id = ? AND object_person_id = ? AND relation_type = ?
                    """,
                    (subject_id, object_id, relation_type),
                ).fetchone()
                if not duplicate and relation_type in {"conjoint", "frere_soeur"}:
                    duplicate = conn.execute(
                        """
                        SELECT 1 FROM relationships
                        WHERE subject_person_id = ? AND object_person_id = ? AND relation_type = ?
                        """,
                        (object_id, subject_id, relation_type),
                    ).fetchone()
                if duplicate:
                    skipped += 1
                    errors.append(f"Ligne relation {index + 1}: relation déjà existante.")
                    continue
                conn.execute(
                    """
                    INSERT INTO relationships
                    (subject_person_id, object_person_id, object_clan_id, relation_type, comment, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        subject_id,
                        object_id,
                        None,
                        relation_type,
                        data.get(f"rel_comment_{index}", "").strip(),
                        data.get(f"rel_confidentiality_{index}", "familial"),
                        data.get(f"rel_certainty_{index}", "a_verifier"),
                    ),
                )
                created += 1
        self.handle_bulk(bulk_flash("Relations", created, skipped, errors))

    def handle_bulk_template(self) -> None:
        self.send_file(
            "modele_import_memoire_coutumiere.xlsx",
            build_bulk_template(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def handle_bulk_excel(self) -> None:
        if self.command != "POST":
            self.handle_bulk()
            return
        file_bytes = self.uploaded_file("excel_file")
        if not file_bytes:
            self.handle_bulk("Aucun fichier Excel reçu.")
            return
        try:
            report = import_bulk_workbook(file_bytes)
        except Exception as exc:
            self.handle_bulk(f"Import Excel impossible: {exc}")
            return
        self.handle_bulk(report)

    def quick_relation_persons(self) -> list[sqlite3.Row]:
        with db_connect() as conn:
            return conn.execute(
                """
                SELECT
                  id,
                  gender,
                  first_names || ' - ' || COALESCE(NULLIF(birth_name, ''), 'Inconnu') ||
                    CASE WHEN other_names IS NOT NULL AND other_names != '' THEN ' | autres: ' || other_names ELSE '' END ||
                    CASE WHEN current_name IS NOT NULL AND current_name != '' THEN ' (' || current_name || ')' ELSE '' END ||
                    CASE WHEN birth_date IS NOT NULL AND birth_date != '' THEN ' | naissance: ' || birth_date ELSE '' END ||
                    ' | fiche #' || id AS label
                FROM persons
                ORDER BY birth_name, first_names, current_name
                """
            ).fetchall()

    def handle_quick_relations(self, flash: str = "") -> None:
        persons = self.quick_relation_persons()
        self.send_html("Relations rapides", quick_relations_form(persons), flash=flash)

    def handle_quick_relations_couple(self) -> None:
        if self.command != "POST":
            self.handle_quick_relations()
            return
        data = self.form_multi()
        parent1 = first_value(data, "parent1_id")
        parent2 = first_value(data, "parent2_id")
        children = data.get("child_ids", [])
        certainty = first_value(data, "certainty", "confirme")
        confidentiality = first_value(data, "confidentiality", "familial")
        comment = first_value(data, "comment")
        created = 0
        skipped = 0
        errors: list[str] = []
        if not parent1 or not children:
            self.handle_quick_relations("Couple + enfants: parent 1 et au moins un enfant sont obligatoires.")
            return
        with db_connect() as conn:
            if parent2 and parent1 != parent2:
                made, reason = insert_spouse_relation(conn, parent1, parent2, certainty, confidentiality, comment)
                created += int(made)
                skipped += 0 if made else 1
                if reason:
                    errors.append(reason)
            for child_id in children:
                if child_id in {parent1, parent2}:
                    skipped += 1
                    errors.append("Un parent ne peut pas aussi être sélectionné comme enfant.")
                    continue
                for parent_id in [parent1, parent2]:
                    if not parent_id:
                        continue
                    made, reason = insert_parent_child_relation(conn, parent_id, child_id, certainty, confidentiality, comment)
                    created += int(made)
                    skipped += 0 if made else 1
                    if reason:
                        errors.append(reason)
        self.handle_quick_relations(bulk_flash("Relations rapides", created, skipped, errors))

    def handle_quick_relations_central(self) -> None:
        if self.command != "POST":
            self.handle_quick_relations()
            return
        data = self.form_multi()
        central = first_value(data, "central_id")
        father = first_value(data, "father_id")
        mother = first_value(data, "mother_id")
        spouses = data.get("spouse_ids", [])
        children = data.get("central_child_ids", [])
        siblings = data.get("sibling_ids", [])
        certainty = first_value(data, "certainty", "confirme")
        confidentiality = first_value(data, "confidentiality", "familial")
        comment = first_value(data, "comment")
        created = 0
        skipped = 0
        errors: list[str] = []
        if not central:
            self.handle_quick_relations("Personne centrale obligatoire.")
            return
        with db_connect() as conn:
            for parent_id in [father, mother]:
                if not parent_id:
                    continue
                if parent_id == central:
                    skipped += 1
                    errors.append("La personne centrale ne peut pas être son propre parent.")
                    continue
                made, reason = insert_parent_child_relation(conn, parent_id, central, certainty, confidentiality, comment)
                created += int(made)
                skipped += 0 if made else 1
                if reason:
                    errors.append(reason)
            for spouse_id in spouses:
                if spouse_id == central:
                    skipped += 1
                    errors.append("La personne centrale ne peut pas être son propre conjoint.")
                    continue
                made, reason = insert_spouse_relation(conn, central, spouse_id, certainty, confidentiality, comment)
                created += int(made)
                skipped += 0 if made else 1
                if reason:
                    errors.append(reason)
            for child_id in children:
                if child_id == central:
                    skipped += 1
                    errors.append("La personne centrale ne peut pas être son propre enfant.")
                    continue
                made, reason = insert_parent_child_relation(conn, central, child_id, certainty, confidentiality, comment)
                created += int(made)
                skipped += 0 if made else 1
                if reason:
                    errors.append(reason)
            for sibling_id in siblings:
                if sibling_id == central:
                    skipped += 1
                    errors.append("La personne centrale ne peut pas être son propre frère/sœur.")
                    continue
                made, reason = insert_sibling_relation(conn, central, sibling_id, certainty, confidentiality, comment)
                created += int(made)
                skipped += 0 if made else 1
                if reason:
                    errors.append(reason)
        self.handle_quick_relations(bulk_flash("Relations rapides", created, skipped, errors))

    def handle_visualisations(self) -> None:
        with db_connect() as conn:
            persons = conn.execute(
                """
                SELECT
                  p.id,
                  p.first_names,
                  p.birth_name,
                  p.current_name,
                  GROUP_CONCAT(DISTINCT cn.name) AS customary_names,
                  COUNT(DISTINCT pgl.id) AS group_count,
                  COUNT(DISTINCT rel.id) AS relation_count
                FROM persons p
                LEFT JOIN person_customary_name_links pcnl ON pcnl.person_id = p.id
                LEFT JOIN customary_names cn ON cn.id = pcnl.customary_name_id
                LEFT JOIN person_group_links pgl ON pgl.person_id = p.id
                LEFT JOIN relationships rel ON rel.subject_person_id = p.id OR rel.object_person_id = p.id
                GROUP BY p.id
                ORDER BY p.birth_name, p.first_names, p.current_name
                """
            ).fetchall()
            groups = conn.execute(
                """
                SELECT
                  g.id,
                  g.name,
                  g.group_type,
                  COUNT(DISTINCT pgl.id) AS person_count,
                  COUNT(DISTINCT gr.id) AS relation_count
                FROM customary_groups g
                LEFT JOIN person_group_links pgl ON pgl.group_id = g.id
                LEFT JOIN group_relations gr ON gr.group_a_id = g.id OR gr.group_b_id = g.id
                GROUP BY g.id
                ORDER BY g.name
                """
            ).fetchall()
        person_visualisation_table = rows_table(
            ["Nom de naissance", "Prénom", "Nom marital", "Nom coutumier", "Groupes associés", "Relations connues", "Action"],
            [
                [
                    e(row["birth_name"]),
                    f'<a href="/persons/{row["id"]}">{e(row["first_names"])}</a>',
                    e(row["current_name"]),
                    e(row["customary_names"]),
                    str(row["group_count"]),
                    str(row["relation_count"]),
                    f"<a class='button secondary' href='/persons/{row['id']}/tree'>Voir l'arbre</a>",
                ]
                for row in persons
            ],
        )
        body = f"""
        <section class="panel">
          <h2>Visualisations</h2>
          <div class="actions"><a class="button" href="/visualisations/extended-tree">Ouvrir un arbre étendu</a></div>
          <p class="empty">Cette page sert de point d'entrée pour lire les liens autrement que dans les tableaux : d'abord les arbres familiaux des personnes, puis les cartes coutumières des groupes, lignages ou noms cités.</p>
        </section>
        <section class="panel">
          <h2>Arbres généalogiques</h2>
          {person_visualisation_table}
        </section>
        <section class="panel">
          <h2>Cartes coutumières</h2>
          {rows_table(["Groupe / lignage", "Type", "Personnes liées", "Relations connues", "Action"], [
              [f'<a href="/groups/{row["id"]}">{e(row["name"])}</a>', e(group_type_label(row["group_type"])), str(row["person_count"]), str(row["relation_count"]), f'<a class="button secondary" href="/groups/{row["id"]}/map">Voir la carte</a>']
              for row in groups
          ])}
        </section>
        """
        self.send_html("Visualisations", body)

    def handle_visualisation_extended_tree(self) -> None:
        query = parse_qs(urlparse(self.path).query)
        person_id = (query.get("person_id") or [""])[0]
        if person_id:
            params = {
                "mode": (query.get("mode") or ["descendance"])[0],
                "generations": (query.get("generations") or ["8"])[0],
            }
            if (query.get("same_name") or [""])[0] == "1":
                params["same_name"] = "1"
            if (query.get("spouses") or [""])[0] == "1":
                params["spouses"] = "1"
            if (query.get("name_color") or [""])[0] == "1":
                params["name_color"] = "1"
            if (query.get("branch_color") or [""])[0] == "1":
                params["branch_color"] = "1"
            self.redirect(f"/persons/{person_id}/extended-tree?{urlencode(params)}")
            return
        persons = self.quick_relation_persons()
        body = extended_tree_picker_form(persons)
        self.send_html("Choisir un arbre étendu", body)

    def handle_persons(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT p.*, GROUP_CONCAT(DISTINCT cn.name) AS customary_names
                FROM persons p
                LEFT JOIN person_customary_name_links pcnl ON pcnl.person_id = p.id
                LEFT JOIN customary_names cn ON cn.id = pcnl.customary_name_id
                GROUP BY p.id
                ORDER BY p.birth_name, p.first_names, p.current_name
                """
            ).fetchall()
        table = rows_table(
            ["Nom de naissance", "Prénom", "Sexe", "Nom marital", "Nom coutumier", "Naissance", "Statut", "Confidentialité", "Action"],
            [
                [
                    e(row["birth_name"]),
                    f'<a href="/persons/{row["id"]}">{e(row["first_names"])}</a>',
                    e(sex_label(row["gender"])),
                    e(row["current_name"]),
                    e(row["customary_names"]),
                    e(row["birth_date"]),
                    "Décédé" if row["death_date"] else "Vivant / inconnu",
                    badge(row["confidentiality"]),
                    delete_button(f'/persons/{row["id"]}/delete', "Personne", f'{row["first_names"]} - {row["birth_name"] or "Inconnu"}'),
                ]
                for row in rows
            ],
        )
        body = f"""
        <section class="panel">
          <div class="actions"><a class="button" href="/persons/new">Ajouter une personne</a></div>
          {table}
        </section>
        """
        self.send_html("Personnes", body)

    def handle_person_new(self) -> None:
        with db_connect() as conn:
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO persons
                    (first_names, current_name, birth_name, other_names, gender, birth_date, birth_place, death_date,
                     current_clan_id, origin_clan_id, notes, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["first_names"],
                        data.get("current_name", ""),
                        data["birth_name"],
                        data.get("other_names"),
                        data.get("gender"),
                        data.get("birth_date"),
                        data.get("birth_place"),
                        data.get("death_date"),
                        None,
                        None,
                        data.get("notes"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/persons")
                return
        body = person_form()
        self.send_html("Ajouter une personne", body)

    def handle_person_edit(self, path: str) -> None:
        try:
            person_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            person = conn.execute("SELECT * FROM persons WHERE id = ?", (person_id,)).fetchone()
            if not person:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE persons
                    SET first_names = ?, current_name = ?, birth_name = ?, other_names = ?, gender = ?,
                        birth_date = ?, birth_place = ?, death_date = ?, notes = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["first_names"],
                        data.get("current_name", ""),
                        data["birth_name"],
                        data.get("other_names"),
                        data.get("gender"),
                        data.get("birth_date"),
                        data.get("birth_place"),
                        data.get("death_date"),
                        data.get("notes"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        person_id,
                    ),
                )
                self.redirect(f"/persons/{person_id}")
                return
        fields = [
            text_field("first_names", "Prénom principal *", row_value(person, "first_names"), True),
            text_field("birth_name", "Nom de naissance *", row_value(person, "birth_name"), True),
            text_field("current_name", "Nom marital / nom actuel", row_value(person, "current_name")),
            text_field("other_names", "Autres noms", row_value(person, "other_names")),
            select_field("gender", "Sexe", SEXES, row_value(person, "gender")),
            text_field("birth_date", "Date ou période de naissance", row_value(person, "birth_date")),
            text_field("birth_place", "Lieu de naissance", row_value(person, "birth_place")),
            text_field("death_date", "Date ou période de décès", row_value(person, "death_date")),
            select_field("confidentiality", "Confidentialité de la fiche *", CONFIDENTIALITY, row_value(person, "confidentiality"), True),
            select_field("certainty", "Certitude de la fiche *", CERTAINTY, row_value(person, "certainty"), True),
            textarea_field("notes", "Notes", row_value(person, "notes")),
        ]
        self.send_html("Modifier une personne", edit_form_panel("Modifier la personne", fields, f"/persons/{person_id}"))

    def handle_person_delete(self, path: str) -> None:
        if self.command != "POST":
            self.redirect("/persons")
            return
        try:
            person_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            conn.execute("DELETE FROM relationships WHERE subject_person_id = ? OR object_person_id = ?", (person_id, person_id))
            conn.execute("DELETE FROM person_group_links WHERE person_id = ?", (person_id,))
            conn.execute("DELETE FROM person_customary_name_links WHERE person_id = ?", (person_id,))
            conn.execute("DELETE FROM event_person_links WHERE person_id = ?", (person_id,))
            conn.execute("UPDATE customary_events SET main_person_id = NULL WHERE main_person_id = ?", (person_id,))
            conn.execute("UPDATE customary_functions SET person_id = NULL WHERE person_id = ?", (person_id,))
            conn.execute("DELETE FROM persons WHERE id = ?", (person_id,))
        self.redirect("/persons")

    def handle_person_detail(self, path: str) -> None:
        try:
            person_id = int(path.rsplit("/", 1)[-1])
        except ValueError:
            self.not_found()
            return
        with db_connect() as conn:
            person = conn.execute(
                """
                SELECT p.*, c.name AS current_clan, oc.name AS origin_clan
                FROM persons p
                LEFT JOIN clans c ON c.id = p.current_clan_id
                LEFT JOIN clans oc ON oc.id = p.origin_clan_id
                WHERE p.id = ?
                """,
                (person_id,),
            ).fetchone()
            if not person:
                self.not_found()
                return
            materialize_inferred_sibling_relationships(conn, person_id)
            relationships = conn.execute(
                """
                SELECT
                  r.*,
                  sp.gender AS subject_gender,
                  op.gender AS object_gender,
                  sp.first_names || ' - ' || COALESCE(NULLIF(sp.birth_name, ''), 'Inconnu') ||
                    CASE WHEN sp.current_name IS NOT NULL AND sp.current_name != '' THEN ' (' || sp.current_name || ')' ELSE '' END AS subject_person,
                  op.first_names || ' - ' || COALESCE(NULLIF(op.birth_name, ''), 'Inconnu') ||
                    CASE WHEN op.current_name IS NOT NULL AND op.current_name != '' THEN ' (' || op.current_name || ')' ELSE '' END AS object_person,
                  CASE
                    WHEN r.subject_person_id = ? THEN op.first_names || ' - ' || COALESCE(NULLIF(op.birth_name, ''), 'Inconnu') ||
                      CASE WHEN op.current_name IS NOT NULL AND op.current_name != '' THEN ' (' || op.current_name || ')' ELSE '' END
                    ELSE sp.first_names || ' - ' || COALESCE(NULLIF(sp.birth_name, ''), 'Inconnu') ||
                      CASE WHEN sp.current_name IS NOT NULL AND sp.current_name != '' THEN ' (' || sp.current_name || ')' ELSE '' END
                  END AS related_person
                FROM relationships r
                JOIN persons sp ON sp.id = r.subject_person_id
                LEFT JOIN persons op ON op.id = r.object_person_id
                WHERE r.subject_person_id = ? OR r.object_person_id = ?
                ORDER BY r.relation_type, related_person
                """,
                (person_id, person_id, person_id),
            ).fetchall()
            events = conn.execute(
                """
                SELECT ev.*, 'principal' AS person_role
                FROM customary_events ev
                WHERE ev.main_person_id = ?
                UNION
                SELECT ev.*, epl.role AS person_role
                FROM event_person_links epl
                JOIN customary_events ev ON ev.id = epl.event_id
                WHERE epl.person_id = ?
                  AND epl.role IN ('epoux','epouse','enfant_redonne','oncle_maternel_concerne','enfant_naissance','defunt')
                ORDER BY event_date
                """,
                (person_id, person_id),
            ).fetchall()
            group_links = conn.execute(
                """
                SELECT pgl.*, g.name AS group_name, g.group_type
                FROM person_group_links pgl
                JOIN customary_groups g ON g.id = pgl.group_id
                WHERE pgl.person_id = ?
                ORDER BY pgl.link_type, g.name
                """,
                (person_id,),
            ).fetchall()
            customary_name_links = conn.execute(
                """
                SELECT pcnl.*, cn.name AS customary_name, cn.name_type
                FROM person_customary_name_links pcnl
                JOIN customary_names cn ON cn.id = pcnl.customary_name_id
                WHERE pcnl.person_id = ?
                ORDER BY pcnl.link_type, cn.name
                """,
                (person_id,),
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>{e(person["first_names"])} {e(person["current_name"])}</h2>
          <div class="actions"><a class="button" href="/persons/{person_id}/edit">Modifier</a> <a class="button" href="/persons/{person_id}/tree">Voir l'arbre</a> <a class="button" href="/persons/{person_id}/extended-tree">Arbre étendu</a></div>
          <dl class="detail-list">
            <dt>Nom de naissance</dt><dd>{e(person["birth_name"])}</dd>
            <dt>Nom marital / nom actuel</dt><dd>{e(person["current_name"])}</dd>
            <dt>Autres noms</dt><dd>{e(person["other_names"])}</dd>
            <dt>Sexe</dt><dd>{e(sex_label(person["gender"]))}</dd>
            <dt>Naissance</dt><dd>{e(person["birth_date"])} - {e(person["birth_place"])}</dd>
            <dt>Décès</dt><dd>{e(person["death_date"])}</dd>
            <dt>Certitude</dt><dd>{badge(person["certainty"])}</dd>
            <dt>Confidentialité</dt><dd>{badge(person["confidentiality"])}</dd>
            <dt>Notes</dt><dd>{e(person["notes"])}</dd>
          </dl>
        </section>
        <section class="panel">
          <h2>Événements personnels / structurants</h2>
          {rows_table(["Date", "Type", "Titre", "Rôle", "Certitude", "Confidentialité"], [
              [e(ev["event_date"]), e(event_label(ev["event_type"])), e(ev["title"]), e("Personne principale" if ev["person_role"] == "principal" else event_person_role_label(ev["person_role"])), badge(ev["certainty"]), badge(ev["confidentiality"])]
              for ev in events
          ])}
        </section>
        <section class="panel">
          <h2>Appartenance coutumière</h2>
          <div class="actions"><a class="button" href="/persons/{person_id}/groups/new">Modifier l'appartenance</a></div>
          <h3>Noms coutumiers</h3>
          {rows_table(["Nom coutumier", "Type", "Lien", "Certitude", "Confidentialité", "Commentaire", "Action"], [
              [f'<a href="/customary-names/{cn["customary_name_id"]}">{e(cn["customary_name"])}</a>', e(group_type_label(cn["name_type"])), e(person_group_link_label(cn["link_type"])), badge(cn["certainty"]), badge(cn["confidentiality"]), e(cn["comment"]), f'<a class="button secondary" href="/person-customary-names/{cn["id"]}/edit">Modifier</a>']
              for cn in customary_name_links
          ])}
          <h3>Groupes / lignages</h3>
          {rows_table(["Groupe", "Type", "Lien", "Certitude", "Confidentialité", "Commentaire", "Action"], [
              [f'<a href="/groups/{gl["group_id"]}">{e(gl["group_name"])}</a>', e(group_type_label(gl["group_type"])), e(person_group_link_label(gl["link_type"])), badge(gl["certainty"]), badge(gl["confidentiality"]), e(gl["comment"]), f'<a class="button secondary" href="/person-groups/{gl["id"]}/edit">Modifier</a>']
              for gl in group_links
          ])}
        </section>
        <section class="panel">
          <h2>Relations</h2>
          <div class="actions"><a class="button" href="/persons/{person_id}/relationships/new">Ajouter une relation</a></div>
          {rows_table(["Relation", "Avec", "Certitude", "Confidentialité", "Commentaire", "Action"], [
              [e(relation_sentence_for_view(person_id, person["gender"], r)), e(r["related_person"]), badge(r["certainty"]), badge(r["confidentiality"]), e(r["comment"]), f'<a class="button secondary" href="/relationships/{r["id"]}/edit">Modifier</a> {delete_button(f"/relationships/{r["id"]}/delete", "Relation", relation_sentence_for_view(person_id, person["gender"], r))}']
              for r in relationships
          ])}
        </section>
        """
        self.send_html("Fiche personne", body)

    def handle_person_tree(self, path: str) -> None:
        try:
            person_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        query = parse_qs(urlparse(self.path).query)
        color_mode = graph_color_mode_from_query(query)
        with db_connect() as conn:
            person = conn.execute(
                """
                SELECT p.*, c.name AS current_clan
                FROM persons p
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE p.id = ?
                """,
                (person_id,),
            ).fetchone()
            if not person:
                self.not_found()
                return
            parents = conn.execute(
                """
                SELECT p.id, p.first_names, p.current_name, p.birth_name, p.gender, p.birth_date, p.death_date, c.name AS current_clan, r.relation_type, r.certainty
                FROM relationships r
                JOIN persons p ON p.id = r.subject_person_id
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE r.object_person_id = ? AND r.relation_type IN ('pere','mere','parent_coutumier')
                UNION
                SELECT p.id, p.first_names, p.current_name, p.birth_name, p.gender, p.birth_date, p.death_date, c.name AS current_clan, r.relation_type, r.certainty
                FROM relationships r
                JOIN persons p ON p.id = r.object_person_id
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE r.subject_person_id = ? AND r.relation_type = 'enfant'
                ORDER BY relation_type, current_name
                """,
                (person_id, person_id),
            ).fetchall()
            children = conn.execute(
                """
                SELECT p.id, p.first_names, p.current_name, p.birth_name, p.gender, p.birth_date, p.death_date, c.name AS current_clan, r.relation_type, r.certainty
                FROM relationships r
                JOIN persons p ON p.id = r.object_person_id
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE r.subject_person_id = ? AND r.relation_type IN ('pere','mere','parent_coutumier')
                UNION
                SELECT p.id, p.first_names, p.current_name, p.birth_name, p.gender, p.birth_date, p.death_date, c.name AS current_clan, r.relation_type, r.certainty
                FROM relationships r
                JOIN persons p ON p.id = r.subject_person_id
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE r.object_person_id = ? AND r.relation_type = 'enfant'
                ORDER BY birth_date, current_name
                """,
                (person_id, person_id),
            ).fetchall()
            spouses = conn.execute(
                """
                SELECT p.id, p.first_names, p.current_name, p.birth_name, p.gender, p.birth_date, p.death_date, c.name AS current_clan, r.relation_type, r.certainty
                FROM relationships r
                JOIN persons p ON p.id = r.object_person_id
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE r.subject_person_id = ? AND r.relation_type = 'conjoint'
                UNION
                SELECT p.id, p.first_names, p.current_name, p.birth_name, p.gender, p.birth_date, p.death_date, c.name AS current_clan, r.relation_type, r.certainty
                FROM relationships r
                JOIN persons p ON p.id = r.subject_person_id
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE r.object_person_id = ? AND r.relation_type = 'conjoint'
                ORDER BY current_name
                """,
                (person_id, person_id),
            ).fetchall()
            group_links = conn.execute(
                """
                SELECT g.id AS group_id, g.name AS group_name, g.group_type, pgl.link_type, pgl.certainty
                FROM person_group_links pgl
                JOIN customary_groups g ON g.id = pgl.group_id
                WHERE pgl.person_id = ?
                ORDER BY pgl.link_type, g.name
                """,
                (person_id,),
            ).fetchall()
            structural_events = conn.execute(
                f"""
                SELECT ev.id, ev.title, ev.event_type, ev.event_date, epl.role, epl.certainty
                FROM event_person_links epl
                JOIN customary_events ev ON ev.id = epl.event_id
                WHERE epl.person_id = ?
                  AND epl.role IN ({",".join("?" for _ in PERSON_EVENT_VISIBLE_ROLES)})
                UNION
                SELECT ev.id, ev.title, ev.event_type, ev.event_date, 'principal' AS role, ev.certainty
                FROM customary_events ev
                WHERE ev.main_person_id = ?
                ORDER BY event_date, title
                """,
                (person_id, *sorted(PERSON_EVENT_VISIBLE_ROLES), person_id),
            ).fetchall()
        svg = render_person_tree(person, parents, spouses, children, group_links, structural_events, color_mode)
        checked_name = " checked" if color_mode == "birth_name" else ""
        checked_branch = " checked" if color_mode == "branch" else ""
        color_form = f"""
          <form method="get" action="#tree-view" class="form-grid">
            <p><label><input type="checkbox" name="name_color" value="1"{checked_name}> Couleur par nom de naissance</label></p>
            <p><label><input type="checkbox" name="branch_color" value="1"{checked_branch}> Couleur par branche descendante</label></p>
            <p class="full"><button type="submit">Actualiser les couleurs</button></p>
          </form>
        """
        body = f"""
        <section class="panel">
          <div class="actions"><a class="button secondary" href="/persons/{person_id}">Retour fiche personne</a></div>
          <h2>Arbre autour de {e(person["first_names"])} {e(person["current_name"])}</h2>
          <p class="empty">Première vue de proximité : parents, conjoints et enfants connus dans la base.</p>
          {color_form}
          {svg}
        </section>
        <section class="panel">
          <h2>Lecture</h2>
          <p class="empty">Les traits pointillés indiquent des liens à vérifier ou probables. Cette vue deviendra plus riche quand davantage de relations seront saisies.</p>
        </section>
        """
        self.send_html("Arbre personne", body)

    def handle_person_extended_tree(self, path: str) -> None:
        try:
            person_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        query = parse_qs(urlparse(self.path).query)
        generations = int((query.get("generations") or ["8"])[0] or "8")
        generations = max(1, min(generations, 8))
        mode = (query.get("mode") or ["descendance"])[0]
        same_name = (query.get("same_name") or ["0"])[0] == "1"
        show_spouses = (query.get("spouses") or ["1"])[0] == "1"
        show_upstream = (query.get("upstream") or ["0"])[0] == "1"
        color_mode = graph_color_mode_from_query(query)
        with db_connect() as conn:
            person = conn.execute(
                """
                SELECT p.*, c.name AS current_clan
                FROM persons p
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE p.id = ?
                """,
                (person_id,),
            ).fetchone()
            if not person:
                self.not_found()
                return
            if mode == "ascendance":
                levels, links, spouses = build_ancestor_levels(conn, person, generations, show_spouses)
            else:
                levels, links, spouses = build_descendant_levels(conn, person, generations, same_name, show_spouses, show_upstream)
        nav_params = {"generations": str(generations)}
        if same_name:
            nav_params["same_name"] = "1"
        if show_spouses:
            nav_params["spouses"] = "1"
        if color_mode == "birth_name":
            nav_params["name_color"] = "1"
        if color_mode == "branch":
            nav_params["branch_color"] = "1"
        return_params = dict(nav_params)
        return_params["mode"] = mode
        generation_start = -1 if show_upstream and mode != "ascendance" and len(levels) > 1 else 0
        svg = render_extended_tree(levels, links, spouses, person_id, mode, color_mode, urlencode(nav_params), generation_start)
        checked_same_name = " checked" if same_name else ""
        checked_spouses = " checked" if show_spouses else ""
        checked_upstream = " checked" if show_upstream else ""
        checked_name = " checked" if color_mode == "birth_name" else ""
        checked_branch = " checked" if color_mode == "branch" else ""
        mode_options = select_options([("descendance", "Descendance"), ("ascendance", "Ascendance")], mode)
        generation_options = select_options([(str(i), str(i)) for i in (3, 5, 8)], str(generations))
        print_params = dict(nav_params)
        print_params["mode"] = mode
        if show_upstream:
            print_params["upstream"] = "1"
        print_href = f"/persons/{person_id}/extended-tree/print?{urlencode(print_params)}"
        body = f"""
        <section class="panel">
          <div class="actions"><a class="button secondary" href="/persons/{person_id}">Retour fiche personne</a> <a class="button secondary" href="/persons/{person_id}/tree">Arbre proche</a> <a class="button" href="{print_href}">Aperçu PDF / impression</a></div>
          <h2>Arbre étendu - {e(person["first_names"])} {e(person["birth_name"])}</h2>
          <form method="get" class="form-grid">
            <p><label>Mode</label><select name="mode">{mode_options}</select></p>
            <p><label>Générations</label><select name="generations">{generation_options}</select></p>
            <p><label><input type="checkbox" name="same_name" value="1"{checked_same_name}> Même nom de naissance seulement</label></p>
            <p><label><input type="checkbox" name="spouses" value="1"{checked_spouses}> Afficher conjoints</label></p>
            <p><label><input type="checkbox" name="upstream" value="1"{checked_upstream}> Afficher parents et frères/sœurs en amont</label></p>
            <p><label><input type="checkbox" name="name_color" value="1"{checked_name}> Couleur par nom de naissance</label></p>
            <p><label><input type="checkbox" name="branch_color" value="1"{checked_branch}> Couleur par branche descendante</label></p>
            <p class="full"><button type="submit">Actualiser</button></p>
          </form>
        </section>
        <section class="panel" id="tree-view">
          {svg}
        </section>
        <section class="panel">
          <h2>Lecture</h2>
          <p class="empty">Les colonnes correspondent aux générations. Cette vue privilégie la lisibilité et le défilement horizontal pour les grands arbres.</p>
        </section>
        """
        self.send_html("Arbre étendu", body)

    def handle_person_extended_tree_print(self, path: str) -> None:
        try:
            person_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        query = parse_qs(urlparse(self.path).query)
        generations = int((query.get("generations") or ["8"])[0] or "8")
        generations = max(1, min(generations, 8))
        mode = (query.get("mode") or ["descendance"])[0]
        same_name = (query.get("same_name") or ["0"])[0] == "1"
        show_spouses = (query.get("spouses") or ["1"])[0] == "1"
        show_upstream = (query.get("upstream") or ["0"])[0] == "1"
        color_mode = graph_color_mode_from_query(query)
        paper = (query.get("paper") or ["A3"])[0].upper()
        if paper not in {"A3", "A4"}:
            paper = "A3"
        orientation = (query.get("orientation") or ["landscape"])[0]
        if orientation not in {"portrait", "landscape"}:
            orientation = "landscape"
        hide_buttons = (query.get("hide_buttons") or ["1"])[-1] == "1"
        with db_connect() as conn:
            person = conn.execute(
                """
                SELECT p.*, c.name AS current_clan
                FROM persons p
                LEFT JOIN clans c ON c.id = p.current_clan_id
                WHERE p.id = ?
                """,
                (person_id,),
            ).fetchone()
            if not person:
                self.not_found()
                return
            if mode == "ascendance":
                levels, links, spouses = build_ancestor_levels(conn, person, generations, show_spouses)
            else:
                levels, links, spouses = build_descendant_levels(conn, person, generations, same_name, show_spouses, show_upstream)
        nav_params = {"generations": str(generations)}
        if same_name:
            nav_params["same_name"] = "1"
        if show_spouses:
            nav_params["spouses"] = "1"
        if show_upstream:
            nav_params["upstream"] = "1"
        if color_mode == "birth_name":
            nav_params["name_color"] = "1"
        if color_mode == "branch":
            nav_params["branch_color"] = "1"
        return_params = dict(nav_params)
        return_params["mode"] = mode
        generation_start = -1 if show_upstream and mode != "ascendance" and len(levels) > 1 else 0
        svg = render_extended_tree(levels, links, spouses, person_id, mode, color_mode, urlencode(nav_params), generation_start, not hide_buttons)
        paper_options = select_options([("A4", "A4"), ("A3", "A3")], paper)
        orientation_options = select_options([("landscape", "Paysage"), ("portrait", "Portrait")], orientation)
        checked_hide_buttons = " checked" if hide_buttons else ""
        hide_button_css = ".print-page .mini-button { display: none; }" if hide_buttons else ""
        hidden_values = {
            "mode": mode,
            "generations": str(generations),
            **({"same_name": "1"} if same_name else {}),
            **({"spouses": "1"} if show_spouses else {}),
            **({"upstream": "1"} if show_upstream else {}),
            **({"name_color": "1"} if color_mode == "birth_name" else {}),
            **({"branch_color": "1"} if color_mode == "branch" else {}),
        }
        hidden_fields = "".join(f'<input type="hidden" name="{e(key)}" value="{e(value)}">' for key, value in hidden_values.items())
        screen_width, screen_height = (420, 297) if paper == "A3" else (297, 210)
        if orientation == "portrait":
            screen_width, screen_height = screen_height, screen_width
        body = f"""
        <style>
          .print-toolbar {{ margin-bottom: 16px; }}
          .print-page {{
            width: {screen_width}mm;
            height: {screen_height}mm;
            margin: 0 auto;
            padding: 10mm;
            box-sizing: border-box;
            background: #fff;
            border: 1px solid var(--line);
            overflow: hidden;
          }}
          .print-title {{ height: 18mm; margin: 0 0 4mm; }}
          .print-title h2 {{ margin: 0 0 2mm; }}
          .print-title p {{ margin: 0; color: var(--muted); }}
          .print-page .graph-wrap {{ height: calc(100% - 24mm); overflow: hidden; padding: 0; }}
          .print-page .graph-svg {{ width: 100%; height: 100%; min-width: 0; border: 0; background: #fff; }}
          {hide_button_css}
          @page {{ size: {paper} {orientation}; margin: 8mm; }}
          @media print {{
            header, .print-toolbar {{ display: none !important; }}
            main {{ max-width: none; margin: 0; padding: 0; }}
            body {{ background: #fff; }}
            .panel {{ border: 0; box-shadow: none; margin: 0; padding: 0; }}
            .print-page {{
              width: {screen_width - 16}mm;
              height: {screen_height - 16}mm;
              border: 0;
              margin: 0;
              padding: 0;
              overflow: hidden;
            }}
            .print-page .graph-wrap {{ height: calc(100% - 24mm); }}
            .print-page .graph-svg {{ display: block; }}
          }}
        </style>
        <section class="panel print-toolbar">
          <div class="actions"><a class="button secondary" href="/persons/{person_id}/extended-tree?{urlencode(return_params)}#tree-view">Retour arbre</a> <button type="button" onclick="window.print()">Imprimer / enregistrer en PDF</button></div>
          <h2>Aperçu avant impression</h2>
          <form method="get" class="form-grid">
            {hidden_fields}
            <p><label>Format papier</label><select name="paper">{paper_options}</select></p>
            <p><label>Orientation</label><select name="orientation">{orientation_options}</select></p>
            <p><label><input type="hidden" name="hide_buttons" value="0"><input type="checkbox" name="hide_buttons" value="1"{checked_hide_buttons}> Masquer les boutons dans l'aperçu/PDF</label></p>
            <p class="full"><button type="submit">Actualiser l'aperçu</button></p>
          </form>
          <p class="empty">Dans la fenêtre d'impression, choisis "Enregistrer au format PDF". L'arbre est automatiquement réduit pour tenir dans la feuille.</p>
        </section>
        <section class="panel">
          <div class="print-page">
            <div class="print-title">
              <h2>Arbre étendu - {e(person["first_names"])} {e(person["birth_name"])}</h2>
              <p>{e(paper)} {e("paysage" if orientation == "landscape" else "portrait")} - {e("descendance" if mode == "descendance" else "ascendance")}</p>
            </div>
            {svg}
          </div>
        </section>
        """
        self.send_html("Aperçu PDF", body)

    def handle_relationship_new(self, path: str) -> None:
        try:
            person_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            person = conn.execute(
                "SELECT id, first_names, current_name, birth_name FROM persons WHERE id = ?",
                (person_id,),
            ).fetchone()
            if not person:
                self.not_found()
                return
            persons = conn.execute(
                """
                SELECT
                  p.id,
                  p.first_names || ' - ' || COALESCE(NULLIF(p.birth_name, ''), 'Inconnu') ||
                    CASE WHEN p.current_name IS NOT NULL AND p.current_name != '' THEN ' (' || p.current_name || ')' ELSE '' END ||
                    CASE
                      WHEN COALESCE(NULLIF(p.birth_date, ''), NULLIF(p.birth_place, ''), NULLIF(p.death_date, '')) IS NOT NULL THEN
                        ' | ' ||
                        TRIM(
                          CASE WHEN p.birth_date IS NOT NULL AND p.birth_date != '' THEN 'naissance: ' || p.birth_date ELSE '' END ||
                          CASE WHEN p.birth_place IS NOT NULL AND p.birth_place != '' THEN ' à ' || p.birth_place ELSE '' END ||
                          CASE WHEN p.death_date IS NOT NULL AND p.death_date != '' THEN ' | décès: ' || p.death_date ELSE '' END
                        )
                      ELSE ''
                    END ||
                    ' | fiche #' || p.id AS label,
                  COALESCE(GROUP_CONCAT(DISTINCT cn.name), '') AS customary_names
                FROM persons p
                LEFT JOIN person_customary_name_links pcnl ON pcnl.person_id = p.id
                LEFT JOIN customary_names cn ON cn.id = pcnl.customary_name_id
                WHERE p.id != ?
                GROUP BY p.id
                ORDER BY p.birth_name, p.first_names
                """,
                (person_id,),
            ).fetchall()
            if self.command == "POST":
                data = self.form()
                object_person_id = none_if_empty(data.get("object_person_id"))
                if not object_person_id:
                    body = relationship_form(person, persons, "Choisis une personne à relier.")
                    self.send_html("Ajouter une relation", body)
                    return
                conn.execute(
                    """
                    INSERT INTO relationships
                    (subject_person_id, object_person_id, object_clan_id, relation_type, comment, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        person_id,
                        object_person_id,
                        None,
                        data["relation_type"],
                        data.get("comment"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect(f"/persons/{person_id}")
                return
        self.send_html("Ajouter une relation", relationship_form(person, persons))

    def handle_relationship_edit(self, path: str) -> None:
        try:
            relationship_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            relation = conn.execute("SELECT * FROM relationships WHERE id = ?", (relationship_id,)).fetchone()
            if not relation:
                self.not_found()
                return
            subject_person = conn.execute(
                "SELECT id, first_names, current_name, birth_name FROM persons WHERE id = ?",
                (relation["subject_person_id"],),
            ).fetchone()
            if not subject_person:
                self.not_found()
                return
            persons = conn.execute(
                """
                SELECT
                  p.id,
                  p.first_names || ' - ' || COALESCE(NULLIF(p.birth_name, ''), 'Inconnu') ||
                    CASE WHEN p.current_name IS NOT NULL AND p.current_name != '' THEN ' (' || p.current_name || ')' ELSE '' END ||
                    CASE
                      WHEN COALESCE(NULLIF(p.birth_date, ''), NULLIF(p.birth_place, ''), NULLIF(p.death_date, '')) IS NOT NULL THEN
                        ' | ' ||
                        TRIM(
                          CASE WHEN p.birth_date IS NOT NULL AND p.birth_date != '' THEN 'naissance: ' || p.birth_date ELSE '' END ||
                          CASE WHEN p.birth_place IS NOT NULL AND p.birth_place != '' THEN ' à ' || p.birth_place ELSE '' END ||
                          CASE WHEN p.death_date IS NOT NULL AND p.death_date != '' THEN ' | décès: ' || p.death_date ELSE '' END
                        )
                      ELSE ''
                    END ||
                    ' | fiche #' || p.id AS label,
                  COALESCE(GROUP_CONCAT(DISTINCT cn.name), '') AS customary_names
                FROM persons p
                LEFT JOIN person_customary_name_links pcnl ON pcnl.person_id = p.id
                LEFT JOIN customary_names cn ON cn.id = pcnl.customary_name_id
                WHERE p.id != ?
                GROUP BY p.id
                ORDER BY p.birth_name, p.first_names
                """,
                (relation["subject_person_id"],),
            ).fetchall()
            if self.command == "POST":
                data = self.form()
                object_person_id = none_if_empty(data.get("object_person_id"))
                conn.execute(
                    """
                    UPDATE relationships
                    SET object_person_id = ?, object_clan_id = ?, relation_type = ?, comment = ?,
                        confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        object_person_id,
                        None,
                        data["relation_type"],
                        data.get("comment"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        relationship_id,
                    ),
                )
                self.redirect(f"/persons/{relation['subject_person_id']}")
                return
        fields = [
            relation_type_select("relation_type", f"Rôle de {person_display(subject_person)} envers la personne liée *", row_value(relation, "relation_type")),
            person_relation_select("object_person_id", "Personne liée *", persons, row_value(relation, "object_person_id")),
            customary_name_display_field(persons, row_value(relation, "object_person_id")),
            relationship_preview(person_display(subject_person), row_value(relation, "relation_type")),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(relation, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(relation, "certainty"), True),
            textarea_field("comment", "Commentaire", row_value(relation, "comment")),
        ]
        self.send_html("Modifier une relation", edit_form_panel("Modifier la relation", fields, f"/persons/{relation['subject_person_id']}"))

    def handle_relationship_delete(self, path: str) -> None:
        try:
            relationship_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            relation = conn.execute("SELECT * FROM relationships WHERE id = ?", (relationship_id,)).fetchone()
            if not relation:
                self.not_found()
                return
            redirect_person_id = relation["subject_person_id"] or relation["object_person_id"]
            if relation["relation_type"] == "frere_soeur" and relation["object_person_id"]:
                left_id, right_id = sorted((int(relation["subject_person_id"]), int(relation["object_person_id"])))
                conn.execute(
                    """
                    INSERT OR IGNORE INTO inferred_relationship_exclusions
                    (person_a_id, person_b_id, relation_type)
                    VALUES (?, ?, ?)
                    """,
                    (left_id, right_id, "frere_soeur"),
                )
            conn.execute("DELETE FROM relationships WHERE id = ?", (relationship_id,))
        self.redirect(f"/persons/{redirect_person_id}")

    def handle_inferred_relationship_delete(self) -> None:
        if self.command != "POST":
            self.redirect("/persons")
            return
        data = self.form()
        person_a_id = none_if_empty(data.get("person_a_id"))
        person_b_id = none_if_empty(data.get("person_b_id"))
        relation_type = data.get("relation_type", "frere_soeur")
        redirect_person_id = person_a_id or person_b_id or ""
        if not person_a_id or not person_b_id or person_a_id == person_b_id:
            self.redirect(f"/persons/{redirect_person_id}" if redirect_person_id else "/persons")
            return
        left_id, right_id = sorted((int(person_a_id), int(person_b_id)))
        with db_connect() as conn:
            conn.execute(
                """
                INSERT OR IGNORE INTO inferred_relationship_exclusions
                (person_a_id, person_b_id, relation_type)
                VALUES (?, ?, ?)
                """,
                (left_id, right_id, relation_type),
            )
        self.redirect(f"/persons/{redirect_person_id}")

    def handle_person_group_new(self, path: str) -> None:
        try:
            person_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            person = conn.execute("SELECT * FROM persons WHERE id = ?", (person_id,)).fetchone()
            if not person:
                self.not_found()
                return
            groups = conn.execute("SELECT id, name FROM customary_groups ORDER BY name").fetchall()
            customary_names = conn.execute("SELECT id, name FROM customary_names ORDER BY name").fetchall()
            if self.command == "POST":
                data = self.form()
                customary_name_id = none_if_empty(data.get("customary_name_id"))
                if customary_name_id:
                    conn.execute(
                        """
                        INSERT INTO person_customary_name_links
                        (person_id, customary_name_id, link_type, comment, confidentiality, certainty)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            person_id,
                            customary_name_id,
                            data.get("customary_name_link_type", "rattache_a"),
                            data.get("customary_name_comment"),
                            data.get("customary_name_confidentiality", "familial"),
                            data.get("customary_name_certainty", "a_verifier"),
                        ),
                    )
                group_id = none_if_empty(data.get("group_id"))
                if group_id:
                    conn.execute(
                        """
                        INSERT INTO person_group_links
                        (person_id, group_id, link_type, comment, confidentiality, certainty)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            person_id,
                            group_id,
                            data.get("link_type", "rattache_a"),
                            data.get("comment"),
                            data.get("confidentiality", "familial"),
                            data.get("certainty", "a_verifier"),
                        ),
                    )
                self.redirect(f"/persons/{person_id}")
                return
        fields = [
            select_field("customary_name_id", "Nom coutumier à associer", customary_names),
            select_field("customary_name_link_type", "Type de lien avec le nom coutumier *", PERSON_GROUP_LINK_TYPES, "rattache_a", True),
            select_field("customary_name_confidentiality", "Confidentialité du lien nom coutumier *", CONFIDENTIALITY, "familial", True),
            select_field("customary_name_certainty", "Certitude du lien nom coutumier *", CERTAINTY, "a_verifier", True),
            textarea_field("customary_name_comment", "Commentaire sur le lien nom coutumier", ""),
            select_field("group_id", "Groupe / lignage à associer si nécessaire", groups),
            select_field("link_type", "Type de lien avec le groupe", PERSON_GROUP_LINK_TYPES, "rattache_a"),
            select_field("confidentiality", "Confidentialité du lien groupe", CONFIDENTIALITY, "familial"),
            select_field("certainty", "Certitude du lien groupe", CERTAINTY, "a_verifier"),
            textarea_field("comment", "Commentaire sur le lien groupe", ""),
        ]
        title = f'Modifier l\'appartenance coutumière de {person["first_names"]} {person["current_name"]}'
        self.send_html("Appartenance coutumière", edit_form_panel(title, fields, f"/persons/{person_id}"))

    def handle_person_group_edit(self, path: str) -> None:
        try:
            link_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            link = conn.execute("SELECT * FROM person_group_links WHERE id = ?", (link_id,)).fetchone()
            if not link:
                self.not_found()
                return
            groups = conn.execute("SELECT id, name FROM customary_groups ORDER BY name").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE person_group_links
                    SET group_id = ?, link_type = ?, comment = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["group_id"],
                        data.get("link_type", "rattache_a"),
                        data.get("comment"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        link_id,
                    ),
                )
                self.redirect(f"/persons/{link['person_id']}")
                return
        fields = [
            select_field("group_id", "Groupe / lignage *", groups, row_value(link, "group_id"), True),
            select_field("link_type", "Type de lien *", PERSON_GROUP_LINK_TYPES, row_value(link, "link_type"), True),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(link, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(link, "certainty"), True),
            textarea_field("comment", "Commentaire", row_value(link, "comment")),
        ]
        self.send_html("Modifier le lien groupe", edit_form_panel("Modifier le lien personne-groupe", fields, f"/persons/{link['person_id']}"))

    def handle_person_customary_name_edit(self, path: str) -> None:
        try:
            link_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            link = conn.execute("SELECT * FROM person_customary_name_links WHERE id = ?", (link_id,)).fetchone()
            if not link:
                self.not_found()
                return
            customary_names = conn.execute("SELECT id, name FROM customary_names ORDER BY name").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE person_customary_name_links
                    SET customary_name_id = ?, link_type = ?, comment = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["customary_name_id"],
                        data.get("link_type", "rattache_a"),
                        data.get("comment"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        link_id,
                    ),
                )
                self.redirect(f"/persons/{link['person_id']}")
                return
        fields = [
            select_field("customary_name_id", "Nom coutumier *", customary_names, row_value(link, "customary_name_id"), True),
            select_field("link_type", "Type de lien *", PERSON_GROUP_LINK_TYPES, row_value(link, "link_type"), True),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(link, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(link, "certainty"), True),
            textarea_field("comment", "Commentaire", row_value(link, "comment")),
        ]
        self.send_html("Modifier le lien nom coutumier", edit_form_panel("Modifier le lien personne-nom coutumier", fields, f"/persons/{link['person_id']}"))

    def handle_clans(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT c.*, (SELECT COUNT(*) FROM persons p WHERE p.current_clan_id = c.id OR p.origin_clan_id = c.id) AS person_count
                FROM clans c
                ORDER BY c.name
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <div class="actions"><a class="button" href="/clans/new">Ajouter un clan</a></div>
          {rows_table(["Nom", "Variantes", "Région", "Personnes liées", "Certitude", "Confidentialité", "Action"], [
              [f'<a href="/clans/{row["id"]}">{e(row["name"])}</a>', e(row["variants"]), e(row["region"]), e(row["person_count"]), badge(row["certainty"]), badge(row["confidentiality"]), delete_button(f'/clans/{row["id"]}/delete', "Clan", f'{row["name"]} - {row["region"] or ""}')]
              for row in rows
          ])}
        </section>
        """
        self.send_html("Clans", body)

    def handle_clan_new(self) -> None:
        if self.command == "POST":
            data = self.form()
            with db_connect() as conn:
                conn.execute(
                    """
                    INSERT INTO clans (name, variants, region, description, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["name"],
                        data.get("variants"),
                        data.get("region"),
                        data.get("description"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
            self.redirect("/clans")
            return
        self.send_html("Ajouter un clan", clan_form())

    def handle_clan_edit(self, path: str) -> None:
        try:
            clan_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            clan = conn.execute("SELECT * FROM clans WHERE id = ?", (clan_id,)).fetchone()
            if not clan:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE clans
                    SET name = ?, variants = ?, region = ?, description = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["name"],
                        data.get("variants"),
                        data.get("region"),
                        data.get("description"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        clan_id,
                    ),
                )
                self.redirect(f"/clans/{clan_id}")
                return
        fields = [
            text_field("name", "Nom du clan *", row_value(clan, "name"), True),
            text_field("variants", "Variantes", row_value(clan, "variants")),
            text_field("region", "Aire / région", row_value(clan, "region")),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(clan, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(clan, "certainty"), True),
            textarea_field("description", "Description", row_value(clan, "description")),
        ]
        self.send_html("Modifier un clan", edit_form_panel("Modifier le clan", fields, f"/clans/{clan_id}"))

    def handle_clan_delete(self, path: str) -> None:
        if self.command != "POST":
            self.redirect("/clans")
            return
        try:
            clan_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            conn.execute("UPDATE persons SET current_clan_id = NULL WHERE current_clan_id = ?", (clan_id,))
            conn.execute("UPDATE persons SET origin_clan_id = NULL WHERE origin_clan_id = ?", (clan_id,))
            conn.execute("UPDATE relationships SET object_clan_id = NULL WHERE object_clan_id = ?", (clan_id,))
            conn.execute("UPDATE lands SET clan_id = NULL WHERE clan_id = ?", (clan_id,))
            conn.execute("DELETE FROM event_clans WHERE clan_id = ?", (clan_id,))
            conn.execute("DELETE FROM clans WHERE id = ?", (clan_id,))
        self.redirect("/clans")

    def handle_clan_detail(self, path: str) -> None:
        try:
            clan_id = int(path.rsplit("/", 1)[-1])
        except ValueError:
            self.not_found()
            return
        with db_connect() as conn:
            clan = conn.execute("SELECT * FROM clans WHERE id = ?", (clan_id,)).fetchone()
            if not clan:
                self.not_found()
                return
            persons = conn.execute(
                """
                SELECT id, first_names, current_name, birth_date, confidentiality, certainty
                FROM persons
                WHERE current_clan_id = ? OR origin_clan_id = ?
                ORDER BY current_name, first_names
                """,
                (clan_id, clan_id),
            ).fetchall()
            lands = conn.execute("SELECT * FROM lands WHERE clan_id = ? ORDER BY name", (clan_id,)).fetchall()
        body = f"""
        <section class="panel">
          <h2>{e(clan["name"])}</h2>
          <div class="actions"><a class="button" href="/clans/{clan_id}/edit">Modifier</a></div>
          <dl class="detail-list">
            <dt>Variantes</dt><dd>{e(clan["variants"])}</dd>
            <dt>Région</dt><dd>{e(clan["region"])}</dd>
            <dt>Certitude</dt><dd>{badge(clan["certainty"])}</dd>
            <dt>Confidentialité</dt><dd>{badge(clan["confidentiality"])}</dd>
            <dt>Description</dt><dd>{e(clan["description"])}</dd>
          </dl>
        </section>
        <section class="panel">
          <h2>Personnes liées</h2>
          {rows_table(["Nom", "Naissance", "Certitude", "Confidentialité"], [
              [f'<a href="/persons/{p["id"]}">{e(p["first_names"])} {e(p["current_name"])}</a>', e(p["birth_date"]), badge(p["certainty"]), badge(p["confidentiality"])]
              for p in persons
          ])}
        </section>
        <section class="panel">
          <h2>Terres associées</h2>
          {rows_table(["Nom", "Statut", "Confidentialité"], [
              [e(land["name"]), badge(land["status"]), badge(land["confidentiality"])]
              for land in lands
          ])}
        </section>
        """
        self.send_html("Fiche clan", body)

    def handle_groups(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT g.*, s.title AS source_title
                FROM customary_groups g
                LEFT JOIN sources s ON s.id = g.source_id
                ORDER BY g.name
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>Groupes coutumiers</h2>
          <p class="empty">Ce module sert à enregistrer prudemment les clans, lignages, branches, noms de famille ou noms cités dans les viva sans les confondre trop vite.</p>
          <div class="actions"><a class="button" href="/groups/new">Ajouter un groupe</a></div>
          {rows_table(["Nom", "Type", "Variantes", "Région", "Source", "Certitude", "Confidentialité", "Action"], [
              [f'<a href="/groups/{row["id"]}">{e(row["name"])}</a>', e(group_type_label(row["group_type"])), e(row["variants"]), e(row["region"]), e(row["source_title"]), badge(row["certainty"]), badge(row["confidentiality"]), delete_button(f'/groups/{row["id"]}/delete', "Groupe", f'{row["name"]} - {group_type_label(row["group_type"])}')]
              for row in rows
          ])}
        </section>
        """
        self.send_html("Groupes coutumiers", body)

    def handle_group_new(self) -> None:
        with db_connect() as conn:
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO customary_groups
                    (name, group_type, variants, region, description, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["name"],
                        data.get("group_type", "a_verifier"),
                        data.get("variants"),
                        data.get("region"),
                        data.get("description"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/groups")
                return
        self.send_html("Ajouter un groupe", group_form(sources))

    def handle_group_edit(self, path: str) -> None:
        try:
            group_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            group = conn.execute("SELECT * FROM customary_groups WHERE id = ?", (group_id,)).fetchone()
            if not group:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE customary_groups
                    SET name = ?, group_type = ?, variants = ?, region = ?, description = ?, source_id = ?,
                        confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["name"],
                        data.get("group_type", "a_verifier"),
                        data.get("variants"),
                        data.get("region"),
                        data.get("description"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        group_id,
                    ),
                )
                self.redirect(f"/groups/{group_id}")
                return
        fields = [
            text_field("name", "Nom *", row_value(group, "name"), True),
            select_field("group_type", "Type *", GROUP_TYPES, row_value(group, "group_type"), True),
            text_field("variants", "Variantes", row_value(group, "variants")),
            text_field("region", "Région / aire", row_value(group, "region")),
            select_field("source_id", "Source", sources, row_value(group, "source_id")),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(group, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(group, "certainty"), True),
            textarea_field("description", "Description", row_value(group, "description")),
        ]
        self.send_html("Modifier un groupe", edit_form_panel("Modifier le groupe", fields, f"/groups/{group_id}"))

    def handle_group_delete(self, path: str) -> None:
        if self.command != "POST":
            self.redirect("/groups")
            return
        try:
            group_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            conn.execute("DELETE FROM person_group_links WHERE group_id = ?", (group_id,))
            conn.execute("DELETE FROM group_relations WHERE group_a_id = ? OR group_b_id = ?", (group_id, group_id))
            conn.execute("UPDATE viva_entries SET group_a_id = NULL WHERE group_a_id = ?", (group_id,))
            conn.execute("UPDATE viva_entries SET group_b_id = NULL WHERE group_b_id = ?", (group_id,))
            conn.execute("UPDATE lineage_genealogies SET group_id = NULL WHERE group_id = ?", (group_id,))
            conn.execute("UPDATE customary_functions SET group_id = NULL WHERE group_id = ?", (group_id,))
            conn.execute("DELETE FROM research_items WHERE subject_type = 'group' AND subject_id = ?", (group_id,))
            conn.execute("DELETE FROM customary_groups WHERE id = ?", (group_id,))
        self.redirect("/groups")

    def handle_group_detail(self, path: str) -> None:
        try:
            group_id = int(path.rsplit("/", 1)[-1])
        except ValueError:
            self.not_found()
            return
        with db_connect() as conn:
            group = conn.execute(
                """
                SELECT g.*, s.title AS source_title
                FROM customary_groups g
                LEFT JOIN sources s ON s.id = g.source_id
                WHERE g.id = ?
                """,
                (group_id,),
            ).fetchone()
            if not group:
                self.not_found()
                return
            relations = conn.execute(
                """
                SELECT gr.*, b.name AS other_group, s.title AS source_title
                FROM group_relations gr
                LEFT JOIN customary_groups b ON b.id = gr.group_b_id
                LEFT JOIN sources s ON s.id = gr.source_id
                WHERE gr.group_a_id = ?
                ORDER BY gr.relation_type, b.name
                """,
                (group_id,),
            ).fetchall()
            viva_entries = conn.execute(
                """
                SELECT ve.*, vl.title AS viva_title
                FROM viva_entries ve
                LEFT JOIN viva_lists vl ON vl.id = ve.viva_list_id
                WHERE ve.group_a_id = ? OR ve.group_b_id = ?
                ORDER BY vl.title, ve.position
                """,
                (group_id, group_id),
            ).fetchall()
            research_items = conn.execute(
                """
                SELECT ri.*, s.title AS source_title
                FROM research_items ri
                LEFT JOIN sources s ON s.id = ri.source_id
                WHERE ri.subject_type = 'group' AND ri.subject_id = ?
                ORDER BY ri.created_at DESC
                """,
                (group_id,),
            ).fetchall()
            lineage_genealogies = conn.execute(
                """
                SELECT lg.*, s.title AS source_title
                FROM lineage_genealogies lg
                LEFT JOIN sources s ON s.id = lg.source_id
                WHERE lg.group_id = ?
                ORDER BY lg.created_at DESC
                """,
                (group_id,),
            ).fetchall()
            customary_functions = conn.execute(
                """
                SELECT cf.*, s.title AS source_title
                FROM customary_functions cf
                LEFT JOIN sources s ON s.id = cf.source_id
                WHERE cf.group_id = ?
                ORDER BY cf.created_at DESC
                """,
                (group_id,),
            ).fetchall()
            linked_persons = conn.execute(
                """
                SELECT pgl.*, p.first_names, p.current_name, p.birth_date
                FROM person_group_links pgl
                JOIN persons p ON p.id = pgl.person_id
                WHERE pgl.group_id = ?
                ORDER BY p.current_name, p.first_names
                """,
                (group_id,),
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>{e(group["name"])}</h2>
          <div class="actions"><a class="button" href="/groups/{group_id}/edit">Modifier</a> <a class="button" href="/groups/{group_id}/map">Voir la carte</a></div>
          <dl class="detail-list">
            <dt>Type</dt><dd>{e(group_type_label(group["group_type"]))}</dd>
            <dt>Variantes</dt><dd>{e(group["variants"])}</dd>
            <dt>Région</dt><dd>{e(group["region"])}</dd>
            <dt>Source</dt><dd>{e(group["source_title"])}</dd>
            <dt>Certitude</dt><dd>{badge(group["certainty"])}</dd>
            <dt>Confidentialité</dt><dd>{badge(group["confidentiality"])}</dd>
            <dt>Description</dt><dd>{e(group["description"])}</dd>
          </dl>
        </section>
        <section class="panel">
          <h2>Personnes associées</h2>
          {rows_table(["Personne", "Naissance", "Lien", "Certitude", "Commentaire"], [
              [f'<a href="/persons/{p["person_id"]}">{e(p["first_names"])} {e(p["current_name"])}</a>', e(p["birth_date"]), e(person_group_link_label(p["link_type"])), badge(p["certainty"]), e(p["comment"])]
              for p in linked_persons
          ])}
        </section>
        <section class="panel">
          <h2>Relations entre groupes</h2>
          {rows_table(["Relation", "Avec", "Source", "Certitude", "Commentaire"], [
              [e(group_relation_label(r["relation_type"])), e(r["other_group"]), e(r["source_title"]), badge(r["certainty"]), e(r["context"])]
              for r in relations
          ])}
        </section>
        <section class="panel">
          <h2>Présence dans les viva</h2>
          {rows_table(["Liste", "Position", "Section", "Texte", "Traduction", "Note"], [
              [e(v["viva_title"]), e(v["position"]), e(v["section_name"]), e(v["raw_text"]), e(v["translation"]), e(v["note"])]
              for v in viva_entries
          ])}
        </section>
        <section class="panel">
          <h2>Informations et hypotheses</h2>
          {rows_table(["Titre", "Auteur / vue", "Information", "Interpretation", "Source", "Certitude"], [
              [e(r["title"]), e(r["author_view"]), e(r["statement"]), e(r["interpretation"]), e(r["source_title"]), badge(r["certainty"])]
              for r in research_items
          ])}
        </section>
        <section class="panel">
          <h2>Genealogies de lignage</h2>
          {rows_table(["Titre", "Auteur / collecteur", "Chaine", "Interpretation", "Source", "Certitude"], [
              [e(lg["title"]), e(lg["author_or_collector"]), e(lg["chain_text"]), e(lg["interpretation"]), e(lg["source_title"]), badge(lg["certainty"])]
              for lg in lineage_genealogies
          ])}
        </section>
        <section class="panel">
          <h2>Fonctions coutumieres</h2>
          {rows_table(["Titre", "Type", "Lieu", "Description", "Source", "Certitude"], [
              [e(cf["title"]), e(function_type_label(cf["function_type"])), e(cf["place"]), e(cf["description"]), e(cf["source_title"]), badge(cf["certainty"])]
              for cf in customary_functions
          ])}
        </section>
        """
        self.send_html("Fiche groupe", body)

    def handle_group_map(self, path: str) -> None:
        try:
            group_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            group = conn.execute(
                "SELECT * FROM customary_groups WHERE id = ?",
                (group_id,),
            ).fetchone()
            if not group:
                self.not_found()
                return
            relations = conn.execute(
                """
                SELECT gr.*, a.name AS group_a, b.name AS group_b
                FROM group_relations gr
                JOIN customary_groups a ON a.id = gr.group_a_id
                JOIN customary_groups b ON b.id = gr.group_b_id
                WHERE gr.group_a_id = ? OR gr.group_b_id = ?
                ORDER BY gr.relation_type
                """,
                (group_id, group_id),
            ).fetchall()
            viva_links = conn.execute(
                """
                SELECT ve.raw_text, ve.certainty, a.name AS group_a, b.name AS group_b
                FROM viva_entries ve
                LEFT JOIN customary_groups a ON a.id = ve.group_a_id
                LEFT JOIN customary_groups b ON b.id = ve.group_b_id
                WHERE ve.group_a_id = ? OR ve.group_b_id = ?
                ORDER BY ve.position
                """,
                (group_id, group_id),
            ).fetchall()
            functions = conn.execute(
                """
                SELECT title, function_type, place, certainty
                FROM customary_functions
                WHERE group_id = ?
                ORDER BY function_type, title
                """,
                (group_id,),
            ).fetchall()
            linked_persons = conn.execute(
                """
                SELECT p.first_names || ' ' || p.current_name AS person_name, pgl.link_type, pgl.certainty
                FROM person_group_links pgl
                JOIN persons p ON p.id = pgl.person_id
                WHERE pgl.group_id = ?
                ORDER BY p.current_name, p.first_names
                """,
                (group_id,),
            ).fetchall()
        svg = render_group_map(group, relations, viva_links, functions, linked_persons)
        body = f"""
        <section class="panel">
          <div class="actions"><a class="button secondary" href="/groups/{group_id}">Retour fiche groupe</a></div>
          <h2>Carte autour de {e(group["name"])}</h2>
          <p class="empty">Carte relationnelle : groupes liés, appariements viva et fonctions ou lieux associés.</p>
          {svg}
        </section>
        <section class="panel">
          <h2>Lecture</h2>
          <p class="empty">Cette carte n'est pas un arbre familial. Elle sert à visualiser les relations entre noms, lignages, lieux, fonctions et sources.</p>
        </section>
        """
        self.send_html("Carte groupe", body)

    def handle_customary_names(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT cn.*, s.title AS source_title
                FROM customary_names cn
                LEFT JOIN sources s ON s.id = cn.source_id
                ORDER BY cn.name
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <p class="empty">Ce module sert à enregistrer les noms coutumiers cités notamment dans les viva, sans les réduire trop vite à un clan, un groupe ou un nom de famille.</p>
          <div class="actions"><a class="button" href="/customary-names/new">Ajouter un nom coutumier</a></div>
          {rows_table(["Nom coutumier", "Type", "Variantes", "Région", "Source", "Certitude", "Confidentialité", "Action"], [
              [f'<a href="/customary-names/{row["id"]}">{e(row["name"])}</a>', e(group_type_label(row["name_type"])), e(row["variants"]), e(row["region"]), e(row["source_title"]), badge(row["certainty"]), badge(row["confidentiality"]), delete_button(f'/customary-names/{row["id"]}/delete', "Nom coutumier", f'{row["name"]} - {group_type_label(row["name_type"])}')]
              for row in rows
          ])}
        </section>
        """
        self.send_html("Noms coutumiers", body)

    def handle_customary_name_new(self) -> None:
        with db_connect() as conn:
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO customary_names
                    (name, name_type, variants, region, description, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["name"],
                        data.get("name_type", "a_verifier"),
                        data.get("variants"),
                        data.get("region"),
                        data.get("description"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/customary-names")
                return
        self.send_html("Ajouter un nom coutumier", customary_name_form(sources))

    def handle_customary_name_edit(self, path: str) -> None:
        try:
            custom_name_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            custom_name = conn.execute("SELECT * FROM customary_names WHERE id = ?", (custom_name_id,)).fetchone()
            if not custom_name:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE customary_names
                    SET name = ?, name_type = ?, variants = ?, region = ?, description = ?, source_id = ?,
                        confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["name"],
                        data.get("name_type", "a_verifier"),
                        data.get("variants"),
                        data.get("region"),
                        data.get("description"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        custom_name_id,
                    ),
                )
                self.redirect(f"/customary-names/{custom_name_id}")
                return
        fields = [
            text_field("name", "Nom coutumier *", row_value(custom_name, "name"), True),
            select_field("name_type", "Type *", GROUP_TYPES, row_value(custom_name, "name_type"), True),
            text_field("variants", "Variantes", row_value(custom_name, "variants")),
            text_field("region", "Région / aire", row_value(custom_name, "region")),
            select_field("source_id", "Source", sources, row_value(custom_name, "source_id")),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(custom_name, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(custom_name, "certainty"), True),
            textarea_field("description", "Description", row_value(custom_name, "description")),
        ]
        self.send_html("Modifier un nom coutumier", edit_form_panel("Modifier le nom coutumier", fields, f"/customary-names/{custom_name_id}"))

    def handle_customary_name_delete(self, path: str) -> None:
        if self.command != "POST":
            self.redirect("/customary-names")
            return
        try:
            custom_name_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            conn.execute("DELETE FROM person_customary_name_links WHERE customary_name_id = ?", (custom_name_id,))
            conn.execute("DELETE FROM customary_name_relations WHERE name_a_id = ? OR name_b_id = ?", (custom_name_id, custom_name_id))
            conn.execute("DELETE FROM customary_names WHERE id = ?", (custom_name_id,))
        self.redirect("/customary-names")

    def handle_customary_name_detail(self, path: str) -> None:
        try:
            custom_name_id = int(path.rsplit("/", 1)[-1])
        except ValueError:
            self.not_found()
            return
        with db_connect() as conn:
            custom_name = conn.execute(
                """
                SELECT cn.*, s.title AS source_title
                FROM customary_names cn
                LEFT JOIN sources s ON s.id = cn.source_id
                WHERE cn.id = ?
                """,
                (custom_name_id,),
            ).fetchone()
            if not custom_name:
                self.not_found()
                return
            relations = conn.execute(
                """
                SELECT cnr.*, a.name AS name_a, b.name AS name_b, s.title AS source_title
                FROM customary_name_relations cnr
                JOIN customary_names a ON a.id = cnr.name_a_id
                JOIN customary_names b ON b.id = cnr.name_b_id
                LEFT JOIN sources s ON s.id = cnr.source_id
                WHERE cnr.name_a_id = ? OR cnr.name_b_id = ?
                ORDER BY cnr.relation_type, b.name, a.name
                """,
                (custom_name_id, custom_name_id),
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>{e(custom_name["name"])}</h2>
          <div class="actions"><a class="button" href="/customary-names/{custom_name_id}/edit">Modifier</a> <a class="button" href="/customary-names/{custom_name_id}/relations/new">Associer un allié</a></div>
          <dl class="detail-list">
            <dt>Type</dt><dd>{e(group_type_label(custom_name["name_type"]))}</dd>
            <dt>Variantes</dt><dd>{e(custom_name["variants"])}</dd>
            <dt>Région</dt><dd>{e(custom_name["region"])}</dd>
            <dt>Source</dt><dd>{e(custom_name["source_title"])}</dd>
            <dt>Certitude</dt><dd>{badge(custom_name["certainty"])}</dd>
            <dt>Confidentialité</dt><dd>{badge(custom_name["confidentiality"])}</dd>
            <dt>Description</dt><dd>{e(custom_name["description"])}</dd>
          </dl>
        </section>
        <section class="panel">
          <h2>Noms coutumiers alliés / associés</h2>
          {rows_table(["Nom allié", "Relation", "Source", "Certitude", "Confidentialité", "Commentaire"], [
              [f'<a href="/customary-names/{r["name_b_id"] if r["name_a_id"] == custom_name_id else r["name_a_id"]}">{e(r["name_b"] if r["name_a_id"] == custom_name_id else r["name_a"])}</a>', e(group_relation_label(r["relation_type"])), e(r["source_title"]), badge(r["certainty"]), badge(r["confidentiality"]), e(r["context"])]
              for r in relations
          ])}
        </section>
        """
        self.send_html("Nom coutumier", body)

    def handle_customary_name_relation_new(self, path: str) -> None:
        try:
            custom_name_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            custom_name = conn.execute("SELECT * FROM customary_names WHERE id = ?", (custom_name_id,)).fetchone()
            if not custom_name:
                self.not_found()
                return
            other_names = conn.execute(
                "SELECT id, name FROM customary_names WHERE id != ? ORDER BY name",
                (custom_name_id,),
            ).fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                other_id = none_if_empty(data.get("name_b_id"))
                if not other_id:
                    fields = customary_name_relation_fields(custom_name, other_names, sources, data)
                    self.send_html("Associer un allié", edit_form_panel(f'Associer un allié à {custom_name["name"]}', fields, f"/customary-names/{custom_name_id}"))
                    return
                conn.execute(
                    """
                    INSERT INTO customary_name_relations
                    (name_a_id, name_b_id, relation_type, context, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        custom_name_id,
                        other_id,
                        data.get("relation_type", "allie_a"),
                        data.get("context"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect(f"/customary-names/{custom_name_id}")
                return
        fields = customary_name_relation_fields(custom_name, other_names, sources)
        self.send_html("Associer un allié", edit_form_panel(f'Associer un allié à {custom_name["name"]}', fields, f"/customary-names/{custom_name_id}"))

    def handle_viva(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT vl.*, s.title AS source_title,
                       (SELECT COUNT(*) FROM viva_entries ve WHERE ve.viva_list_id = vl.id) AS entry_count
                FROM viva_lists vl
                LEFT JOIN sources s ON s.id = vl.source_id
                ORDER BY vl.title
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>Listes viva</h2>
          <p class="empty">Une liste viva archive une récitation ou un extrait structuré : ordre, appariements, regroupements et source.</p>
          <div class="actions"><a class="button" href="/viva/new">Ajouter une liste viva</a></div>
          {rows_table(["Titre", "Aire", "Auteur / collecteur", "Date", "Entrées", "Source", "Certitude", "Action"], [
              [f'<a href="/viva/{row["id"]}">{e(row["title"])}</a>', e(row["area"]), e(row["collector_author"]), e(row["collection_date"]), e(row["entry_count"]), e(row["source_title"]), badge(row["certainty"]), delete_button(f'/viva/{row["id"]}/delete', "Liste viva", f'{row["title"]} - {row["area"] or ""}')]
              for row in rows
          ])}
        </section>
        """
        self.send_html("Viva", body)

    def handle_viva_new(self) -> None:
        with db_connect() as conn:
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO viva_lists
                    (title, area, collector_author, collection_date, description, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["title"],
                        data.get("area"),
                        data.get("collector_author"),
                        data.get("collection_date"),
                        data.get("description"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/viva")
                return
        self.send_html("Ajouter une liste viva", viva_form(sources))

    def handle_viva_edit(self, path: str) -> None:
        try:
            viva_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            viva = conn.execute("SELECT * FROM viva_lists WHERE id = ?", (viva_id,)).fetchone()
            if not viva:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE viva_lists
                    SET title = ?, area = ?, collector_author = ?, collection_date = ?, description = ?,
                        source_id = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["title"],
                        data.get("area"),
                        data.get("collector_author"),
                        data.get("collection_date"),
                        data.get("description"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        viva_id,
                    ),
                )
                self.redirect(f"/viva/{viva_id}")
                return
        fields = [
            text_field("title", "Titre *", row_value(viva, "title"), True),
            text_field("area", "Aire / région", row_value(viva, "area")),
            text_field("collector_author", "Auteur ou collecteur", row_value(viva, "collector_author")),
            text_field("collection_date", "Date ou période", row_value(viva, "collection_date")),
            select_field("source_id", "Source", sources, row_value(viva, "source_id")),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(viva, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(viva, "certainty"), True),
            textarea_field("description", "Description", row_value(viva, "description")),
        ]
        self.send_html("Modifier une liste viva", edit_form_panel("Modifier la liste viva", fields, f"/viva/{viva_id}"))

    def handle_viva_delete(self, path: str) -> None:
        if self.command != "POST":
            self.redirect("/viva")
            return
        try:
            viva_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            conn.execute("DELETE FROM viva_entries WHERE viva_list_id = ?", (viva_id,))
            conn.execute("DELETE FROM viva_lists WHERE id = ?", (viva_id,))
        self.redirect("/viva")

    def handle_viva_detail(self, path: str) -> None:
        try:
            viva_id = int(path.rsplit("/", 1)[-1])
        except ValueError:
            self.not_found()
            return
        with db_connect() as conn:
            viva = conn.execute(
                """
                SELECT vl.*, s.title AS source_title
                FROM viva_lists vl
                LEFT JOIN sources s ON s.id = vl.source_id
                WHERE vl.id = ?
                """,
                (viva_id,),
            ).fetchone()
            if not viva:
                self.not_found()
                return
            entries = conn.execute(
                """
                SELECT ve.*, a.name AS group_a, b.name AS group_b
                FROM viva_entries ve
                LEFT JOIN customary_groups a ON a.id = ve.group_a_id
                LEFT JOIN customary_groups b ON b.id = ve.group_b_id
                WHERE ve.viva_list_id = ?
                ORDER BY ve.position, ve.id
                """,
                (viva_id,),
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>{e(viva["title"])}</h2>
          <div class="actions"><a class="button" href="/viva/{viva_id}/edit">Modifier</a></div>
          <dl class="detail-list">
            <dt>Aire</dt><dd>{e(viva["area"])}</dd>
            <dt>Auteur / collecteur</dt><dd>{e(viva["collector_author"])}</dd>
            <dt>Date</dt><dd>{e(viva["collection_date"])}</dd>
            <dt>Source</dt><dd>{e(viva["source_title"])}</dd>
            <dt>Certitude</dt><dd>{badge(viva["certainty"])}</dd>
            <dt>Confidentialité</dt><dd>{badge(viva["confidentiality"])}</dd>
            <dt>Description</dt><dd>{e(viva["description"])}</dd>
          </dl>
        </section>
        <section class="panel">
          <h2>Entrées</h2>
          {rows_table(["Position", "Section", "Texte", "Groupe A", "Groupe B", "Traduction", "Note"], [
              [e(v["position"]), e(v["section_name"]), e(v["raw_text"]), e(v["group_a"]), e(v["group_b"]), e(v["translation"]), e(v["note"])]
              for v in entries
          ])}
        </section>
        """
        self.send_html("Fiche viva", body)

    def handle_research(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT ri.*, s.title AS source_title, g.name AS group_name
                FROM research_items ri
                LEFT JOIN sources s ON s.id = ri.source_id
                LEFT JOIN customary_groups g ON ri.subject_type = 'group' AND g.id = ri.subject_id
                ORDER BY ri.created_at DESC
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>Informations et hypothèses</h2>
          <p class="empty">Ce module sépare l'information relevée, l'interprétation et la source. Il sert à conserver plusieurs versions sans trancher trop vite.</p>
          <div class="actions"><a class="button" href="/research/new">Ajouter une information</a></div>
          {rows_table(["Titre", "Sujet", "Auteur / vue", "Preuve", "Source", "Certitude", "Action"], [
              [e(row["title"]), e(row["group_name"] or subject_type_label(row["subject_type"])), e(row["author_view"]), e(evidence_label(row["evidence_level"])), e(row["source_title"]), badge(row["certainty"]), f'<a class="button secondary" href="/research/{row["id"]}/edit">Modifier</a>']
              for row in rows
          ])}
        </section>
        """
        self.send_html("Infos", body)

    def handle_research_new(self) -> None:
        with db_connect() as conn:
            groups = conn.execute("SELECT id, name FROM customary_groups ORDER BY name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                subject_type = data.get("subject_type", "general")
                subject_id = none_if_empty(data.get("subject_id")) if subject_type == "group" else None
                conn.execute(
                    """
                    INSERT INTO research_items
                    (subject_type, subject_id, title, statement, interpretation, author_view, source_id, evidence_level, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        subject_type,
                        subject_id,
                        data["title"],
                        data["statement"],
                        data.get("interpretation"),
                        data.get("author_view"),
                        none_if_empty(data.get("source_id")),
                        data.get("evidence_level", "document_ecrit"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/research")
                return
        self.send_html("Ajouter une information", research_form(groups, sources))

    def handle_research_edit(self, path: str) -> None:
        try:
            item_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            groups = conn.execute("SELECT id, name FROM customary_groups ORDER BY name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            item = conn.execute("SELECT * FROM research_items WHERE id = ?", (item_id,)).fetchone()
            if not item:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                subject_type = data.get("subject_type", "general")
                subject_id = none_if_empty(data.get("subject_id")) if subject_type == "group" else None
                conn.execute(
                    """
                    UPDATE research_items
                    SET subject_type = ?, subject_id = ?, title = ?, statement = ?, interpretation = ?,
                        author_view = ?, source_id = ?, evidence_level = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        subject_type,
                        subject_id,
                        data["title"],
                        data["statement"],
                        data.get("interpretation"),
                        data.get("author_view"),
                        none_if_empty(data.get("source_id")),
                        data.get("evidence_level", "document_ecrit"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        item_id,
                    ),
                )
                self.redirect("/research")
                return
        fields = [
            select_field("subject_type", "Type de sujet *", SUBJECT_TYPES, row_value(item, "subject_type"), True),
            select_field("subject_id", "Groupe concerné", groups, row_value(item, "subject_id")),
            text_field("title", "Titre *", row_value(item, "title"), True),
            text_field("author_view", "Auteur / version", row_value(item, "author_view")),
            select_field("source_id", "Source", sources, row_value(item, "source_id")),
            select_field("evidence_level", "Niveau de preuve *", EVIDENCE_LEVELS, row_value(item, "evidence_level"), True),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(item, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(item, "certainty"), True),
            textarea_field("statement", "Information relevée *", row_value(item, "statement"), True),
            textarea_field("interpretation", "Interprétation / prudence", row_value(item, "interpretation")),
        ]
        self.send_html("Modifier une information", edit_form_panel("Modifier l'information", fields, "/research"))

    def handle_lineages(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT lg.*, g.name AS group_name, s.title AS source_title
                FROM lineage_genealogies lg
                LEFT JOIN customary_groups g ON g.id = lg.group_id
                LEFT JOIN sources s ON s.id = lg.source_id
                ORDER BY lg.created_at DESC
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>Généalogies de lignage</h2>
          <p class="empty">Ce module archive des chaînes de noms ou généalogies de lignage, distinctes des fiches de personnes vivantes.</p>
          <div class="actions"><a class="button" href="/lineages/new">Ajouter une généalogie</a></div>
          {rows_table(["Titre", "Groupe", "Auteur / collecteur", "Chaîne", "Source", "Certitude", "Action"], [
              [e(row["title"]), e(row["group_name"]), e(row["author_or_collector"]), e(row["chain_text"]), e(row["source_title"]), badge(row["certainty"]), f'<a class="button secondary" href="/lineages/{row["id"]}/edit">Modifier</a>']
              for row in rows
          ])}
        </section>
        """
        self.send_html("Généalogies", body)

    def handle_lineage_new(self) -> None:
        with db_connect() as conn:
            groups = conn.execute("SELECT id, name FROM customary_groups ORDER BY name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO lineage_genealogies
                    (group_id, title, author_or_collector, chain_text, interpretation, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        none_if_empty(data.get("group_id")),
                        data["title"],
                        data.get("author_or_collector"),
                        data["chain_text"],
                        data.get("interpretation"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/lineages")
                return
        self.send_html("Ajouter une généalogie", lineage_form(groups, sources))

    def handle_lineage_edit(self, path: str) -> None:
        try:
            lineage_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            groups = conn.execute("SELECT id, name FROM customary_groups ORDER BY name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            lineage = conn.execute("SELECT * FROM lineage_genealogies WHERE id = ?", (lineage_id,)).fetchone()
            if not lineage:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE lineage_genealogies
                    SET group_id = ?, title = ?, author_or_collector = ?, chain_text = ?, interpretation = ?,
                        source_id = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        none_if_empty(data.get("group_id")),
                        data["title"],
                        data.get("author_or_collector"),
                        data["chain_text"],
                        data.get("interpretation"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        lineage_id,
                    ),
                )
                self.redirect("/lineages")
                return
        fields = [
            select_field("group_id", "Groupe / lignage", groups, row_value(lineage, "group_id")),
            text_field("title", "Titre *", row_value(lineage, "title"), True),
            text_field("author_or_collector", "Auteur / collecteur", row_value(lineage, "author_or_collector")),
            select_field("source_id", "Source", sources, row_value(lineage, "source_id")),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(lineage, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(lineage, "certainty"), True),
            textarea_field("chain_text", "Chaîne de noms *", row_value(lineage, "chain_text"), True),
            textarea_field("interpretation", "Interprétation", row_value(lineage, "interpretation")),
        ]
        self.send_html("Modifier une généalogie", edit_form_panel("Modifier la généalogie", fields, "/lineages"))

    def handle_functions(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT cf.*, g.name AS group_name, p.first_names || ' ' || p.current_name AS person_name, s.title AS source_title
                FROM customary_functions cf
                LEFT JOIN customary_groups g ON g.id = cf.group_id
                LEFT JOIN persons p ON p.id = cf.person_id
                LEFT JOIN sources s ON s.id = cf.source_id
                ORDER BY cf.created_at DESC
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <h2>Fonctions coutumières</h2>
          <p class="empty">Ce module sert à noter des rôles : garde de lieu, maître d'un élément, guerrier, serviteur, porte-parole, maître des ignames, etc.</p>
          <div class="actions"><a class="button" href="/functions/new">Ajouter une fonction</a></div>
          {rows_table(["Titre", "Type", "Groupe", "Personne", "Lieu", "Source", "Certitude", "Action"], [
              [e(row["title"]), e(function_type_label(row["function_type"])), e(row["group_name"]), e(row["person_name"]), e(row["place"]), e(row["source_title"]), badge(row["certainty"]), f'<a class="button secondary" href="/functions/{row["id"]}/edit">Modifier</a>']
              for row in rows
          ])}
        </section>
        """
        self.send_html("Fonctions", body)

    def handle_function_new(self) -> None:
        with db_connect() as conn:
            groups = conn.execute("SELECT id, name FROM customary_groups ORDER BY name").fetchall()
            persons = conn.execute("SELECT id, first_names || ' ' || current_name AS label FROM persons ORDER BY current_name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO customary_functions
                    (group_id, person_id, function_type, title, place, description, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        none_if_empty(data.get("group_id")),
                        none_if_empty(data.get("person_id")),
                        data.get("function_type", "autre"),
                        data["title"],
                        data.get("place"),
                        data["description"],
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/functions")
                return
        self.send_html("Ajouter une fonction", function_form(groups, persons, sources))

    def handle_function_edit(self, path: str) -> None:
        try:
            function_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            groups = conn.execute("SELECT id, name FROM customary_groups ORDER BY name").fetchall()
            persons = conn.execute("SELECT id, first_names || ' ' || current_name AS label FROM persons ORDER BY current_name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            function = conn.execute("SELECT * FROM customary_functions WHERE id = ?", (function_id,)).fetchone()
            if not function:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE customary_functions
                    SET group_id = ?, person_id = ?, function_type = ?, title = ?, place = ?, description = ?,
                        source_id = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        none_if_empty(data.get("group_id")),
                        none_if_empty(data.get("person_id")),
                        data.get("function_type", "autre"),
                        data["title"],
                        data.get("place"),
                        data["description"],
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        function_id,
                    ),
                )
                self.redirect("/functions")
                return
        fields = [
            select_field("group_id", "Groupe concerné", groups, row_value(function, "group_id")),
            select_field("person_id", "Personne concernée", persons, row_value(function, "person_id")),
            select_field("function_type", "Type de fonction *", FUNCTION_TYPES, row_value(function, "function_type"), True),
            text_field("title", "Titre *", row_value(function, "title"), True),
            text_field("place", "Lieu", row_value(function, "place")),
            select_field("source_id", "Source", sources, row_value(function, "source_id")),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(function, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(function, "certainty"), True),
            textarea_field("description", "Description *", row_value(function, "description"), True),
        ]
        self.send_html("Modifier une fonction", edit_form_panel("Modifier la fonction", fields, "/functions"))

    def handle_events(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT ev.*, p.first_names || ' ' || p.current_name AS person_name,
                       (
                         SELECT GROUP_CONCAT(pp.first_names || ' ' || pp.current_name || ' (' || epl.role || ')', ', ')
                         FROM event_person_links epl
                         JOIN persons pp ON pp.id = epl.person_id
                         WHERE epl.event_id = ev.id
                       ) AS linked_people
                FROM customary_events ev
                LEFT JOIN persons p ON p.id = ev.main_person_id
                ORDER BY ev.created_at DESC
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <div class="actions"><a class="button" href="/events/new">Ajouter un événement</a></div>
          {rows_table(["Date", "Type", "Titre", "Personne principale", "Personnes associées", "Certitude", "Confidentialité", "Action"], [
              [e(row["event_date"]), e(event_label(row["event_type"])), e(row["title"]), e(row["person_name"]), e(row["linked_people"]), badge(row["certainty"]), badge(row["confidentiality"]), f'<a class="button secondary" href="/events/{row["id"]}/edit">Modifier</a> <a class="button secondary" href="/events/{row["id"]}/people/new">Associer personne</a> {delete_button(f"/events/{row["id"]}/delete", "Événement", f"{row["title"]} - {event_label(row["event_type"])}")}']
              for row in rows
          ])}
        </section>
        """
        self.send_html("Événements", body)

    def handle_event_new(self) -> None:
        with db_connect() as conn:
            persons = conn.execute("SELECT id, first_names || ' ' || current_name AS label FROM persons ORDER BY current_name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO customary_events
                    (event_type, title, event_date, place, main_person_id, description, effects, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["event_type"],
                        data["title"],
                        data.get("event_date"),
                        data.get("place"),
                        none_if_empty(data.get("main_person_id")),
                        data["description"],
                        data.get("effects"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/events")
                return
        self.send_html("Ajouter un événement", event_form(persons, sources))

    def handle_event_edit(self, path: str) -> None:
        try:
            event_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            persons = conn.execute("SELECT id, first_names || ' ' || current_name AS label FROM persons ORDER BY current_name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            event = conn.execute("SELECT * FROM customary_events WHERE id = ?", (event_id,)).fetchone()
            if not event:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE customary_events
                    SET event_type = ?, title = ?, event_date = ?, place = ?, main_person_id = ?,
                        description = ?, effects = ?, source_id = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["event_type"],
                        data["title"],
                        data.get("event_date"),
                        data.get("place"),
                        none_if_empty(data.get("main_person_id")),
                        data["description"],
                        data.get("effects"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        event_id,
                    ),
                )
                self.redirect("/events")
                return
        fields = [
            select_field("event_type", "Type *", EVENT_TYPES, row_value(event, "event_type"), True),
            text_field("title", "Titre *", row_value(event, "title"), True),
            text_field("event_date", "Date ou période", row_value(event, "event_date")),
            text_field("place", "Lieu", row_value(event, "place")),
            select_field("main_person_id", "Personne principale", persons, row_value(event, "main_person_id")),
            select_field("source_id", "Source", sources, row_value(event, "source_id")),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(event, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(event, "certainty"), True),
            textarea_field("description", "Description *", row_value(event, "description"), True),
            textarea_field("effects", "Effets", row_value(event, "effects")),
        ]
        self.send_html("Modifier un événement", edit_form_panel("Modifier l'événement", fields, "/events"))

    def handle_event_delete(self, path: str) -> None:
        if self.command != "POST":
            self.redirect("/events")
            return
        try:
            event_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            conn.execute("DELETE FROM event_person_links WHERE event_id = ?", (event_id,))
            conn.execute("DELETE FROM event_clans WHERE event_id = ?", (event_id,))
            conn.execute("DELETE FROM customary_events WHERE id = ?", (event_id,))
        self.redirect("/events")

    def handle_event_person_new(self, path: str) -> None:
        try:
            event_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            event = conn.execute("SELECT * FROM customary_events WHERE id = ?", (event_id,)).fetchone()
            if not event:
                self.not_found()
                return
            persons = conn.execute("SELECT id, first_names || ' ' || current_name AS label FROM persons ORDER BY current_name").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO event_person_links
                    (event_id, person_id, role, comment, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        event_id,
                        data["person_id"],
                        data["role"],
                        data.get("comment"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                    ),
                )
                self.redirect("/events")
                return
        fields = [
            select_field("person_id", "Personne *", persons, None, True),
            select_field("role", "Rôle dans l'événement *", EVENT_PERSON_ROLES, "epoux" if event["event_type"] == "mariage" else "participant", True),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, "familial", True),
            select_field("certainty", "Certitude *", CERTAINTY, "a_verifier", True),
            textarea_field("comment", "Commentaire", ""),
        ]
        self.send_html("Associer une personne", edit_form_panel(f'Associer une personne à : {event["title"]}', fields, "/events"))

    def handle_event_person_edit(self, path: str) -> None:
        try:
            link_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            link = conn.execute("SELECT * FROM event_person_links WHERE id = ?", (link_id,)).fetchone()
            if not link:
                self.not_found()
                return
            persons = conn.execute("SELECT id, first_names || ' ' || current_name AS label FROM persons ORDER BY current_name").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE event_person_links
                    SET person_id = ?, role = ?, comment = ?, confidentiality = ?, certainty = ?
                    WHERE id = ?
                    """,
                    (
                        data["person_id"],
                        data["role"],
                        data.get("comment"),
                        data.get("confidentiality", "familial"),
                        data.get("certainty", "a_verifier"),
                        link_id,
                    ),
                )
                self.redirect("/events")
                return
        fields = [
            select_field("person_id", "Personne *", persons, row_value(link, "person_id"), True),
            select_field("role", "Rôle dans l'événement *", EVENT_PERSON_ROLES, row_value(link, "role"), True),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(link, "confidentiality"), True),
            select_field("certainty", "Certitude *", CERTAINTY, row_value(link, "certainty"), True),
            textarea_field("comment", "Commentaire", row_value(link, "comment")),
        ]
        self.send_html("Modifier un rôle événement", edit_form_panel("Modifier le rôle dans l'événement", fields, "/events"))

    def handle_sources(self) -> None:
        with db_connect() as conn:
            rows = conn.execute("SELECT * FROM sources ORDER BY created_at DESC").fetchall()
        body = f"""
        <section class="panel">
          <div class="actions"><a class="button" href="/sources/new">Ajouter une source</a></div>
          {rows_table(["Titre", "Type", "Témoin", "Date recueil", "Fiabilité", "Consentement", "Confidentialité"], [
              [e(row["title"]), e(source_label(row["source_type"])), e(row["witness_name"]), e(row["collection_date"]), badge(row["reliability"]), e(row["consent"]), badge(row["confidentiality"]), f'<a class="button secondary" href="/sources/{row["id"]}/edit">Modifier</a>']
              for row in rows
          ])}
        </section>
        """
        self.send_html("Sources", body)

    def handle_source_new(self) -> None:
        if self.command == "POST":
            data = self.form()
            with db_connect() as conn:
                conn.execute(
                    """
                    INSERT INTO sources
                    (source_type, title, witness_name, collection_date, collected_by, summary, consent, confidentiality, reliability)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["source_type"],
                        data["title"],
                        data.get("witness_name"),
                        data.get("collection_date"),
                        data.get("collected_by"),
                        data["summary"],
                        data.get("consent", "a_confirmer"),
                        data.get("confidentiality", "familial"),
                        data.get("reliability", "a_verifier"),
                    ),
                )
            self.redirect("/sources")
            return
        self.send_html("Ajouter une source", source_form())

    def handle_source_edit(self, path: str) -> None:
        try:
            source_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            source = conn.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE sources
                    SET source_type = ?, title = ?, witness_name = ?, collection_date = ?, collected_by = ?,
                        summary = ?, consent = ?, confidentiality = ?, reliability = ?
                    WHERE id = ?
                    """,
                    (
                        data["source_type"],
                        data["title"],
                        data.get("witness_name"),
                        data.get("collection_date"),
                        data.get("collected_by"),
                        data["summary"],
                        data.get("consent", "a_confirmer"),
                        data.get("confidentiality", "familial"),
                        data.get("reliability", "a_verifier"),
                        source_id,
                    ),
                )
                self.redirect("/sources")
                return
        fields = [
            select_field("source_type", "Type *", SOURCE_TYPES, row_value(source, "source_type"), True),
            text_field("title", "Titre *", row_value(source, "title"), True),
            text_field("witness_name", "Témoin / transmetteur", row_value(source, "witness_name")),
            text_field("collection_date", "Date de recueil", row_value(source, "collection_date")),
            text_field("collected_by", "Recueilli par", row_value(source, "collected_by")),
            select_field("consent", "Consentement *", CONSENT, row_value(source, "consent"), True),
            select_field("reliability", "Fiabilité *", CERTAINTY, row_value(source, "reliability"), True),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(source, "confidentiality"), True),
            textarea_field("summary", "Résumé *", row_value(source, "summary"), True),
        ]
        self.send_html("Modifier une source", edit_form_panel("Modifier la source", fields, "/sources"))

    def handle_lands(self) -> None:
        with db_connect() as conn:
            rows = conn.execute(
                """
                SELECT l.*, c.name AS clan_name
                FROM lands l
                LEFT JOIN clans c ON c.id = l.clan_id
                ORDER BY l.name
                """
            ).fetchall()
        body = f"""
        <section class="panel">
          <div class="actions"><a class="button" href="/lands/new">Ajouter une terre</a></div>
          {rows_table(["Nom", "Type", "Localisation", "Clan", "Statut", "Confidentialité"], [
              [e(row["name"]), e(row["place_type"]), e(row["location_text"]), e(row["clan_name"]), badge(row["status"]), badge(row["confidentiality"]), f'<a class="button secondary" href="/lands/{row["id"]}/edit">Modifier</a>']
              for row in rows
          ])}
        </section>
        """
        self.send_html("Terres", body)

    def handle_land_new(self) -> None:
        with db_connect() as conn:
            clans = conn.execute("SELECT id, name FROM clans ORDER BY name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    INSERT INTO lands
                    (name, place_type, location_text, clan_id, known_rights, status, source_id, confidentiality)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data["name"],
                        data.get("place_type"),
                        data.get("location_text"),
                        none_if_empty(data.get("clan_id")),
                        data.get("known_rights"),
                        data.get("status", "a_verifier"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "sensible"),
                    ),
                )
                self.redirect("/lands")
                return
        self.send_html("Ajouter une terre", land_form(clans, sources))

    def handle_land_edit(self, path: str) -> None:
        try:
            land_id = int(path.strip("/").split("/")[1])
        except (ValueError, IndexError):
            self.not_found()
            return
        with db_connect() as conn:
            clans = conn.execute("SELECT id, name FROM clans ORDER BY name").fetchall()
            sources = conn.execute("SELECT id, title AS label FROM sources ORDER BY title").fetchall()
            land = conn.execute("SELECT * FROM lands WHERE id = ?", (land_id,)).fetchone()
            if not land:
                self.not_found()
                return
            if self.command == "POST":
                data = self.form()
                conn.execute(
                    """
                    UPDATE lands
                    SET name = ?, place_type = ?, location_text = ?, clan_id = ?, known_rights = ?,
                        status = ?, source_id = ?, confidentiality = ?
                    WHERE id = ?
                    """,
                    (
                        data["name"],
                        data.get("place_type"),
                        data.get("location_text"),
                        none_if_empty(data.get("clan_id")),
                        data.get("known_rights"),
                        data.get("status", "a_verifier"),
                        none_if_empty(data.get("source_id")),
                        data.get("confidentiality", "sensible"),
                        land_id,
                    ),
                )
                self.redirect("/lands")
                return
        fields = [
            text_field("name", "Nom du lieu ou de la terre *", row_value(land, "name"), True),
            text_field("place_type", "Type de lieu", row_value(land, "place_type")),
            select_field("clan_id", "Clan associé", clans, row_value(land, "clan_id")),
            select_field("source_id", "Source", sources, row_value(land, "source_id")),
            select_field("status", "Statut *", CERTAINTY, row_value(land, "status"), True),
            select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, row_value(land, "confidentiality"), True),
            textarea_field("location_text", "Localisation textuelle", row_value(land, "location_text")),
            textarea_field("known_rights", "Droits connus", row_value(land, "known_rights")),
        ]
        self.send_html("Modifier une terre", edit_form_panel("Modifier la terre", fields, "/lands"))

    def handle_search(self) -> None:
        query = parse_qs(urlparse(self.path).query).get("q", [""])[0].strip()
        results = ""
        if query:
            term = f"%{query}%"
            with db_connect() as conn:
                persons = conn.execute(
                    """
                    SELECT DISTINCT p.id, p.first_names, p.current_name, p.birth_name
                    FROM persons p
                    LEFT JOIN person_group_links pgl ON pgl.person_id = p.id
                    LEFT JOIN customary_groups g ON g.id = pgl.group_id
                    WHERE p.first_names LIKE ? OR p.current_name LIKE ? OR p.birth_name LIKE ?
                       OR p.other_names LIKE ? OR g.name LIKE ? OR g.variants LIKE ?
                    """,
                    (term, term, term, term, term, term),
                ).fetchall()
                clans = conn.execute(
                    "SELECT id, name, variants, region FROM clans WHERE name LIKE ? OR variants LIKE ? OR region LIKE ?",
                    (term, term, term),
                ).fetchall()
                events = conn.execute(
                    """
                    SELECT DISTINCT ev.id, ev.title, ev.event_type, ev.event_date
                    FROM customary_events ev
                    LEFT JOIN event_person_links epl ON epl.event_id = ev.id
                    LEFT JOIN persons p ON p.id = epl.person_id
                    WHERE ev.title LIKE ? OR ev.description LIKE ? OR ev.effects LIKE ?
                       OR p.first_names LIKE ? OR p.current_name LIKE ?
                    """,
                    (term, term, term, term, term),
                ).fetchall()
                sources = conn.execute(
                    "SELECT title, source_type, witness_name FROM sources WHERE title LIKE ? OR summary LIKE ? OR witness_name LIKE ?",
                    (term, term, term),
                ).fetchall()
                groups = conn.execute(
                    "SELECT id, name, group_type, variants, region FROM customary_groups WHERE name LIKE ? OR variants LIKE ? OR region LIKE ? OR description LIKE ?",
                    (term, term, term, term),
                ).fetchall()
                customary_names = conn.execute(
                    "SELECT id, name, name_type, variants, region FROM customary_names WHERE name LIKE ? OR variants LIKE ? OR region LIKE ? OR description LIKE ?",
                    (term, term, term, term),
                ).fetchall()
                viva = conn.execute(
                    "SELECT id, title, area, collector_author FROM viva_lists WHERE title LIKE ? OR area LIKE ? OR collector_author LIKE ? OR description LIKE ?",
                    (term, term, term, term),
                ).fetchall()
                research = conn.execute(
                    "SELECT title, author_view, statement, certainty FROM research_items WHERE title LIKE ? OR statement LIKE ? OR interpretation LIKE ? OR author_view LIKE ?",
                    (term, term, term, term),
                ).fetchall()
                lineages = conn.execute(
                    "SELECT title, author_or_collector, chain_text, certainty FROM lineage_genealogies WHERE title LIKE ? OR author_or_collector LIKE ? OR chain_text LIKE ? OR interpretation LIKE ?",
                    (term, term, term, term),
                ).fetchall()
                functions = conn.execute(
                    "SELECT title, function_type, place, certainty FROM customary_functions WHERE title LIKE ? OR description LIKE ? OR place LIKE ?",
                    (term, term, term),
                ).fetchall()
            results = f"""
            <h2>Résultats</h2>
            <h3>Personnes</h3>
            {rows_table(["Prénom", "Nom de naissance", "Nom marital", "Action"], [[f'<a href="/persons/{p["id"]}">{e(p["first_names"])}</a>', e(p["birth_name"]), e(p["current_name"]), delete_button(f'/persons/{p["id"]}/delete', "Personne", f'{p["first_names"]} - {p["birth_name"] or "Inconnu"}')] for p in persons])}
            <h3>Clans</h3>
            {rows_table(["Nom", "Variantes", "Région", "Action"], [[f'<a href="/clans/{c["id"]}">{e(c["name"])}</a>', e(c["variants"]), e(c["region"]), delete_button(f'/clans/{c["id"]}/delete', "Clan", f'{c["name"]} - {c["region"] or ""}')] for c in clans])}
            <h3>Groupes coutumiers</h3>
            {rows_table(["Nom", "Type", "Variantes", "Région", "Action"], [[f'<a href="/groups/{g["id"]}">{e(g["name"])}</a>', e(group_type_label(g["group_type"])), e(g["variants"]), e(g["region"]), delete_button(f'/groups/{g["id"]}/delete', "Groupe", f'{g["name"]} - {group_type_label(g["group_type"])}')] for g in groups])}
            <h3>Noms coutumiers</h3>
            {rows_table(["Nom", "Type", "Variantes", "Région", "Action"], [[f'<a href="/customary-names/{cn["id"]}">{e(cn["name"])}</a>', e(group_type_label(cn["name_type"])), e(cn["variants"]), e(cn["region"]), delete_button(f'/customary-names/{cn["id"]}/delete', "Nom coutumier", f'{cn["name"]} - {group_type_label(cn["name_type"])}')] for cn in customary_names])}
            <h3>Listes viva</h3>
            {rows_table(["Titre", "Aire", "Auteur / collecteur", "Action"], [[f'<a href="/viva/{v["id"]}">{e(v["title"])}</a>', e(v["area"]), e(v["collector_author"]), delete_button(f'/viva/{v["id"]}/delete', "Liste viva", f'{v["title"]} - {v["area"] or ""}')] for v in viva])}
            <h3>Informations et hypotheses</h3>
            {rows_table(["Titre", "Auteur / vue", "Information", "Certitude"], [[e(r["title"]), e(r["author_view"]), e(r["statement"]), badge(r["certainty"])] for r in research])}
            <h3>Genealogies de lignage</h3>
            {rows_table(["Titre", "Auteur / collecteur", "Chaine", "Certitude"], [[e(l["title"]), e(l["author_or_collector"]), e(l["chain_text"]), badge(l["certainty"])] for l in lineages])}
            <h3>Fonctions coutumieres</h3>
            {rows_table(["Titre", "Type", "Lieu", "Certitude"], [[e(f["title"]), e(function_type_label(f["function_type"])), e(f["place"]), badge(f["certainty"])] for f in functions])}
            <h3>Événements</h3>
            {rows_table(["Date", "Type", "Titre", "Action"], [[e(ev["event_date"]), e(event_label(ev["event_type"])), e(ev["title"]), delete_button(f'/events/{ev["id"]}/delete', "Événement", f'{ev["title"]} - {event_label(ev["event_type"])}')] for ev in events])}
            <h3>Sources</h3>
            {rows_table(["Titre", "Type", "Témoin"], [[e(s["title"]), e(source_label(s["source_type"])), e(s["witness_name"])] for s in sources])}
            """
        body = f"""
        <section class="panel">
          <form method="get" action="/search">
            <label for="q">Recherche globale</label>
            <div style="display:flex; gap:10px;">
              <input id="q" name="q" value="{e(query)}" placeholder="Nom, clan, lieu, source, mot clé">
              <button type="submit">Rechercher</button>
            </div>
          </form>
        </section>
        <section class="panel">{results or '<p class="empty">Saisis un mot pour rechercher dans la base.</p>'}</section>
        """
        self.send_html("Recherche", body)


def setup_form(error: str = "") -> str:
    return f"""
    <section class="panel" style="max-width:560px;margin:0 auto;">
      <h2>Créer le mot de passe local</h2>
      <p>Ce mot de passe protégera l'accès au prototype sur ce PC.</p>
      {f'<p class="flash">{e(error)}</p>' if error else ''}
      <form method="post" action="/setup">
        <p><label>Mot de passe</label><input type="password" name="password" required minlength="8"></p>
        <p><label>Confirmer</label><input type="password" name="confirm" required minlength="8"></p>
        <button type="submit">Créer le mot de passe</button>
      </form>
    </section>
    """


def login_form(error: str = "") -> str:
    return f"""
    <section class="panel" style="max-width:480px;margin:0 auto;">
      <h2>Connexion</h2>
      {f'<p class="flash">{e(error)}</p>' if error else ''}
      <form method="post" action="/login">
        <p><label>Mot de passe</label><input type="password" name="password" required></p>
        <button type="submit">Entrer</button>
      </form>
    </section>
    """


def metric(label: str, value: int) -> str:
    return f'<div class="metric"><strong>{value}</strong><span>{e(label)}</span></div>'


def badge(value: str | None) -> str:
    if not value:
        return ""
    label = dict(CONFIDENTIALITY + CERTAINTY).get(value, value)
    return f'<span class="badge {e(value)}">{e(label)}</span>'


def rows_table(headers: list[str], rows: list[list[str]]) -> str:
    if not rows:
        return '<p class="empty">Aucune donnée pour le moment.</p>'
    header_html = "".join(f"<th>{e(h)}</th>" for h in headers)
    body_html = "".join("<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>" for row in rows)
    return f"<table><thead><tr>{header_html}</tr></thead><tbody>{body_html}</tbody></table>"


def delete_button(path: str, label: str, details: str) -> str:
    message = f"Confirmer la suppression ?\\n\\n{label}\\n{details}"
    return f"""
    <form method="post" action="{e(path)}" style="display:inline;" onsubmit="return confirm('{e(js_string(message))}');">
      <button type="submit" class="button secondary">Supprimer</button>
    </form>
    """


def js_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'").replace("\r", "").replace("\n", "\\n")


def js_literal(value: str) -> str:
    return "'" + js_string(value).replace('"', '\\"') + "'"


def bulk_indices(data: dict[str, str], prefix: str) -> list[int]:
    indexes: list[int] = []
    for key in data:
        if key.startswith(prefix):
            suffix = key[len(prefix) :]
            if suffix.isdigit():
                indexes.append(int(suffix))
    return sorted(set(indexes))


def bulk_flash(label: str, created: int, skipped: int, errors: list[str]) -> str:
    parts = [f"{label}: {created} ligne(s) créée(s)."]
    if skipped:
        parts.append(f"{skipped} ligne(s) ignorée(s).")
    if errors:
        parts.append("À vérifier: " + " / ".join(errors[:8]))
        if len(errors) > 8:
            parts.append(f"... et {len(errors) - 8} autre(s) message(s).")
    return " ".join(parts)


def first_value(data: dict[str, list[str]], key: str, default: str = "") -> str:
    values = data.get(key) or []
    return values[0] if values else default


def relationship_exists(conn: sqlite3.Connection, subject_id: str, object_id: str, relation_types: tuple[str, ...]) -> bool:
    placeholders = ",".join("?" for _ in relation_types)
    return bool(
        conn.execute(
            f"""
            SELECT 1 FROM relationships
            WHERE subject_person_id = ? AND object_person_id = ?
              AND relation_type IN ({placeholders})
            """,
            (subject_id, object_id, *relation_types),
        ).fetchone()
    )


def insert_relationship_once(
    conn: sqlite3.Connection,
    subject_id: str,
    object_id: str,
    relation_type: str,
    certainty: str,
    confidentiality: str,
    comment: str,
) -> tuple[bool, str | None]:
    if subject_id == object_id:
        return False, "Une personne ne peut pas être reliée à elle-même."
    if relationship_exists(conn, subject_id, object_id, (relation_type,)):
        return False, "Relation déjà existante."
    conn.execute(
        """
        INSERT INTO relationships
        (subject_person_id, object_person_id, object_clan_id, relation_type, comment, confidentiality, certainty)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (subject_id, object_id, None, relation_type, comment, confidentiality, certainty),
    )
    return True, None


def insert_spouse_relation(
    conn: sqlite3.Connection,
    person_a_id: str,
    person_b_id: str,
    certainty: str,
    confidentiality: str,
    comment: str,
) -> tuple[bool, str | None]:
    if relationship_exists(conn, person_a_id, person_b_id, ("conjoint",)) or relationship_exists(conn, person_b_id, person_a_id, ("conjoint",)):
        return False, "Relation de conjoint déjà existante."
    return insert_relationship_once(conn, person_a_id, person_b_id, "conjoint", certainty, confidentiality, comment)


def insert_sibling_relation(
    conn: sqlite3.Connection,
    person_a_id: str,
    person_b_id: str,
    certainty: str,
    confidentiality: str,
    comment: str,
) -> tuple[bool, str | None]:
    if relationship_exists(conn, person_a_id, person_b_id, ("frere_soeur",)) or relationship_exists(conn, person_b_id, person_a_id, ("frere_soeur",)):
        return False, "Relation frère/sœur déjà existante."
    return insert_relationship_once(conn, person_a_id, person_b_id, "frere_soeur", certainty, confidentiality, comment)


def materialize_inferred_sibling_relationships(conn: sqlite3.Connection, person_id: int) -> None:
    inferred = conn.execute(
        """
        WITH parent_links(parent_id) AS (
          SELECT r.subject_person_id
          FROM relationships r
          WHERE r.object_person_id = ? AND r.relation_type IN ('pere','mere','parent_coutumier')
          UNION
          SELECT r.object_person_id
          FROM relationships r
          WHERE r.subject_person_id = ? AND r.relation_type = 'enfant'
        ),
        sibling_links(person_id, parent_id) AS (
          SELECT r.object_person_id, r.subject_person_id
          FROM relationships r
          JOIN parent_links pl ON pl.parent_id = r.subject_person_id
          WHERE r.relation_type IN ('pere','mere','parent_coutumier') AND r.object_person_id != ?
          UNION
          SELECT r.subject_person_id, r.object_person_id
          FROM relationships r
          JOIN parent_links pl ON pl.parent_id = r.object_person_id
          WHERE r.relation_type = 'enfant' AND r.subject_person_id != ?
        )
        SELECT
          p.id AS sibling_id,
          GROUP_CONCAT(DISTINCT pp.first_names || ' - ' || COALESCE(NULLIF(pp.birth_name, ''), 'Inconnu')) AS common_parents
        FROM sibling_links sl
        JOIN persons p ON p.id = sl.person_id
        JOIN persons pp ON pp.id = sl.parent_id
        WHERE p.id != ?
          AND NOT EXISTS (
            SELECT 1
            FROM inferred_relationship_exclusions ire
            WHERE ire.relation_type = 'frere_soeur'
              AND ire.person_a_id = MIN(?, p.id)
              AND ire.person_b_id = MAX(?, p.id)
          )
        GROUP BY p.id
        """,
        (person_id, person_id, person_id, person_id, person_id, person_id, person_id),
    ).fetchall()
    for row in inferred:
        sibling_id = str(row["sibling_id"])
        if relationship_exists(conn, str(person_id), sibling_id, ("frere_soeur",)) or relationship_exists(conn, sibling_id, str(person_id), ("frere_soeur",)):
            continue
        conn.execute(
            """
            INSERT INTO relationships
            (subject_person_id, object_person_id, object_clan_id, relation_type, comment, confidentiality, certainty)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                person_id,
                sibling_id,
                None,
                "frere_soeur",
                f"Relation frère/sœur déduite automatiquement par parent(s) commun(s): {row['common_parents'] or ''}",
                "familial",
                "probable",
            ),
        )


def parent_child_exists(conn: sqlite3.Connection, parent_id: str, child_id: str) -> bool:
    return (
        relationship_exists(conn, parent_id, child_id, ("pere", "mere", "parent_coutumier"))
        or relationship_exists(conn, child_id, parent_id, ("enfant",))
    )


def insert_parent_child_relation(
    conn: sqlite3.Connection,
    parent_id: str,
    child_id: str,
    certainty: str,
    confidentiality: str,
    comment: str,
) -> tuple[bool, str | None]:
    if parent_id == child_id:
        return False, "Une personne ne peut pas être son propre parent."
    if parent_child_exists(conn, parent_id, child_id):
        return False, "Relation parent/enfant déjà existante."
    gender = conn.execute("SELECT gender FROM persons WHERE id = ?", (parent_id,)).fetchone()
    if gender and is_male(gender["gender"]):
        return insert_relationship_once(conn, parent_id, child_id, "pere", certainty, confidentiality, comment)
    if gender and is_female(gender["gender"]):
        return insert_relationship_once(conn, parent_id, child_id, "mere", certainty, confidentiality, comment)
    return insert_relationship_once(conn, child_id, parent_id, "enfant", certainty, confidentiality, comment)


def excel_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def excel_header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {excel_text(value).lower(): index for index, value in enumerate(row) if excel_text(value)}


def excel_cell(row: tuple[Any, ...], headers: dict[str, int], key: str) -> str:
    index = headers.get(key)
    if index is None or index >= len(row):
        return ""
    return excel_text(row[index])


def option_value_from_text(options: list[tuple[str, str]], value: str, default: str) -> str:
    text = value.strip()
    if not text:
        return default
    lower = text.lower()
    for option_value, label in options:
        if lower in {option_value.lower(), label.lower()}:
            return option_value
    return default


def resolve_person_reference(conn: sqlite3.Connection, person_id: str, first_names: str, birth_name: str) -> tuple[str | None, str | None]:
    if person_id:
        row = conn.execute("SELECT id FROM persons WHERE id = ?", (person_id,)).fetchone()
        return (person_id, None) if row else (None, f"fiche #{person_id} introuvable")
    if not first_names or not birth_name:
        return None, "id ou prénom + nom de naissance obligatoire"
    rows = conn.execute(
        """
        SELECT id FROM persons
        WHERE lower(first_names) = lower(?) AND lower(birth_name) = lower(?)
        """,
        (first_names, birth_name),
    ).fetchall()
    if len(rows) == 1:
        return str(rows[0]["id"]), None
    if not rows:
        return None, f"personne introuvable: {first_names} - {birth_name}"
    return None, f"homonyme ambigu: {first_names} - {birth_name}; utilise l'id de fiche"


def build_bulk_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Personnes"
    sheets = {
        "Personnes": [
            "nom_naissance",
            "prenom",
            "autres_prenoms",
            "sexe",
            "nom_marital",
            "date_naissance",
            "lieu_naissance",
            "date_deces",
            "notes",
        ],
        "Noms coutumiers": [
            "nom_coutumier",
            "type",
            "variantes",
            "region",
            "description",
        ],
        "Relations": [
            "personne_a_id",
            "personne_a_prenom",
            "personne_a_nom_naissance",
            "relation",
            "personne_b_id",
            "personne_b_prenom",
            "personne_b_nom_naissance",
            "certitude",
            "confidentialite",
            "commentaire",
        ],
    }
    for title, headers in sheets.items():
        ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = 22
        ws.freeze_panes = "A2"
    ws = wb.create_sheet("Exemples")
    ws.append(["feuille", "exemple"])
    ws.append(["Personnes", "NOM_EXEMPLE | Prenom | autres prenoms | M | | vers 1960 | Lieu exemple | | note libre"])
    ws.append(["Noms coutumiers", "Nom coutumier exemple | type_a_verifier | Variante | Region exemple | note libre"])
    ws.append(["Relations", "personne_a_id ou prénom+nom | pere | personne_b_id ou prénom+nom | confirme | familial"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    for ws in wb.worksheets:
        if ws.title == "Personnes":
            validation = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
            ws.add_data_validation(validation)
            validation.add("C2:C500")
        if ws.title == "Relations":
            relation_values = ",".join(value for value, _ in RELATION_TYPES)
            certainty_values = ",".join(value for value, _ in CERTAINTY)
            confidentiality_values = ",".join(value for value, _ in CONFIDENTIALITY)
            for cell_range, values in (("D2:D500", relation_values), ("H2:H500", certainty_values), ("I2:I500", confidentiality_values)):
                validation = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
                ws.add_data_validation(validation)
                validation.add(cell_range)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def import_bulk_workbook(file_bytes: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    created = {"Personnes": 0, "Noms coutumiers": 0, "Relations": 0}
    skipped = 0
    errors: list[str] = []
    with db_connect() as conn:
        if "Personnes" in wb.sheetnames:
            ws = wb["Personnes"]
            headers = excel_header_map(tuple(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))))
            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                birth_name = excel_cell(row, headers, "nom_naissance")
                first_names = excel_cell(row, headers, "prenom")
                if not any(excel_text(value) for value in row):
                    continue
                if not birth_name or not first_names:
                    skipped += 1
                    errors.append(f"Personnes ligne {row_number}: nom_naissance et prenom obligatoires.")
                    continue
                gender = excel_cell(row, headers, "sexe").upper()
                if gender not in {"", "M", "F"}:
                    skipped += 1
                    errors.append(f"Personnes ligne {row_number}: sexe invalide.")
                    continue
                conn.execute(
                    """
                    INSERT INTO persons
                    (first_names, current_name, birth_name, other_names, gender, birth_date, birth_place, death_date,
                     current_clan_id, origin_clan_id, notes, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        first_names,
                        excel_cell(row, headers, "nom_marital"),
                        birth_name,
                        excel_cell(row, headers, "autres_prenoms"),
                        gender,
                        excel_cell(row, headers, "date_naissance"),
                        excel_cell(row, headers, "lieu_naissance"),
                        excel_cell(row, headers, "date_deces"),
                        None,
                        None,
                        excel_cell(row, headers, "notes"),
                        "familial",
                        "a_verifier",
                    ),
                )
                created["Personnes"] += 1

        if "Noms coutumiers" in wb.sheetnames:
            ws = wb["Noms coutumiers"]
            headers = excel_header_map(tuple(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))))
            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = excel_cell(row, headers, "nom_coutumier")
                if not any(excel_text(value) for value in row):
                    continue
                if not name:
                    skipped += 1
                    errors.append(f"Noms coutumiers ligne {row_number}: nom_coutumier obligatoire.")
                    continue
                exists = conn.execute("SELECT 1 FROM customary_names WHERE lower(name) = lower(?)", (name,)).fetchone()
                if exists:
                    skipped += 1
                    errors.append(f"Noms coutumiers ligne {row_number}: {name} existe déjà.")
                    continue
                conn.execute(
                    """
                    INSERT INTO customary_names
                    (name, name_type, variants, region, description, source_id, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        name,
                        option_value_from_text(GROUP_TYPES, excel_cell(row, headers, "type"), "a_verifier"),
                        excel_cell(row, headers, "variantes"),
                        excel_cell(row, headers, "region"),
                        excel_cell(row, headers, "description"),
                        None,
                        "familial",
                        "a_verifier",
                    ),
                )
                created["Noms coutumiers"] += 1

        if "Relations" in wb.sheetnames:
            ws = wb["Relations"]
            headers = excel_header_map(tuple(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))))
            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(excel_text(value) for value in row):
                    continue
                relation_type = option_value_from_text(RELATION_TYPES, excel_cell(row, headers, "relation"), "")
                subject_id, subject_error = resolve_person_reference(
                    conn,
                    excel_cell(row, headers, "personne_a_id"),
                    excel_cell(row, headers, "personne_a_prenom"),
                    excel_cell(row, headers, "personne_a_nom_naissance"),
                )
                object_id, object_error = resolve_person_reference(
                    conn,
                    excel_cell(row, headers, "personne_b_id"),
                    excel_cell(row, headers, "personne_b_prenom"),
                    excel_cell(row, headers, "personne_b_nom_naissance"),
                )
                if subject_error or object_error or not relation_type:
                    skipped += 1
                    errors.append(f"Relations ligne {row_number}: {subject_error or object_error or 'relation invalide'}.")
                    continue
                if subject_id == object_id:
                    skipped += 1
                    errors.append(f"Relations ligne {row_number}: une personne ne peut pas être reliée à elle-même.")
                    continue
                duplicate = conn.execute(
                    "SELECT 1 FROM relationships WHERE subject_person_id = ? AND object_person_id = ? AND relation_type = ?",
                    (subject_id, object_id, relation_type),
                ).fetchone()
                if duplicate:
                    skipped += 1
                    errors.append(f"Relations ligne {row_number}: relation déjà existante.")
                    continue
                conn.execute(
                    """
                    INSERT INTO relationships
                    (subject_person_id, object_person_id, object_clan_id, relation_type, comment, confidentiality, certainty)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        subject_id,
                        object_id,
                        None,
                        relation_type,
                        excel_cell(row, headers, "commentaire"),
                        option_value_from_text(CONFIDENTIALITY, excel_cell(row, headers, "confidentialite"), "familial"),
                        option_value_from_text(CERTAINTY, excel_cell(row, headers, "certitude"), "a_verifier"),
                    ),
                )
                created["Relations"] += 1

    parts = [f"Import Excel terminé: {value} {key.lower()} créée(s)" for key, value in created.items() if value]
    if not parts:
        parts.append("Import Excel terminé: aucune ligne créée.")
    if skipped:
        parts.append(f"{skipped} ligne(s) ignorée(s).")
    if errors:
        parts.append("À vérifier: " + " / ".join(errors[:8]))
        if len(errors) > 8:
            parts.append(f"... et {len(errors) - 8} autre(s) message(s).")
    return " ".join(parts)


def bulk_cell(content: str) -> str:
    return f"<td>{content}</td>"


def bulk_input(name: str, placeholder: str = "", width: int = 150) -> str:
    return f'<input name="{e(name)}" placeholder="{e(placeholder)}" style="min-width:{width}px;">'


def bulk_select(name: str, options: list[tuple[str, str]] | list[sqlite3.Row], selected: str = "", width: int = 150) -> str:
    if options and isinstance(options[0], sqlite3.Row):
        options_html = option_list(options, selected)
    else:
        options_html = select_options(options, selected)  # type: ignore[arg-type]
    return f'<select name="{e(name)}" style="min-width:{width}px;">{options_html}</select>'


def person_select(name: str, label: str, persons: list[sqlite3.Row], required: bool = False) -> str:
    return f'<p><label>{e(label)}</label><select name="{e(name)}"{" required" if required else ""}>{option_list(persons, empty_label="Choisir une personne")}</select></p>'


def person_multi_select(name: str, label: str, persons: list[sqlite3.Row], height: int = 10) -> str:
    options = []
    for person in persons:
        options.append(f'<option value="{e(str(person["id"]))}">{e(person["label"])}</option>')
    return f'<p class="full"><label>{e(label)}</label><select name="{e(name)}" multiple size="{height}">{"".join(options)}</select></p>'


def searchable_person_multi(name: str, label: str, persons: list[sqlite3.Row], component_id: str) -> str:
    js_people = ",".join(
        "{id:'" + js_string(str(person["id"])) + "',label:'" + js_string(person["label"]) + "'}"
        for person in persons
    )
    return f"""
    <div class="full person-picker" id="{e(component_id)}">
      <label>{e(label)}</label>
      <div style="display:grid; grid-template-columns: minmax(220px, 1fr) auto; gap:8px; align-items:end;">
        <p style="margin:0;"><input type="search" class="person-picker-search" placeholder="Rechercher par prénom, nom ou numéro de fiche"></p>
        <button type="button" class="person-picker-add">Ajouter</button>
      </div>
      <select class="person-picker-results" size="6" style="margin-top:8px;"></select>
      <div class="person-picker-selected empty" style="margin-top:8px;">Aucun enfant sélectionné.</div>
      <div class="person-picker-hidden"></div>
    </div>
    <script>
      (function() {{
        var root = document.getElementById({js_literal(component_id)});
        if (!root) return;
        var people = [{js_people}];
        var fieldName = {js_literal(name)};
        var search = root.querySelector('.person-picker-search');
        var results = root.querySelector('.person-picker-results');
        var add = root.querySelector('.person-picker-add');
        var selectedBox = root.querySelector('.person-picker-selected');
        var hiddenBox = root.querySelector('.person-picker-hidden');
        var selected = [];

        function normalize(value) {{
          return (value || '').toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '');
        }}

        function renderResults() {{
          var term = normalize(search.value);
          var matches = people.filter(function(person) {{
            return !term || normalize(person.label).indexOf(term) !== -1;
          }}).slice(0, 80);
          results.innerHTML = matches.map(function(person) {{
            return '<option value="' + person.id + '">' + person.label.replace(/&/g, '&amp;').replace(/</g, '&lt;') + '</option>';
          }}).join('');
          if (results.options.length) results.selectedIndex = 0;
        }}

        function renderSelected() {{
          hiddenBox.innerHTML = selected.map(function(person) {{
            return '<input type="hidden" name="' + fieldName + '" value="' + person.id + '">';
          }}).join('');
          if (!selected.length) {{
            selectedBox.className = 'person-picker-selected empty';
            selectedBox.innerHTML = 'Aucun enfant sélectionné.';
            return;
          }}
          selectedBox.className = 'person-picker-selected';
          selectedBox.innerHTML = selected.map(function(person) {{
            return '<span class="badge" style="margin:3px;">' +
              person.label.replace(/&/g, '&amp;').replace(/</g, '&lt;') +
              ' <button type="button" class="person-picker-remove" data-id="' + person.id + '" style="padding:2px 6px; margin-left:6px;">Retirer</button></span>';
          }}).join('');
        }}

        function addSelected() {{
          var id = results.value;
          if (!id || selected.some(function(person) {{ return person.id === id; }})) return;
          var person = people.find(function(item) {{ return item.id === id; }});
          if (!person) return;
          selected.push(person);
          renderSelected();
        }}

        search.addEventListener('input', renderResults);
        add.addEventListener('click', addSelected);
        results.addEventListener('dblclick', addSelected);
        selectedBox.addEventListener('click', function(event) {{
          if (!event.target.classList.contains('person-picker-remove')) return;
          var id = event.target.getAttribute('data-id');
          selected = selected.filter(function(person) {{ return person.id !== id; }});
          renderSelected();
        }});
        renderResults();
        renderSelected();
      }})();
    </script>
    """


def searchable_person_single(name: str, label: str, persons: list[sqlite3.Row], component_id: str, required: bool = False) -> str:
    js_people = ",".join(
        "{id:'" + js_string(str(person["id"])) + "',label:'" + js_string(person["label"]) + "'}"
        for person in persons
    )
    required_js = "true" if required else "false"
    return f"""
    <div class="full person-picker" id="{e(component_id)}">
      <label>{e(label)}</label>
      <input type="hidden" name="{e(name)}" class="person-picker-value">
      <div style="display:grid; grid-template-columns: minmax(220px, 1fr) auto; gap:8px; align-items:end;">
        <p style="margin:0;"><input type="search" class="person-picker-search" placeholder="Rechercher par prénom, nom, autre prénom ou numéro de fiche"></p>
        <button type="button" class="person-picker-add">Choisir</button>
      </div>
      <select class="person-picker-results" size="7" style="margin-top:8px;"></select>
      <div class="person-picker-selected empty" style="margin-top:8px;">Aucune personne choisie.</div>
      <button type="button" class="person-picker-clear secondary" style="margin-top:8px;">Effacer</button>
    </div>
    <script>
      (function() {{
        var root = document.getElementById({js_literal(component_id)});
        if (!root) return;
        var people = [{js_people}];
        var required = {required_js};
        var hidden = root.querySelector('.person-picker-value');
        var search = root.querySelector('.person-picker-search');
        var results = root.querySelector('.person-picker-results');
        var add = root.querySelector('.person-picker-add');
        var clear = root.querySelector('.person-picker-clear');
        var selectedBox = root.querySelector('.person-picker-selected');

        function escapeHtml(value) {{
          return (value || '').replace(/&/g, '&amp;').replace(/</g, '&lt;');
        }}
        function normalize(value) {{
          return (value || '').toLowerCase().normalize('NFD').replace(/[\\u0300-\\u036f]/g, '');
        }}
        function renderResults() {{
          var term = normalize(search.value);
          var matches = people.filter(function(person) {{
            return !term || normalize(person.label).indexOf(term) !== -1 || normalize('#' + person.id).indexOf(term) !== -1;
          }}).slice(0, 80);
          results.innerHTML = matches.map(function(person) {{
            return '<option value="' + person.id + '">' + escapeHtml(person.label) + '</option>';
          }}).join('');
          if (results.options.length) results.selectedIndex = 0;
        }}
        function chooseSelected() {{
          var id = results.value;
          if (!id) return;
          var person = people.find(function(item) {{ return item.id === id; }});
          if (!person) return;
          hidden.value = person.id;
          selectedBox.className = 'person-picker-selected';
          selectedBox.innerHTML = '<span class="badge">' + escapeHtml(person.label) + '</span>';
        }}
        function clearSelected() {{
          hidden.value = '';
          selectedBox.className = 'person-picker-selected empty';
          selectedBox.innerHTML = 'Aucune personne choisie.';
        }}
        var form = root.closest('form');
        if (form) {{
          form.addEventListener('submit', function(event) {{
            if (required && !hidden.value) {{
              event.preventDefault();
              alert('Choisis une personne pour le champ: {js_string(label)}.');
            }}
          }});
        }}
        search.addEventListener('input', renderResults);
        add.addEventListener('click', chooseSelected);
        clear.addEventListener('click', clearSelected);
        results.addEventListener('dblclick', chooseSelected);
        renderResults();
      }})();
    </script>
    """


def quick_relations_form(persons: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Relations rapides</h2>
      <p class="empty">Utilise ces formulaires pour créer plusieurs liens familiaux d'un seul coup. Les relations déjà existantes sont ignorées, et les frères/sœurs peuvent être déduits automatiquement quand les parents sont les mêmes.</p>
    </section>
    <section class="panel">
      <h2>Couple + enfants</h2>
      <form method="post" action="/quick-relations/couple" class="form-grid" onsubmit="return confirm('Créer les relations couple + enfants ?');">
        {searchable_person_single("parent1_id", "Parent 1 *", persons, "couple-parent1-picker", True)}
        {searchable_person_single("parent2_id", "Parent 2 / conjoint éventuel", persons, "couple-parent2-picker")}
        {searchable_person_multi("child_ids", "Enfants du couple / du parent", persons, "couple-child-picker")}
        <p><label>Certitude</label><select name="certainty">{select_options(CERTAINTY, "confirme")}</select></p>
        <p><label>Confidentialité</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p class="full"><label>Commentaire commun</label><textarea name="comment" placeholder="Optionnel: source orale, précision, doute..."></textarea></p>
        <p class="full"><button type="submit">Créer ces relations</button></p>
      </form>
    </section>
    <section class="panel">
      <h2>Personne centrale</h2>
      <form method="post" action="/quick-relations/central" class="form-grid" onsubmit="return confirm('Créer les relations autour de cette personne ?');">
        {searchable_person_single("central_id", "Personne centrale *", persons, "central-person-picker", True)}
        {searchable_person_single("father_id", "Père / parent 1", persons, "central-father-picker")}
        {searchable_person_single("mother_id", "Mère / parent 2", persons, "central-mother-picker")}
        {searchable_person_multi("spouse_ids", "Conjoints / conjointes", persons, "central-spouse-picker")}
        {searchable_person_multi("central_child_ids", "Enfants", persons, "central-child-picker")}
        {searchable_person_multi("sibling_ids", "Frères / sœurs directs si parents inconnus", persons, "central-sibling-picker")}
        <p><label>Certitude</label><select name="certainty">{select_options(CERTAINTY, "confirme")}</select></p>
        <p><label>Confidentialité</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p class="full"><label>Commentaire commun</label><textarea name="comment" placeholder="Optionnel"></textarea></p>
        <p class="full"><button type="submit">Créer les relations autour de la personne</button></p>
      </form>
    </section>
    """


def extended_tree_picker_form(persons: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <div class="actions"><a class="button secondary" href="/visualisations">Retour visualisations</a></div>
      <h2>Ouvrir un arbre étendu</h2>
      <form method="get" action="/visualisations/extended-tree" class="form-grid">
        {searchable_person_single("person_id", "Personne de départ *", persons, "extended-tree-person-picker", True)}
        <p><label>Mode</label><select name="mode">{select_options([("descendance", "Descendance"), ("ascendance", "Ascendance")], "descendance")}</select></p>
        <p><label>Générations</label><select name="generations">{select_options([("3", "3"), ("5", "5"), ("8", "8")], "8")}</select></p>
        <p><label><input type="checkbox" name="same_name" value="1"> Même nom de naissance seulement</label></p>
        <p><label><input type="checkbox" name="spouses" value="1" checked> Afficher conjoints</label></p>
        <p><label><input type="checkbox" name="name_color" value="1"> Couleur par nom de naissance</label></p>
        <p><label><input type="checkbox" name="branch_color" value="1"> Couleur par branche descendante</label></p>
        <p class="full"><button type="submit">Afficher l'arbre étendu</button></p>
      </form>
    </section>
    """


def bulk_form(persons: list[sqlite3.Row]) -> str:
    person_rows = []
    for index in range(30):
        person_rows.append(
            "<tr>"
            + bulk_cell(bulk_input(f"person_birth_name_{index}", "ex: NOM_EXEMPLE", 150))
            + bulk_cell(bulk_input(f"person_first_names_{index}", "ex: Jean-Claude", 150))
            + bulk_cell(bulk_input(f"person_other_names_{index}", "", 150))
            + bulk_cell(bulk_select(f"person_gender_{index}", SEXES, "", 120))
            + bulk_cell(bulk_input(f"person_current_name_{index}", "", 150))
            + bulk_cell(bulk_input(f"person_birth_date_{index}", "ex: vers 1950", 130))
            + bulk_cell(bulk_input(f"person_birth_place_{index}", "", 140))
            + bulk_cell(bulk_input(f"person_death_date_{index}", "", 130))
            + bulk_cell(bulk_input(f"person_notes_{index}", "", 240))
            + "</tr>"
        )
    customary_rows = []
    for index in range(20):
        customary_rows.append(
            "<tr>"
            + bulk_cell(bulk_input(f"customary_name_{index}", "ex: Nom coutumier exemple", 160))
            + bulk_cell(bulk_select(f"customary_type_{index}", GROUP_TYPES, "a_verifier", 170))
            + bulk_cell(bulk_input(f"customary_variants_{index}", "", 160))
            + bulk_cell(bulk_input(f"customary_region_{index}", "ex: Kouaoua", 150))
            + bulk_cell(bulk_input(f"customary_description_{index}", "", 280))
            + "</tr>"
        )
    relationship_rows = []
    for index in range(40):
        relationship_rows.append(
            "<tr>"
            + bulk_cell(bulk_select(f"rel_subject_{index}", persons, "", 260))
            + bulk_cell(bulk_select(f"rel_type_{index}", RELATION_TYPES, "pere", 170))
            + bulk_cell(bulk_select(f"rel_object_{index}", persons, "", 260))
            + bulk_cell(bulk_select(f"rel_certainty_{index}", CERTAINTY, "a_verifier", 140))
            + bulk_cell(bulk_select(f"rel_confidentiality_{index}", CONFIDENTIALITY, "familial", 130))
            + bulk_cell(bulk_input(f"rel_comment_{index}", "", 260))
            + "</tr>"
        )
    return f"""
    <section class="panel">
      <h2>Saisie en masse</h2>
      <p class="empty">Remplis uniquement les lignes utiles. Les lignes totalement vides sont ignorées. Pour les relations, lis la ligne comme: personne A est le rôle choisi de personne B.</p>
      <div class="actions">
        <a class="button" href="/bulk/template.xlsx">Télécharger le modèle Excel</a>
      </div>
      <form method="post" action="/bulk/excel" enctype="multipart/form-data" style="margin-top:12px;" onsubmit="return confirm('Importer ce fichier Excel dans la base ?');">
        <p><label>Importer un fichier Excel rempli</label><input type="file" name="excel_file" accept=".xlsx" required></p>
        <p><button type="submit">Importer Excel</button></p>
      </form>
    </section>
    <section class="panel">
      <h2>Personnes</h2>
      <form method="post" action="/bulk/persons" onsubmit="return confirm('Enregistrer toutes les personnes renseignées ?');">
        <div style="overflow:auto;">
          <table>
            <thead><tr><th>Nom de naissance *</th><th>Prénom *</th><th>Autres prénoms / noms</th><th>Sexe</th><th>Nom marital</th><th>Naissance</th><th>Lieu naissance</th><th>Décès</th><th>Notes</th></tr></thead>
            <tbody>{''.join(person_rows)}</tbody>
          </table>
        </div>
        <p><button type="submit">Enregistrer les personnes</button></p>
      </form>
    </section>
    <section class="panel">
      <h2>Noms coutumiers</h2>
      <form method="post" action="/bulk/customary-names" onsubmit="return confirm('Enregistrer tous les noms coutumiers renseignés ?');">
        <div style="overflow:auto;">
          <table>
            <thead><tr><th>Nom coutumier *</th><th>Type</th><th>Variantes</th><th>Région / aire</th><th>Description</th></tr></thead>
            <tbody>{''.join(customary_rows)}</tbody>
          </table>
        </div>
        <p><button type="submit">Enregistrer les noms coutumiers</button></p>
      </form>
    </section>
    <section class="panel">
      <h2>Relations entre personnes</h2>
      <form method="post" action="/bulk/relationships" onsubmit="return confirm('Enregistrer toutes les relations renseignées ?');">
        <div style="overflow:auto;">
          <table>
            <thead><tr><th>Personne A *</th><th>Rôle de A envers B *</th><th>Personne B *</th><th>Certitude</th><th>Confidentialité</th><th>Commentaire</th></tr></thead>
            <tbody>{''.join(relationship_rows)}</tbody>
          </table>
        </div>
        <p><button type="submit">Enregistrer les relations</button></p>
      </form>
    </section>
    """


def graph_style() -> str:
    return """
    <style>
      .graph-wrap { overflow-x: auto; padding: 8px 0; }
      .graph-svg { width: 100%; min-width: 760px; height: auto; border: 1px solid var(--line); background: #fbfcfe; }
      .graph-node rect { fill: #fff; stroke: #22577a; stroke-width: 1.5; rx: 8; }
      .graph-node.center rect { stroke: #111827 !important; stroke-width: 3.2; }
      .graph-node.person rect { fill: #eef6fb; stroke: #22577a; }
      .graph-node.group rect { fill: #f3f0ff; stroke: #6554c0; }
      .graph-node.event rect { fill: #fff4f0; stroke: #b45309; }
      .graph-node.place rect { fill: #fff8e7; stroke: #8a5b00; }
      .graph-node text { font-size: 13px; fill: #202124; }
      .graph-node .title { font-weight: 700; fill: #174766; }
      .graph-node .sub { font-size: 11px; fill: #5f6368; }
      .graph-node .mini-button rect { fill: #fff; stroke: #22577a; stroke-width: 1; rx: 4; }
      .graph-node .mini-button text { font-size: 10px; fill: #22577a; }
      .graph-link { stroke: #7c8794; stroke-width: 1.4; fill: none; }
      .graph-link.a_verifier, .graph-link.probable { stroke-dasharray: 6 5; }
      .graph-link.conteste { stroke: #9a3412; stroke-dasharray: 3 4; }
      .graph-label { font-size: 11px; fill: #5f6368; }
    </style>
    """


def stable_color_seed(value: Any) -> int:
    text = ("" if value is None else str(value)).strip().lower()
    if not text:
        return 0
    return sum((index + 1) * ord(char) for index, char in enumerate(text))


def dynamic_graph_color(value: Any) -> tuple[str, str]:
    seed = stable_color_seed(value)
    hue = seed % 360
    fill = f"hsl({hue}, 85%, 94%)"
    stroke = f"hsl({hue}, 58%, 42%)"
    return fill, stroke


def graph_person_style(person: sqlite3.Row, color_mode: str = "gender", branch_key: Any = None) -> str:
    if color_mode == "birth_name":
        fill, stroke = dynamic_graph_color(row_value(person, "birth_name"))
    elif color_mode == "branch":
        fill, stroke = dynamic_graph_color(branch_key if branch_key is not None else row_value(person, "id"))
    else:
        gender = row_value(person, "gender")
        if is_male(gender):
            fill, stroke = "#e8f4ff", "#1f6fa8"
        elif is_female(gender):
            fill, stroke = "#fff0f3", "#b8325a"
        else:
            fill, stroke = "#f1f5f9", "#64748b"
    return f"fill:{fill};stroke:{stroke};"


def graph_color_mode_from_query(query: dict[str, list[str]]) -> str:
    if (query.get("branch_color") or [""])[0] == "1":
        return "branch"
    if (query.get("name_color") or [""])[0] == "1":
        return "birth_name"
    return "gender"


def graph_color_controls(color_mode: str, include_branch: bool = True) -> str:
    checked_name = " checked" if color_mode == "birth_name" else ""
    checked_branch = " checked" if color_mode == "branch" else ""
    branch_control = (
        f'<p><label><input type="checkbox" name="branch_color" value="1"{checked_branch}> Couleur par branche descendante</label></p>'
        if include_branch
        else ""
    )
    return f"""
      <p><label><input type="checkbox" name="name_color" value="1"{checked_name}> Couleur par nom de naissance</label></p>
      {branch_control}
    """


def shorten(value: Any, limit: int) -> str:
    text = "" if value is None else str(value)
    return text if len(text) <= limit else text[: max(0, limit - 3)] + "..."


def wrap_svg_text(value: Any, limit: int, max_lines: int) -> list[str]:
    words = ("" if value is None else str(value)).split()
    if not words:
        return []
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = word if not current else f"{current} {word}"
        if len(candidate) <= limit:
            current = candidate
            continue
        if current:
            lines.append(current)
        current = word
        if len(lines) >= max_lines:
            return lines[:max_lines]
    if current and len(lines) < max_lines:
        lines.append(current)
    return lines[:max_lines]


def svg_box(
    x: int,
    y: int,
    title: str,
    subtitle: str = "",
    href: str | None = None,
    css_class: str = "",
    width: int = 170,
    height: int = 62,
) -> str:
    content = f"""
    <g class="graph-node {e(css_class)}">
      <rect x="{x}" y="{y}" width="{width}" height="{height}"></rect>
      <text x="{x + 12}" y="{y + 25}">{e(shorten(title, 24))}</text>
      <text class="sub" x="{x + 12}" y="{y + 45}">{e(shorten(subtitle, 28))}</text>
    </g>
    """
    if href:
        return f'<a href="{e(href)}">{content}</a>'
    return content


def person_tree_title(person: sqlite3.Row) -> str:
    return row_value(person, "first_names") or "Personne"


def person_tree_lines(person: sqlite3.Row) -> list[str]:
    birth_name = row_value(person, "birth_name")
    current_name = row_value(person, "current_name")
    birth_date = row_value(person, "birth_date")
    death_date = row_value(person, "death_date")
    lines: list[str] = []
    names = []
    if birth_name:
        names.append(f"N. {birth_name}")
    if current_name and current_name != birth_name:
        names.append(f"Mar. {current_name}")
    if names:
        lines.append(" | ".join(names))
    dates = []
    if birth_date:
        dates.append(f"né(e) {birth_date}")
    if death_date:
        dates.append(f"déc. {death_date}")
    if dates:
        lines.append(" | ".join(dates))
    clan = row_value(person, "current_clan")
    if clan:
        lines.append(clan)
    return lines


def svg_person_box(
    x: int,
    y: int,
    person: sqlite3.Row,
    css_class: str = "person",
    width: int = 190,
    height: int = 86,
    tree_path: str = "tree",
    style: str = "",
    show_tree_nav: bool = False,
    tree_query: str = "",
    show_buttons: bool = True,
) -> str:
    title = person_tree_title(person)
    lines = person_tree_lines(person)[:3]
    person_id = row_value(person, "id")
    title_limit = 18 if show_buttons else 26
    detail_limit = 31 if show_buttons else 36
    text_limit_y = y + height - (34 if show_buttons else 10)
    text_rows: list[tuple[str, str]] = [("title", line) for line in wrap_svg_text(title, title_limit, 2)]
    for line in lines:
        text_rows.extend(("sub", wrapped) for wrapped in wrap_svg_text(line, detail_limit, 2))
    line_svg_parts: list[str] = []
    for index, (text_class, text_value) in enumerate(text_rows):
        text_y = y + 24 + index * 13
        if text_y > text_limit_y:
            break
        line_svg_parts.append(
            f'<text class="{text_class}" x="{x + 12}" y="{text_y}">{e(text_value)}</text>'
        )
    line_svg = "".join(line_svg_parts)
    nav_svg = ""
    tree_href = f"/persons/{e(person_id)}/{e(tree_path)}"
    if show_tree_nav and show_buttons:
        query_suffix = "&" + tree_query if tree_query else ""
        tree_href = f"/persons/{e(person_id)}/extended-tree?mode=descendance{e(query_suffix)}#tree-view"
        nav_svg = f"""
      <a class="mini-button" href="/persons/{e(person_id)}/extended-tree?mode=descendance&upstream=1{e(query_suffix)}#tree-view">
        <rect x="{x + width - 112}" y="{y + height - 29}" width="50" height="22"></rect>
        <text x="{x + width - 87}" y="{y + height - 14}" text-anchor="middle">Amont</text>
      </a>
      <a class="mini-button" href="/persons/{e(person_id)}/extended-tree?mode=descendance{e(query_suffix)}#tree-view">
        <rect x="{x + width - 56}" y="{y + height - 29}" width="44" height="22"></rect>
        <text x="{x + width - 34}" y="{y + height - 14}" text-anchor="middle">Aval</text>
      </a>
        """
    return f"""
    <g class="graph-node {e(css_class)}">
      <rect x="{x}" y="{y}" width="{width}" height="{height}" style="{e(style)}"></rect>
      <a href="{tree_href}">
        {line_svg if show_buttons else ""}
      </a>
      {"" if show_buttons else line_svg}
      {f'''
      <a class="mini-button" href="/persons/{e(person_id)}">
        <rect x="{x + width - 52}" y="{y + 9}" width="40" height="22"></rect>
        <text x="{x + width - 32}" y="{y + 24}" text-anchor="middle">Fiche</text>
      </a>
      ''' if show_buttons else ""}
      {nav_svg}
    </g>
    """


def svg_line(x1: int, y1: int, x2: int, y2: int, label: str = "", certainty: str | None = None) -> str:
    mid_x = (x1 + x2) // 2
    mid_y = (y1 + y2) // 2 - 6
    label_svg = f'<text class="graph-label" x="{mid_x}" y="{mid_y}" text-anchor="middle">{e(shorten(label, 24))}</text>' if label else ""
    return f'<path class="graph-link {e(certainty or "")}" d="M{x1},{y1} L{x2},{y2}"></path>{label_svg}'


def person_row_select(conn: sqlite3.Connection, person_id: str | int) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT p.*, c.name AS current_clan
        FROM persons p
        LEFT JOIN clans c ON c.id = p.current_clan_id
        WHERE p.id = ?
        """,
        (person_id,),
    ).fetchone()


def child_rows_for_parent(conn: sqlite3.Connection, parent_id: str | int, same_birth_name: str | None = None) -> list[sqlite3.Row]:
    rows = conn.execute(
        """
        SELECT DISTINCT p.*, c.name AS current_clan, r.certainty
        FROM relationships r
        JOIN persons p ON p.id = r.object_person_id
        LEFT JOIN clans c ON c.id = p.current_clan_id
        WHERE r.subject_person_id = ? AND r.relation_type IN ('pere','mere','parent_coutumier')
        UNION
        SELECT DISTINCT p.*, c.name AS current_clan, r.certainty
        FROM relationships r
        JOIN persons p ON p.id = r.subject_person_id
        LEFT JOIN clans c ON c.id = p.current_clan_id
        WHERE r.object_person_id = ? AND r.relation_type = 'enfant'
        ORDER BY birth_date, birth_name, first_names
        """,
        (parent_id, parent_id),
    ).fetchall()
    if same_birth_name:
        return [row for row in rows if row_value(row, "birth_name").lower() == same_birth_name.lower()]
    return rows


def parent_rows_for_child(conn: sqlite3.Connection, child_id: str | int) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT DISTINCT p.*, c.name AS current_clan, r.certainty
        FROM relationships r
        JOIN persons p ON p.id = r.subject_person_id
        LEFT JOIN clans c ON c.id = p.current_clan_id
        WHERE r.object_person_id = ? AND r.relation_type IN ('pere','mere','parent_coutumier')
        UNION
        SELECT DISTINCT p.*, c.name AS current_clan, r.certainty
        FROM relationships r
        JOIN persons p ON p.id = r.object_person_id
        LEFT JOIN clans c ON c.id = p.current_clan_id
        WHERE r.subject_person_id = ? AND r.relation_type = 'enfant'
        ORDER BY birth_name, first_names
        """,
        (child_id, child_id),
    ).fetchall()


def spouse_rows_for_person(conn: sqlite3.Connection, person_id: str | int) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT DISTINCT p.*, c.name AS current_clan, r.certainty
        FROM relationships r
        JOIN persons p ON p.id = r.object_person_id
        LEFT JOIN clans c ON c.id = p.current_clan_id
        WHERE r.subject_person_id = ? AND r.relation_type = 'conjoint'
        UNION
        SELECT DISTINCT p.*, c.name AS current_clan, r.certainty
        FROM relationships r
        JOIN persons p ON p.id = r.subject_person_id
        LEFT JOIN clans c ON c.id = p.current_clan_id
        WHERE r.object_person_id = ? AND r.relation_type = 'conjoint'
        ORDER BY birth_name, first_names
        """,
        (person_id, person_id),
    ).fetchall()


def build_descendant_levels(
    conn: sqlite3.Connection,
    root: sqlite3.Row,
    generations: int,
    same_name: bool,
    show_spouses: bool,
    show_upstream: bool = False,
) -> tuple[list[list[sqlite3.Row]], list[tuple[int, int, str | None]], dict[int, list[sqlite3.Row]]]:
    levels: list[list[sqlite3.Row]] = [[root]]
    links: list[tuple[int, int, str | None]] = []
    spouses: dict[int, list[sqlite3.Row]] = {}
    seen = {int(root["id"])}
    filter_name = row_value(root, "birth_name") if same_name else None
    for _ in range(generations):
        current = levels[-1]
        next_level: list[sqlite3.Row] = []
        for parent in current:
            parent_id = int(parent["id"])
            if show_spouses:
                spouses[parent_id] = spouse_rows_for_person(conn, parent_id)
            for child in child_rows_for_parent(conn, parent_id, filter_name):
                child_id = int(child["id"])
                links.append((parent_id, child_id, row_value(child, "certainty")))
                if child_id not in seen:
                    seen.add(child_id)
                    next_level.append(child)
        if not next_level:
            break
        levels.append(next_level)
    if show_upstream:
        parents = parent_rows_for_child(conn, int(root["id"]))
        if parents:
            sibling_by_id: dict[int, sqlite3.Row] = {}
            for parent in parents:
                parent_id = int(parent["id"])
                for sibling in child_rows_for_parent(conn, parent_id):
                    sibling_id = int(sibling["id"])
                    sibling_by_id.setdefault(sibling_id, sibling)
                    links.append((parent_id, sibling_id, row_value(sibling, "certainty")))
            root_id = int(root["id"])
            siblings = [row for sibling_id, row in sibling_by_id.items() if sibling_id != root_id]
            if siblings:
                levels[0] = [root] + siblings
            levels = [parents] + levels
    if show_spouses:
        for level in levels:
            for person in level:
                spouses.setdefault(int(person["id"]), spouse_rows_for_person(conn, int(person["id"])))
    return levels, links, spouses


def build_ancestor_levels(
    conn: sqlite3.Connection,
    root: sqlite3.Row,
    generations: int,
    show_spouses: bool,
) -> tuple[list[list[sqlite3.Row]], list[tuple[int, int, str | None]], dict[int, list[sqlite3.Row]]]:
    levels: list[list[sqlite3.Row]] = [[root]]
    links: list[tuple[int, int, str | None]] = []
    spouses: dict[int, list[sqlite3.Row]] = {}
    seen = {int(root["id"])}
    for _ in range(generations):
        current = levels[-1]
        next_level = []
        for child in current:
            child_id = int(child["id"])
            if show_spouses:
                spouses[child_id] = spouse_rows_for_person(conn, child_id)
            for parent in parent_rows_for_child(conn, child_id):
                parent_id = int(parent["id"])
                links.append((parent_id, child_id, row_value(parent, "certainty")))
                if parent_id not in seen:
                    seen.add(parent_id)
                    next_level.append(parent)
        if not next_level:
            break
        levels.append(next_level)
    if show_spouses:
        for level in levels:
            for person in level:
                spouses.setdefault(int(person["id"]), spouse_rows_for_person(conn, int(person["id"])))
    return levels, links, spouses


def render_extended_tree(
    levels: list[list[sqlite3.Row]],
    links: list[tuple[int, int, str | None]],
    spouses: dict[int, list[sqlite3.Row]],
    root_id: int,
    mode: str = "descendance",
    color_mode: str = "gender",
    tree_query: str = "",
    generation_start: int = 0,
    interactive: bool = True,
) -> str:
    person_width = 190
    person_height = 108 if interactive else 128
    spouse_label_space = 28
    x_gap = 260
    y_gap = person_height + spouse_label_space + 20
    top = 70
    left = 70
    max_rows = max((len(level) for level in levels), default=1)
    width = max(1000, left * 2 + len(levels) * x_gap + 240)
    height = max(420, top * 2 + max_rows * y_gap + 80)
    positions: dict[int, tuple[int, int]] = {}
    nodes: list[str] = []
    edge_svg: list[str] = []
    branch_keys: dict[int, int] = {int(root_id): int(root_id)}
    if color_mode == "branch" and mode == "descendance":
        for parent_id, child_id, _certainty in links:
            if int(parent_id) == int(root_id):
                branch_keys[int(child_id)] = int(child_id)
            elif int(parent_id) in branch_keys:
                branch_keys[int(child_id)] = branch_keys[int(parent_id)]

    for level_index, level in enumerate(levels):
        x = left + level_index * x_gap
        column_height = max(1, len(level)) * y_gap
        y_start = max(top, (height - column_height) // 2)
        generation_number = -level_index if mode == "ascendance" else generation_start + level_index
        nodes.append(f'<text class="graph-label" x="{x + person_width // 2}" y="34" text-anchor="middle">Génération {generation_number}</text>')
        for row_index, person in enumerate(level):
            person_id = int(person["id"])
            y = y_start + row_index * y_gap
            positions[person_id] = (x, y)
            css_class = "center person" if person_id == int(root_id) else "person"
            style = graph_person_style(person, color_mode, branch_keys.get(person_id))
            nodes.append(svg_person_box(x, y, person, css_class, person_width, person_height, "extended-tree", style, interactive, tree_query, interactive))
            if spouses.get(person_id):
                spouse_names = ", ".join(shorten(person_tree_title(spouse), 18) for spouse in spouses[person_id][:3])
                nodes.append(f'<text class="graph-label spouse-label" x="{x + person_width // 2}" y="{y + person_height + 18}" text-anchor="middle">Conjoint(s): {e(spouse_names)}</text>')

    for parent_id, child_id, certainty in links:
        if parent_id not in positions or child_id not in positions:
            continue
        px, py = positions[parent_id]
        cx, cy = positions[child_id]
        edge_svg.append(svg_line(px + person_width, py + person_height // 2, cx, cy + person_height // 2, "", certainty))

    if len(levels) == 1:
        nodes.append(f'<text x="{width // 2}" y="{height // 2}" text-anchor="middle" class="graph-label">Aucune génération liée trouvée avec ces filtres.</text>')

    return f"""
    {graph_style()}
    <div class="graph-wrap">
      <svg class="graph-svg" viewBox="0 0 {width} {height}" role="img" aria-label="Arbre étendu">
        {''.join(edge_svg)}
        {''.join(nodes)}
      </svg>
    </div>
    """


def render_person_tree(
    person: sqlite3.Row,
    parents: list[sqlite3.Row],
    spouses: list[sqlite3.Row],
    children: list[sqlite3.Row],
    group_links: list[sqlite3.Row] | None = None,
    structural_events: list[sqlite3.Row] | None = None,
    color_mode: str = "gender",
) -> str:
    group_links = group_links or []
    structural_events = structural_events or []
    person_width = 190
    person_height = 86
    width = max(1040, max(len(children), len(group_links), len(structural_events), len(parents), 1) * 210 + 140)
    has_lower_layers = bool(group_links or structural_events)
    height = 650 if has_lower_layers else 450
    center_y = 175
    links: list[str] = []
    nodes: list[str] = []
    center_x = (width - person_width) // 2
    center_mid_x = center_x + person_width // 2
    center_bottom_y = center_y + person_height
    root_id = int(person["id"])
    nodes.append(svg_person_box(center_x, center_y, person, "center person", person_width, person_height, "tree", graph_person_style(person, color_mode, root_id)))

    parent_count = max(1, len(parents))
    start_x = (width - parent_count * 210) // 2
    for index, parent in enumerate(parents):
        x = start_x + index * 210
        y = 45
        nodes.append(svg_person_box(x, y, parent, "person", person_width, person_height, "tree", graph_person_style(parent, color_mode, root_id)))
        links.append(svg_line(x + person_width // 2, y + person_height, center_mid_x, center_y, relation_label(parent["relation_type"]), parent["certainty"]))

    for index, spouse in enumerate(spouses[:4]):
        side_offset = 300 + (index // 2) * 210
        x = center_x - side_offset if index % 2 == 0 else center_x + side_offset
        x = max(20, min(width - person_width - 20, x))
        y = center_y
        nodes.append(svg_person_box(x, y, spouse, "person", person_width, person_height, "tree", graph_person_style(spouse, color_mode, int(spouse["id"]))))
        links.append(svg_line(center_x if x < center_x else center_x + person_width, center_y + person_height // 2, x + person_width if x < center_x else x, y + person_height // 2, "conjoint", spouse["certainty"]))

    child_count = max(1, len(children))
    start_x = (width - child_count * 210) // 2
    for index, child in enumerate(children):
        x = start_x + index * 210
        y = 310
        nodes.append(svg_person_box(x, y, child, "person", person_width, person_height, "tree", graph_person_style(child, color_mode, int(child["id"]))))
        links.append(svg_line(center_mid_x, center_bottom_y, x + person_width // 2, y, "enfant", child["certainty"]))

    if structural_events:
        event_count = len(structural_events)
        start_x = (width - event_count * 180) // 2
        for index, event in enumerate(structural_events):
            x = start_x + index * 180
            y = 420
            subtitle = event_person_role_label(event["role"]) if event["role"] != "principal" else event_label(event["event_type"])
            if event["event_date"]:
                subtitle = f'{subtitle} - {event["event_date"]}'
            nodes.append(svg_box(x, y, event["title"], subtitle, f'/events/{event["id"]}', "event"))
            links.append(svg_line(center_mid_x, center_bottom_y, x + 85, y, "événement", event["certainty"]))

    if group_links:
        group_count = len(group_links)
        start_x = (width - group_count * 180) // 2
        for index, group in enumerate(group_links):
            x = start_x + index * 180
            y = 520 if structural_events else 420
            nodes.append(svg_box(x, y, group["group_name"], person_group_link_label(group["link_type"]), f'/groups/{group["group_id"]}', "group"))
            links.append(svg_line(center_mid_x, center_bottom_y, x + 85, y, "groupe", group["certainty"]))

    if not parents and not spouses and not children and not group_links and not structural_events:
        nodes.append(f'<text x="{width // 2}" y="285" text-anchor="middle" class="graph-label">Aucune relation familiale ou coutumière saisie pour le moment.</text>')

    return f"""
    {graph_style()}
    <div class="graph-wrap">
      <svg class="graph-svg" viewBox="0 0 {width} {height}" role="img" aria-label="Arbre genealogique">
        {''.join(links)}
        {''.join(nodes)}
      </svg>
    </div>
    """


def render_group_map(
    group: sqlite3.Row,
    relations: list[sqlite3.Row],
    viva_links: list[sqlite3.Row],
    functions: list[sqlite3.Row],
    linked_persons: list[sqlite3.Row] | None = None,
) -> str:
    width = 1180
    height = 620
    center_x = (width - 170) // 2
    center_y = 220
    nodes: list[str] = [svg_box(center_x, center_y, group["name"], group_type_label(group["group_type"]), f'/groups/{group["id"]}', "center group")]
    links: list[str] = []
    seen: set[str] = {group["name"]}
    related: list[tuple[str, str, str, str | None]] = []

    for relation in relations:
        other_name = relation["group_b"] if relation["group_a_id"] == group["id"] else relation["group_a"]
        if other_name and other_name not in seen:
            related.append((other_name, group_relation_label(relation["relation_type"]), "group", relation["certainty"]))
            seen.add(other_name)
    for viva in viva_links:
        for name in (viva["group_a"], viva["group_b"]):
            if name and name != group["name"] and name not in seen:
                related.append((name, "apparie viva", "group", viva["certainty"]))
                seen.add(name)
    for function in functions:
        label = function["place"] or function["title"]
        if label and label not in seen:
            related.append((label, function_type_label(function["function_type"]), "place", function["certainty"]))
            seen.add(label)
    for person in linked_persons or []:
        label = person["person_name"]
        if label and label not in seen:
            related.append((label, person_group_link_label(person["link_type"]), "person", person["certainty"]))
            seen.add(label)

    positions = [
        (80, 70), (300, 45), (520, 45), (740, 45), (960, 70),
        (80, 210), (930, 210),
        (80, 385), (300, 455), (520, 475), (740, 455), (960, 385),
    ]
    for index, (name, label, kind, certainty) in enumerate(related[:12]):
        x, y = positions[index]
        nodes.append(svg_box(x, y, name, label, None, kind))
        links.append(svg_line(center_x + 85, center_y + 31, x + 85, y + 31, label, certainty))

    if len(related) > 12:
        nodes.append(f'<text x="{width // 2}" y="575" text-anchor="middle" class="graph-label">{len(related) - 12} liens supplémentaires non affichés dans cette première carte.</text>')

    if not related:
        nodes.append(f'<text x="{width // 2}" y="340" text-anchor="middle" class="graph-label">Aucune relation de groupe ou fonction saisie pour le moment.</text>')

    return f"""
    {graph_style()}
    <div class="graph-wrap">
      <svg class="graph-svg" viewBox="0 0 {width} {height}" role="img" aria-label="Carte relationnelle de groupe">
        {''.join(links)}
        {''.join(nodes)}
      </svg>
    </div>
    """


def none_if_empty(value: str | None) -> str | None:
    return value if value else None


def row_value(row: sqlite3.Row | None, key: str, default: str = "") -> str:
    if row is None:
        return default
    value = row[key]
    return "" if value is None else str(value)


def text_field(name: str, label: str, value: str = "", required: bool = False, placeholder: str = "") -> str:
    return f'<p><label>{e(label)}</label><input name="{e(name)}" value="{e(value)}"{" required" if required else ""} placeholder="{e(placeholder)}"></p>'


def textarea_field(name: str, label: str, value: str = "", required: bool = False, placeholder: str = "") -> str:
    return f'<p class="full"><label>{e(label)}</label><textarea name="{e(name)}"{" required" if required else ""} placeholder="{e(placeholder)}">{e(value)}</textarea></p>'


def select_field(name: str, label: str, options: list[tuple[str, str]] | list[sqlite3.Row], selected: Any = None, required: bool = False) -> str:
    if options and isinstance(options[0], sqlite3.Row):
        options_html = option_list(options, selected)
    else:
        options_html = select_options(options, "" if selected is None else str(selected))  # type: ignore[arg-type]
    return f'<p><label>{e(label)}</label><select name="{e(name)}"{" required" if required else ""}>{options_html}</select></p>'


def person_relation_select(name: str, label: str, persons: list[sqlite3.Row], selected: Any = None) -> str:
    selected_text = "" if selected is None else str(selected)
    options = ['<option value="" data-customary-names="">Choisir une personne</option>']
    for person in persons:
        value = str(person["id"])
        mark = " selected" if value == selected_text else ""
        options.append(
            f'<option value="{e(value)}" data-customary-names="{e(person["customary_names"])}"{mark}>{e(person["label"])}</option>'
        )
    return f'<p><label>{e(label)}</label><select id="object_person_id" name="{e(name)}" required>{"".join(options)}</select></p>'


def relation_type_select(name: str, label: str, selected: str = "pere") -> str:
    return f'<p><label>{e(label)}</label><select id="relation_type" name="{e(name)}" required>{select_options(RELATION_TYPES, selected)}</select></p>'


def relationship_preview(subject_label: str, selected_relation: str = "pere") -> str:
    return f"""
    <p class="full">
      <label>Phrase qui sera enregistrée</label>
      <input id="relationship_preview" value="" readonly>
    </p>
    <script>
      (function() {{
        var subject = {js_literal(subject_label)};
        var relation = document.getElementById('relation_type');
        var person = document.getElementById('object_person_id');
        var preview = document.getElementById('relationship_preview');
        if (!relation || !person || !preview) return;
        function updateRelationshipPreview() {{
          var relationText = relation.options[relation.selectedIndex] ? relation.options[relation.selectedIndex].text.toLowerCase() : '';
          var personText = person.options[person.selectedIndex] ? person.options[person.selectedIndex].text : '';
          preview.value = person.value ? subject + ' est ' + relationText + ' de ' + personText : subject + ' est ... de ...';
        }}
        relation.addEventListener('change', updateRelationshipPreview);
        person.addEventListener('change', updateRelationshipPreview);
        updateRelationshipPreview();
      }})();
    </script>
    """


def customary_name_display_field(persons: list[sqlite3.Row], selected: Any = None) -> str:
    selected_text = "" if selected is None else str(selected)
    current = ""
    for person in persons:
        if str(person["id"]) == selected_text:
            current = row_value(person, "customary_names", "")
            break
    return f"""
    <p>
      <label>Nom coutumier associé</label>
      <input id="linked_customary_names" value="{e(current or 'Non renseigné')}" readonly>
    </p>
    <script>
      (function() {{
        var select = document.getElementById('object_person_id');
        var target = document.getElementById('linked_customary_names');
        if (!select || !target) return;
        function updateCustomaryNames() {{
          var option = select.options[select.selectedIndex];
          var value = option ? option.getAttribute('data-customary-names') : '';
          target.value = value || 'Non renseigné';
        }}
        select.addEventListener('change', updateCustomaryNames);
        updateCustomaryNames();
      }})();
    </script>
    """


def edit_form_panel(title: str, fields: list[str], cancel_path: str) -> str:
    return f"""
    <section class="panel">
      <h2>{e(title)}</h2>
      <form method="post" class="form-grid">
        {''.join(fields)}
        <p class="full"><button type="submit">Enregistrer les modifications</button> <a class="button secondary" href="{e(cancel_path)}">Annuler</a></p>
      </form>
    </section>
    """


def event_label(value: str) -> str:
    return dict(EVENT_TYPES).get(value, value)


def event_person_role_label(value: str) -> str:
    return dict(EVENT_PERSON_ROLES).get(value, value)


def source_label(value: str) -> str:
    return dict(SOURCE_TYPES).get(value, value)


def relation_label(value: str) -> str:
    return dict(RELATION_TYPES).get(value, value)


def relation_sentence(subject_name: str | None, relation_type: str, object_name: str | None) -> str:
    subject = subject_name or "Personne non renseignée"
    obj = object_name or "Personne non renseignée"
    return f"{subject} est {relation_label(relation_type).lower()} de {obj}"


def is_female(value: str | None) -> bool:
    return (value or "").strip().lower() in {"f", "femme", "féminin", "feminin", "female"}


def is_male(value: str | None) -> bool:
    return (value or "").strip().lower() in {"m", "homme", "masculin", "male"}


def gendered_label(male_label: str, female_label: str, unknown_label: str, gender: str | None) -> str:
    if is_female(gender):
        return female_label
    if is_male(gender):
        return male_label
    return unknown_label


def relation_label_for_view(viewer_is_subject: bool, relation_type: str, viewer_gender: str | None) -> str:
    if relation_type == "pere":
        return "père" if viewer_is_subject else "enfant"
    if relation_type == "mere":
        return "mère" if viewer_is_subject else "enfant"
    if relation_type == "enfant":
        return "enfant" if viewer_is_subject else gendered_label("père", "mère", "parent", viewer_gender)
    if relation_type == "conjoint":
        return gendered_label("conjoint", "conjointe", "conjoint/conjointe", viewer_gender)
    if relation_type == "frere_soeur":
        return gendered_label("frère", "sœur", "frère/sœur", viewer_gender)
    if relation_type == "oncle_maternel":
        return "oncle maternel" if viewer_is_subject else gendered_label("neveu", "nièce", "neveu/nièce", viewer_gender)
    if relation_type == "tante_maternelle":
        return "tante maternelle" if viewer_is_subject else gendered_label("neveu", "nièce", "neveu/nièce", viewer_gender)
    if relation_type == "parent_coutumier":
        return "parent coutumier" if viewer_is_subject else "enfant coutumier"
    if relation_type == "referent_memoire":
        return "référent mémoire" if viewer_is_subject else "lié au référent mémoire"
    if relation_type == "appartenance_clan":
        return "lié au clan" if viewer_is_subject else "a pour membre"
    if relation_type == "clan_origine":
        return "originaire du clan" if viewer_is_subject else "clan d'origine de"
    return relation_label(relation_type).lower()


def relation_sentence_for_view(viewer_id: int, viewer_gender: str | None, row: sqlite3.Row) -> str:
    viewer_is_subject = int(row["subject_person_id"]) == int(viewer_id)
    subject = row["subject_person"] if viewer_is_subject else row["object_person"]
    obj = row["object_person"] if viewer_is_subject else row["subject_person"]
    relation_text = relation_label_for_view(viewer_is_subject, row["relation_type"], viewer_gender)
    return f"{subject or 'Personne non renseignée'} est {relation_text} de {obj or 'Personne non renseignée'}"


def person_display(person: sqlite3.Row | None) -> str:
    if not person:
        return "Cette personne"
    birth_name = row_value(person, "birth_name", "Inconnu") or "Inconnu"
    current_name = row_value(person, "current_name")
    suffix = f" ({current_name})" if current_name else ""
    return f"{row_value(person, 'first_names')} - {birth_name}{suffix}"


def group_type_label(value: str) -> str:
    return dict(GROUP_TYPES).get(value, value)


def group_relation_label(value: str) -> str:
    return dict(GROUP_RELATION_TYPES).get(value, value)


def person_group_link_label(value: str) -> str:
    return dict(PERSON_GROUP_LINK_TYPES).get(value, value)


def subject_type_label(value: str) -> str:
    return dict(SUBJECT_TYPES).get(value, value)


def evidence_label(value: str) -> str:
    return dict(EVIDENCE_LEVELS).get(value, value)


def function_type_label(value: str) -> str:
    return dict(FUNCTION_TYPES).get(value, value)


def sex_label(value: str | None) -> str:
    return dict(SEXES).get(value or "", value or "")


def person_form() -> str:
    return f"""
    <section class="panel">
      <h2>Nouvelle personne</h2>
      <form method="post" class="form-grid">
        <p><label>Prénom principal *</label><input name="first_names" required></p>
        <p><label>Nom de naissance *</label><input name="birth_name" required placeholder="ex: Inconnu"></p>
        <p><label>Nom marital / nom actuel</label><input name="current_name"></p>
        <p><label>Autres noms</label><input name="other_names"></p>
        <p><label>Sexe</label><select name="gender">{select_options(SEXES, "")}</select></p>
        <p><label>Date ou période de naissance</label><input name="birth_date" placeholder="ex: vers 1975"></p>
        <p><label>Lieu de naissance</label><input name="birth_place"></p>
        <p><label>Date ou période de décès</label><input name="death_date"></p>
        <p><label>Confidentialité de la fiche *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude de la fiche *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Notes</label><textarea name="notes"></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/persons">Annuler</a></p>
      </form>
    </section>
    """


def group_form(sources: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Nouveau groupe coutumier</h2>
      <p class="empty">Utilise ce formulaire quand tu n'es pas sûr qu'un nom soit un clan. On peut le classer provisoirement comme lignage, nom de famille, branche ou nom cité dans un viva.</p>
      <form method="post" class="form-grid">
        <p><label>Nom *</label><input name="name" required></p>
        <p><label>Type *</label><select name="group_type">{select_options(GROUP_TYPES, "a_verifier")}</select></p>
        <p><label>Variantes</label><input name="variants" placeholder="Autres écritures ou prononciations"></p>
        <p><label>Région / aire</label><input name="region"></p>
        <p><label>Source</label><select name="source_id">{option_list(sources)}</select></p>
        <p><label>Confidentialité *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Description</label><textarea name="description"></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/groups">Annuler</a></p>
      </form>
    </section>
    """


def customary_name_form(sources: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Nouveau nom coutumier</h2>
      <p class="empty">Utilise ce formulaire pour les noms chantés, cités dans les viva ou transmis oralement, sans décider trop vite s'il s'agit d'un clan, d'un lignage ou d'un nom de famille.</p>
      <form method="post" class="form-grid">
        <p><label>Nom coutumier *</label><input name="name" required></p>
        <p><label>Type *</label><select name="name_type">{select_options(GROUP_TYPES, "a_verifier")}</select></p>
        <p><label>Variantes</label><input name="variants" placeholder="Autres écritures ou prononciations"></p>
        <p><label>Région / aire</label><input name="region"></p>
        <p><label>Source</label><select name="source_id">{option_list(sources)}</select></p>
        <p><label>Confidentialité *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Description</label><textarea name="description"></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/customary-names">Annuler</a></p>
      </form>
    </section>
    """


def customary_name_relation_fields(
    custom_name: sqlite3.Row,
    other_names: list[sqlite3.Row],
    sources: list[sqlite3.Row],
    data: dict[str, str] | None = None,
) -> list[str]:
    data = data or {}
    return [
        select_field("name_b_id", "Nom coutumier allié / associé *", other_names, data.get("name_b_id"), True),
        select_field("relation_type", "Type de relation *", GROUP_RELATION_TYPES, data.get("relation_type", "allie_a"), True),
        select_field("source_id", "Source", sources, data.get("source_id")),
        select_field("confidentiality", "Confidentialité *", CONFIDENTIALITY, data.get("confidentiality", "familial"), True),
        select_field("certainty", "Certitude *", CERTAINTY, data.get("certainty", "a_verifier"), True),
        textarea_field("context", "Commentaire / contexte", data.get("context", f"Relation associée à {custom_name['name']}")),
    ]


def viva_form(sources: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Nouvelle liste viva</h2>
      <form method="post" class="form-grid">
        <p><label>Titre *</label><input name="title" required></p>
        <p><label>Aire / région</label><input name="area"></p>
        <p><label>Auteur ou collecteur</label><input name="collector_author"></p>
        <p><label>Date ou période</label><input name="collection_date"></p>
        <p><label>Source</label><select name="source_id">{option_list(sources)}</select></p>
        <p><label>Confidentialité *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Description</label><textarea name="description" placeholder="Contexte, personne qui récite, document, variantes connues..."></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/viva">Annuler</a></p>
      </form>
    </section>
    """


def research_form(groups: list[sqlite3.Row], sources: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Nouvelle information / hypothese</h2>
      <form method="post" class="form-grid">
        <p><label>Type de sujet *</label><select name="subject_type">{select_options(SUBJECT_TYPES, "group")}</select></p>
        <p><label>Groupe concerne</label><select name="subject_id">{option_list(groups)}</select></p>
        <p><label>Titre *</label><input name="title" required></p>
        <p><label>Auteur / version</label><input name="author_view" placeholder="ex: Pillon, Guiart, Frimigacci, ancien..."></p>
        <p><label>Source</label><select name="source_id">{option_list(sources)}</select></p>
        <p><label>Niveau de preuve *</label><select name="evidence_level">{select_options(EVIDENCE_LEVELS, "document_ecrit")}</select></p>
        <p><label>Confidentialite *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Information relevee *</label><textarea name="statement" required placeholder="Ce que dit la source, sans trop interpreter"></textarea></p>
        <p class="full"><label>Interpretation / prudence</label><textarea name="interpretation" placeholder="Ce que nous en comprenons, ce qui reste a verifier, contradictions possibles..."></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/research">Annuler</a></p>
      </form>
    </section>
    """


def lineage_form(groups: list[sqlite3.Row], sources: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Nouvelle genealogie de lignage</h2>
      <form method="post" class="form-grid">
        <p><label>Groupe / lignage</label><select name="group_id">{option_list(groups)}</select></p>
        <p><label>Titre *</label><input name="title" required></p>
        <p><label>Auteur / collecteur</label><input name="author_or_collector" placeholder="ex: Frimigacci cite par Pillon"></p>
        <p><label>Source</label><select name="source_id">{option_list(sources)}</select></p>
        <p><label>Confidentialite *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Chaine de noms *</label><textarea name="chain_text" required placeholder="ex: Dinage, gowemeu, Fane..."></textarea></p>
        <p class="full"><label>Interpretation</label><textarea name="interpretation" placeholder="Genealogie de lignage, liste viva, version contestee, points a verifier..."></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/lineages">Annuler</a></p>
      </form>
    </section>
    """


def function_form(groups: list[sqlite3.Row], persons: list[sqlite3.Row], sources: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Nouvelle fonction coutumiere</h2>
      <form method="post" class="form-grid">
        <p><label>Groupe concerne</label><select name="group_id">{option_list(groups)}</select></p>
        <p><label>Personne concernee</label><select name="person_id">{option_list(persons)}</select></p>
        <p><label>Type de fonction *</label><select name="function_type">{select_options(FUNCTION_TYPES, "autre")}</select></p>
        <p><label>Titre *</label><input name="title" required></p>
        <p><label>Lieu</label><input name="place"></p>
        <p><label>Source</label><select name="source_id">{option_list(sources)}</select></p>
        <p><label>Confidentialite *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Description *</label><textarea name="description" required></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/functions">Annuler</a></p>
      </form>
    </section>
    """


def relationship_form(
    person: sqlite3.Row,
    persons: list[sqlite3.Row],
    error: str = "",
) -> str:
    return f"""
    <section class="panel">
      <h2>Nouvelle relation pour {e(person["first_names"])} {e(person["current_name"])}</h2>
      {f'<p class="flash">{e(error)}</p>' if error else ''}
      <form method="post" class="form-grid">
        {relation_type_select("relation_type", f"Rôle de {person_display(person)} envers la personne liée *", "pere")}
        {person_relation_select("object_person_id", "Personne liée *", persons)}
        {customary_name_display_field(persons)}
        {relationship_preview(person_display(person), "pere")}
        <p><label>Confidentialité *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Commentaire</label><textarea name="comment" placeholder="Source, nuance, contexte ou remarque à vérifier"></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/persons/{person["id"]}">Annuler</a></p>
      </form>
    </section>
    """


def clan_form() -> str:
    return f"""
    <section class="panel">
      <h2>Nouveau clan</h2>
      <form method="post" class="form-grid">
        <p><label>Nom du clan *</label><input name="name" required></p>
        <p><label>Variantes</label><input name="variants"></p>
        <p><label>Aire / région</label><input name="region"></p>
        <p><label>Confidentialité *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Description</label><textarea name="description"></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/clans">Annuler</a></p>
      </form>
    </section>
    """


def event_form(persons: list[sqlite3.Row], sources: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Nouvel événement coutumier</h2>
      <form method="post" class="form-grid">
        <p><label>Type *</label><select name="event_type" required>{select_options(EVENT_TYPES, "naissance")}</select></p>
        <p><label>Titre *</label><input name="title" required></p>
        <p><label>Date ou période</label><input name="event_date"></p>
        <p><label>Lieu</label><input name="place"></p>
        <p><label>Personne principale</label><select name="main_person_id">{option_list(persons)}</select></p>
        <p><label>Source</label><select name="source_id">{option_list(sources)}</select></p>
        <p><label>Confidentialité *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p><label>Certitude *</label><select name="certainty">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p class="full"><label>Description *</label><textarea name="description" required></textarea></p>
        <p class="full"><label>Effets</label><textarea name="effects" placeholder="Ex: changement de nom, clan, droits, obligations"></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/events">Annuler</a></p>
      </form>
    </section>
    """


def source_form() -> str:
    return f"""
    <section class="panel">
      <h2>Nouvelle source</h2>
      <form method="post" class="form-grid">
        <p><label>Type *</label><select name="source_type" required>{select_options(SOURCE_TYPES, "temoignage_oral")}</select></p>
        <p><label>Titre *</label><input name="title" required></p>
        <p><label>Témoin / transmetteur</label><input name="witness_name"></p>
        <p><label>Date de recueil</label><input name="collection_date" placeholder="ex: 2026-07-15"></p>
        <p><label>Recueilli par</label><input name="collected_by"></p>
        <p><label>Consentement *</label><select name="consent">{select_options(CONSENT, "a_confirmer")}</select></p>
        <p><label>Fiabilité *</label><select name="reliability">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p><label>Confidentialité *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "familial")}</select></p>
        <p class="full"><label>Résumé *</label><textarea name="summary" required></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/sources">Annuler</a></p>
      </form>
    </section>
    """


def land_form(clans: list[sqlite3.Row], sources: list[sqlite3.Row]) -> str:
    return f"""
    <section class="panel">
      <h2>Nouvelle terre coutumière</h2>
      <p class="empty">Par défaut, une terre est marquée sensible.</p>
      <form method="post" class="form-grid">
        <p><label>Nom du lieu ou de la terre *</label><input name="name" required></p>
        <p><label>Type de lieu</label><input name="place_type" placeholder="terre, village, tribu, île..."></p>
        <p><label>Clan associé</label><select name="clan_id">{option_list(clans)}</select></p>
        <p><label>Source</label><select name="source_id">{option_list(sources)}</select></p>
        <p><label>Statut *</label><select name="status">{select_options(CERTAINTY, "a_verifier")}</select></p>
        <p><label>Confidentialité *</label><select name="confidentiality">{select_options(CONFIDENTIALITY, "sensible")}</select></p>
        <p class="full"><label>Localisation textuelle</label><textarea name="location_text"></textarea></p>
        <p class="full"><label>Droits connus</label><textarea name="known_rights"></textarea></p>
        <p class="full"><button type="submit">Enregistrer</button> <a class="button secondary" href="/lands">Annuler</a></p>
      </form>
    </section>
    """


def main() -> None:
    init_db()
    server = ThreadingHTTPServer((HOST, PORT), AppHandler)
    print(f"{APP_NAME} lancé sur http://{HOST}:{PORT}")
    print("Appuie sur Ctrl+C pour arrêter le serveur.")
    server.serve_forever()


if __name__ == "__main__":
    main()
